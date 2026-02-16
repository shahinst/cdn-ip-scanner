"""
CDN IP Scanner V2.0 - API Routes
Author: shahinst
"""

import json
import time
import threading
import traceback
from datetime import datetime
from flask import Blueprint, request, jsonify, current_app
from app import db, socketio
from app.models import (
    ScanResult, ScanSession, ClosedIP, OperatorRange,
    OperatorMatrix, AppSetting, ScanLog
)
from app.scanner.core import SHScanner, SHNetUtils, SPEED_MODES
from app.scanner.range_fetcher import RangeFetcher
from app.scanner.operators import (
    OPERATORS_BY_COUNTRY, fetch_all_operator_prefixes
)
from app.scanner.v2ray import V2RayConfigParser, V2RayScanner

api_bp = Blueprint('api', __name__)

# Global scanner instances
_scanner = SHScanner()
_v2ray_scanner = V2RayScanner()
_active_session_id = None
_user_stop_requested = False  # set by stop_scan(), checked by run_scan() to exit batch loop
_log_enabled = False
_debug_enabled = False


def _emit_log(level, message, session_id=None):
    """Emit log to WebSocket and optionally save to DB."""
    global _log_enabled, _debug_enabled
    # Skip DEBUG-level logs unless debug is enabled
    if level == 'DEBUG' and not _debug_enabled:
        return
    timestamp = datetime.utcnow().isoformat()
    log_entry = {'level': level, 'message': message, 'timestamp': timestamp}
    try:
        socketio.emit('scan_log', log_entry, namespace='/')
    except Exception:
        pass
    if _log_enabled and session_id:
        try:
            log = ScanLog(session_id=session_id, level=level, message=message)
            db.session.add(log)
            db.session.commit()
        except Exception:
            try:
                db.session.rollback()
            except Exception:
                pass


# ========== Settings ==========

@api_bp.route('/settings', methods=['GET'])
def get_settings():
    defaults = {
        'theme': 'light', 'mode': 'hyper', 'ai_level': 'smart',
        'target_count': '100', 'ping_min': '0', 'ping_max': '9999',
        'scan_ports': '443,80,8443,2053,2083,2087,2096',
        'language': 'en', 'log_enabled': 'false',
        'debug_enabled': 'false',
        'operator_country': 'ir',
    }
    settings = {}
    for key, default in defaults.items():
        settings[key] = AppSetting.get(key, default)
    return jsonify(settings)


@api_bp.route('/settings', methods=['POST'])
def save_settings():
    data = request.json or {}
    for key, value in data.items():
        AppSetting.set(key, str(value))
    global _log_enabled
    _log_enabled = str(data.get('log_enabled', 'true')).lower() == 'true'
    return jsonify({'status': 'ok'})


# ========== Ranges ==========

@api_bp.route('/ranges/fetch', methods=['POST'])
def fetch_ranges():
    source = (request.json or {}).get('source', 'all')
    try:
        ranges = RangeFetcher.fetch_by_source(source)
        return jsonify({'ranges': ranges, 'count': len(ranges), 'source': source})
    except Exception as e:
        return jsonify({
            'ranges': [], 'count': 0, 'source': source,
            'error': str(e) or 'Server cannot reach the internet. Check firewall and DNS.'
        }), 200


@api_bp.route('/ranges/operators', methods=['GET'])
def get_operators():
    country = request.args.get('country', 'ir')
    operators = OPERATORS_BY_COUNTRY.get(country, OPERATORS_BY_COUNTRY['ir'])
    result = {}
    for key, info in operators.items():
        count = OperatorRange.query.filter_by(operator_key=key).count()
        result[key] = {
            'name': info['name'],
            'name_fa': info.get('name_fa', info['name']),
            'asn': info['asn'],
            'prefix_count': count,
        }
    return jsonify(result)


@api_bp.route('/ranges/operators/fetch', methods=['POST'])
def fetch_operator_ranges():
    data = request.json or {}
    operator_key = data.get('operator_key')
    country = data.get('country', 'ir')

    if not operator_key:
        return jsonify({'error': 'operator_key required'}), 400

    cnt, prefixes, err = fetch_all_operator_prefixes(operator_key, country)
    if err:
        return jsonify({'error': err, 'count': 0}), 500

    operators = OPERATORS_BY_COUNTRY.get(country, {})
    op_info = operators.get(operator_key, {})

    OperatorRange.query.filter_by(operator_key=operator_key).delete()
    for p in prefixes:
        db.session.add(OperatorRange(
            operator_key=operator_key,
            operator_name=op_info.get('name', operator_key),
            asn=str(op_info.get('asn', '')),
            prefix=p,
            country=country,
        ))
    db.session.commit()
    return jsonify({'count': cnt, 'operator': operator_key})


@api_bp.route('/ranges/operators/fetch-all', methods=['POST'])
def fetch_all_operators_route():
    """Fetch IP ranges for all operators in the selected country from RIPE/BGP (no local IP)."""
    country = (request.json or {}).get('country', 'ir')
    operators = OPERATORS_BY_COUNTRY.get(country, OPERATORS_BY_COUNTRY['ir'])
    results = {}
    for op_key in operators:
        cnt, prefixes, err = fetch_all_operator_prefixes(op_key, country)
        if not err and prefixes:
            OperatorRange.query.filter_by(operator_key=op_key).delete()
            op_info = operators[op_key]
            for p in prefixes:
                db.session.add(OperatorRange(
                    operator_key=op_key,
                    operator_name=op_info.get('name', op_key),
                    asn=str(op_info.get('asn', '')),
                    prefix=p,
                    country=country,
                ))
            db.session.commit()
        results[op_key] = {'count': cnt, 'error': err}
    return jsonify(results)


# ========== V2Ray Config ==========

@api_bp.route('/v2ray/parse', methods=['POST'])
def parse_v2ray_config():
    data = request.json or {}
    config_str = data.get('config', '')
    parsed = V2RayConfigParser.parse(config_str)
    if not parsed:
        return jsonify({'error': 'Invalid config format'}), 400
    return jsonify({
        'protocol': parsed['protocol'],
        'uuid': parsed['uuid'][:8] + '...',
        'ip': parsed['ip'],
        'port': parsed['port'],
        'params': {k: v for k, v in parsed['params'].items() if k != 'id'},
        'fragment': parsed.get('fragment', ''),
    })


@api_bp.route('/v2ray/build-config', methods=['POST'])
def build_v2ray_config():
    """Build config string with a new IP for download."""
    data = request.json or {}
    config_str = data.get('config', '')
    new_ip = data.get('ip', '')
    if not config_str or not new_ip:
        return jsonify({'error': 'config and ip required'}), 400
    parsed = V2RayConfigParser.parse(config_str)
    if not parsed:
        return jsonify({'error': 'Invalid config format'}), 400
    built = V2RayConfigParser.rebuild_config(parsed, new_ip)
    if not built:
        return jsonify({'error': 'Could not build config'}), 400
    return jsonify({'config': built})


# ========== Scanning ==========

@api_bp.route('/scan/start', methods=['POST'])
def start_scan():
    global _active_session_id, _log_enabled, _debug_enabled

    data = request.json or {}
    ranges_input = data.get('ranges', [])
    scan_method = data.get('scan_method', 'cloud')
    mode = data.get('mode', 'hyper')
    target_count = data.get('target_count', 100)
    ping_min = int(data.get('ping_min', 0))
    ping_max = int(data.get('ping_max', 9999))
    ports_str = data.get('ports', '443,80,8443,2053,2083,2087,2096')
    operator_key = (data.get('operator_key') or '').strip() or None
    if operator_key:
        operator_key = operator_key.lower()
    country = (data.get('country') or 'ir').strip().lower() or 'ir'
    clear_previous = data.get('clear_previous', True)
    v2ray_config = data.get('v2ray_config', '')
    _log_enabled = str(data.get('log_enabled', True)).lower() in ('true', '1', 'yes')
    _debug_enabled = str(data.get('debug_enabled', False)).lower() in ('true', '1', 'yes')

    ports = [int(p.strip()) for p in ports_str.split(',') if p.strip().isdigit()]
    if not ports:
        ports = [443, 80, 8443, 2053, 2083, 2087, 2096]

    if target_count == 'All':
        target_count = None
    else:
        try:
            target_count = int(target_count)
        except (ValueError, TypeError):
            target_count = 100

    # Create session
    session = ScanSession(mode=mode, scan_method=scan_method, status='running')
    db.session.add(session)
    db.session.commit()
    sess_id = session.id
    _active_session_id = sess_id

    if clear_previous:
        ClosedIP.query.delete()
        db.session.commit()

    # CRITICAL: Get the CURRENT app reference - do NOT create_app() in thread
    app = current_app._get_current_object()

    def _scanner_log(level, message):
        with app.app_context():
            _emit_log(level, message, sess_id)

    _scanner.log_callback = _scanner_log
    _scanner.reset()

    def run_scan():
        """Background scan thread - uses the EXISTING app context."""
        with app.app_context():
            try:
                start_time = time.time()
                sess = db.session.get(ScanSession, sess_id)
                if not sess:
                    return

                # Configure scanner
                _scanner.set_mode(mode)
                _scanner.max_latency_ms = ping_max
                _scanner.timeout = min(10, max(2, ping_max / 1000.0 * 1.5))

                _emit_log('INFO', f'Scan started: method={scan_method}, mode={mode}', sess_id)

                # Handle V2Ray scan method
                v2ray_parsed = None
                if scan_method == 'v2ray' and v2ray_config:
                    v2ray_parsed = V2RayConfigParser.parse(v2ray_config)
                    if not v2ray_parsed:
                        _emit_log('ERROR', 'Invalid V2Ray config', sess_id)
                        sess.status = 'completed'
                        db.session.commit()
                        socketio.emit('scan_complete', {'session_id': sess_id, 'total_found': 0}, namespace='/')
                        return
                    _v2ray_scanner.reset()
                    _v2ray_scanner.log_callback = _scanner_log
                    mode_cfg = SPEED_MODES.get(mode, SPEED_MODES['hyper'])
                    _v2ray_scanner.max_workers = max(20, int(50 * mode_cfg['resource_pct']))
                    _emit_log('INFO', f'V2Ray config parsed: {v2ray_parsed["protocol"]}', sess_id)

                # Build IPs from ranges using /24-splitting mechanism
                # For each CIDR: split to /24 blocks → random IPs per block → shuffle
                mode_cfg = SPEED_MODES.get(mode, SPEED_MODES['hyper'])
                ips_per_24 = mode_cfg.get('ips_per_24', 30)
                scan_ranges = list(ranges_input) if ranges_input else []

                if scan_method == 'operators':
                    # با رنج آی‌پی اپراتورها: فقط آی‌پی‌های CDN اسکن می‌شوند (نه رنج اپراتورها).
                    # چک می‌شود هر آی‌پی CDN روی اپراتور انتخاب‌شده پینگ و پورت دارد یا نه.
                    # رنج‌های اپراتور (ایرانسل، همراه اول، ...) فقط برای نمایش نام اپراتور استفاده می‌شوند.
                    if not scan_ranges:
                        _emit_log('INFO', 'Operator mode: fetching CDN IP ranges (not operator ranges)...', sess_id)
                        try:
                            scan_ranges = RangeFetcher.get_vfarid_cdn_ranges()
                            if scan_ranges:
                                _emit_log('INFO', f'Fetched {len(scan_ranges)} CDN ranges (vfarid)', sess_id)
                            else:
                                scan_ranges = RangeFetcher.fetch_by_source('all')
                                _emit_log('INFO', f'Fetched {len(scan_ranges)} CDN ranges (official)', sess_id)
                        except Exception as e:
                            _emit_log('ERROR', f'Failed to fetch CDN ranges: {e}', sess_id)
                    else:
                        _emit_log('INFO', f'Operator mode: using {len(scan_ranges)} CDN ranges (ping+port check)', sess_id)
                    if not scan_ranges:
                        _emit_log('WARN', 'No CDN ranges. Paste CDN ranges or click Fetch Ranges.', sess_id)

                # Batch size per round; we'll keep scanning until found >= target_count
                batch_size = max(target_count * 100, 5000) if target_count else 100000
                max_total_scanned = 500000  # safety: stop after 500k IPs tried
                total_scanned = 0

                found = 0

                # Progress: show % toward target (found/target_count) so user sees goal
                def on_progress(done, total_ips, speed=0, elapsed=0):
                    nonlocal total_scanned
                    if target_count and target_count > 0:
                        pct = min(100.0, (found / target_count) * 100.0)
                    else:
                        pct = (done / total_ips * 100) if total_ips > 0 else 0
                    socketio.emit('scan_progress', {
                        'done': total_scanned + done, 'total': total_scanned + total_ips,
                        'percent': round(pct, 1), 'speed': round(speed, 1),
                        'elapsed': round(elapsed, 1), 'session_id': sess_id,
                    }, namespace='/')

                # اپراتور فقط از انتخاب کاربر (از لیست اپراتورهای همان کشور). از IP لوکال استفاده نمی‌شود.
                # وقتی "دریافت همه آی‌پی‌های اپراتورها" زده شده، رنج‌ها از RIPE/BGP برای همه اپراتورهای آن کشور گرفته می‌شود.
                _session_operator_name = ''
                if scan_method in ('operators', 'v2ray'):
                    if operator_key:
                        operators_dict = OPERATORS_BY_COUNTRY.get(country, OPERATORS_BY_COUNTRY['ir'])
                        if operator_key in operators_dict:
                            _session_operator_name = operators_dict[operator_key].get('name', operator_key)
                            _emit_log('INFO', f'Operator: {_session_operator_name} — CDN IPs with ping on this operator will be listed', sess_id)
                    if not _session_operator_name:
                        _emit_log('INFO', 'Select an operator from the list (operator column will be empty otherwise).', sess_id)

                if not scan_ranges:
                    _emit_log('WARN', 'No CDN ranges. Paste CDN ranges or click Fetch Ranges.', sess_id)
                    sess.status = 'completed'
                    sess.duration = 0
                    db.session.commit()
                    socketio.emit('scan_complete', {'session_id': sess_id, 'total_found': 0, 'duration': 0}, namespace='/')
                    return

                _emit_log('INFO', f'Target: {target_count or "unlimited"} — scan will run until target reached or max IPs tried', sess_id)

                # Emit each result the moment it is found (real-time table update)
                def on_result(result):
                    nonlocal found
                    if target_count and found >= target_count:
                        _scanner.stop()
                        _v2ray_scanner.stop()
                        return
                    ping_val = result.get('ping')
                    if ping_val is not None and (ping_val < ping_min or ping_val > ping_max):
                        return

                    operator_name = result.get('operator') or ''
                    if scan_method in ('operators', 'v2ray'):
                        operator_name = operator_name or _session_operator_name or ''
                    operator_name = (operator_name or '').strip()

                    score = SHScanner.calc_score(result)
                    sr = ScanResult(
                        ip=result['ip'],
                        ping=result.get('ping'),
                        open_ports=json.dumps(result.get('open_ports', [])),
                        score=score,
                        operator=operator_name,
                        scan_session_id=sess_id,
                    )
                    db.session.add(sr)
                    found += 1
                    socketio.emit('scan_result', {
                        'ip': result['ip'],
                        'ping': round(result.get('ping', 0), 1) if result.get('ping') else None,
                        'open_ports': result.get('open_ports', []),
                        'score': round(score, 1),
                        'operator': operator_name,
                        'session_id': sess_id,
                        'is_v2ray': scan_method == 'v2ray',
                    }, namespace='/')
                    _emit_log('INFO', f'Found: {result["ip"]} ping={round(result.get("ping",0),1)}ms ports={result.get("open_ports",[])} op={operator_name}', sess_id)
                    try:
                        db.session.commit()
                    except Exception:
                        db.session.rollback()

                # Run scan in batches until we reach target_count or exhaust max_total_scanned
                batch_num = 0
                global _user_stop_requested
                _user_stop_requested = False
                while True:
                    if _user_stop_requested:
                        _emit_log('INFO', 'Scan stopped by user.', sess_id)
                        break
                    if target_count and found >= target_count:
                        break
                    if total_scanned >= max_total_scanned:
                        _emit_log('INFO', f'Reached max IPs to try ({max_total_scanned}). Stopping.', sess_id)
                        break

                    batch_num += 1
                    all_ips = SHNetUtils.generate_scan_ips(
                        scan_ranges,
                        per_block=ips_per_24,
                        max_total=batch_size,
                        shuffle=True,
                    )
                    if not all_ips:
                        _emit_log('WARN', 'No more IPs to generate from ranges.', sess_id)
                        break

                    _emit_log('INFO', f'Batch {batch_num}: scanning {len(all_ips)} IPs (target {target_count or "—"}, found {found} so far)', sess_id)
                    socketio.emit('scan_status', {
                        'status': 'scanning', 'total': total_scanned + len(all_ips), 'session_id': sess_id
                    }, namespace='/')

                    _scanner.reset()

                    if scan_method == 'v2ray' and v2ray_parsed:
                        _v2ray_scanner.scan_ips(
                            v2ray_parsed, [str(ip) for ip in all_ips],
                            timeout=min(8, ping_max / 1000.0 * 1.5),
                            progress_callback=on_progress,
                            result_callback=on_result,
                        )
                    else:
                        _scanner.batch_scan(
                            all_ips, ports,
                            progress_callback=on_progress,
                            result_callback=on_result,
                            start_time=start_time,
                        )

                    total_scanned += len(all_ips)
                    if target_count and found >= target_count:
                        _emit_log('INFO', f'Target reached: {found} IPs found.', sess_id)
                        break
                    if not target_count:
                        # No target (unlimited): one batch only
                        break
                    if _user_stop_requested:
                        _emit_log('INFO', 'Scan stopped by user.', sess_id)
                        break

                try:
                    db.session.commit()
                except Exception:
                    db.session.rollback()

                elapsed = time.time() - start_time
                sess = db.session.get(ScanSession, sess_id)
                if sess:
                    sess.total_scanned = total_scanned
                    sess.total_found = found
                    sess.duration = round(elapsed, 1)
                    if sess.status == 'running':
                        sess.status = 'completed'
                    sess.completed_at = datetime.utcnow()
                    db.session.commit()

                _emit_log('INFO', f'Scan complete: {found}/{total_scanned} IPs found in {elapsed:.1f}s', sess_id)
                socketio.emit('scan_complete', {
                    'session_id': sess_id,
                    'total_scanned': total_scanned,
                    'total_found': found,
                    'duration': round(elapsed, 1),
                }, namespace='/')

            except Exception as e:
                tb = traceback.format_exc()
                _emit_log('ERROR', f'Scan error: {str(e)}\n{tb}', sess_id)
                try:
                    sess = db.session.get(ScanSession, sess_id)
                    if sess:
                        sess.status = 'error'
                        db.session.commit()
                except Exception:
                    db.session.rollback()
                socketio.emit('scan_error', {'error': str(e), 'session_id': sess_id}, namespace='/')

    thread = threading.Thread(target=run_scan, daemon=True)
    thread.start()

    return jsonify({'session_id': sess_id, 'status': 'started'})


@api_bp.route('/scan/stop', methods=['POST'])
def stop_scan():
    global _active_session_id, _user_stop_requested
    _user_stop_requested = True
    _scanner.stop()
    _v2ray_scanner.stop()
    if _active_session_id:
        session = db.session.get(ScanSession, _active_session_id)
        if session and session.status == 'running':
            session.status = 'stopped'
            session.completed_at = datetime.utcnow()
            db.session.commit()
    return jsonify({'status': 'stopped'})


@api_bp.route('/scan/results', methods=['GET'])
def get_results():
    session_id = request.args.get('session_id', type=int)
    limit = request.args.get('limit', 200, type=int)
    q = ScanResult.query
    if session_id:
        q = q.filter_by(scan_session_id=session_id)
    results = q.order_by(ScanResult.score.desc()).limit(limit).all()
    return jsonify([r.to_dict() for r in results])


@api_bp.route('/scan/sessions', methods=['GET'])
def get_sessions():
    sessions = ScanSession.query.order_by(ScanSession.id.desc()).limit(20).all()
    return jsonify([s.to_dict() for s in sessions])


@api_bp.route('/scan/logs', methods=['GET'])
def get_logs():
    session_id = request.args.get('session_id', type=int)
    limit = request.args.get('limit', 100, type=int)
    q = ScanLog.query
    if session_id:
        q = q.filter_by(session_id=session_id)
    logs = q.order_by(ScanLog.id.desc()).limit(limit).all()
    return jsonify([l.to_dict() for l in logs])


# ========== Export ==========

# Excel/Text export headers by language (standard table columns)
EXPORT_HEADERS = {
    'fa': ('رتبه', 'آدرس IP', 'Ping', 'پورت\u200cها', 'امتیاز', 'اپراتور'),
    'en': ('#', 'IP', 'Ping', 'Ports', 'Score', 'Operator'),
}


@api_bp.route('/export/<fmt>', methods=['GET'])
def export_results(fmt):
    from flask import Response
    session_id = request.args.get('session_id', type=int)
    lang = (request.args.get('lang') or 'en').strip().lower()
    if lang not in EXPORT_HEADERS:
        lang = 'en'
    q = ScanResult.query
    if session_id:
        q = q.filter_by(scan_session_id=session_id)
    results = q.order_by(ScanResult.score.desc()).all()

    if fmt == 'json':
        output = json.dumps([r.to_dict() for r in results], indent=2, ensure_ascii=False)
        return Response(output, mimetype='application/json',
                        headers={'Content-Disposition': 'attachment;filename=scan_results.json'})
    elif fmt == 'txt':
        # Only IPs, one per line (column-style)
        lines = [r.ip for r in results]
        return Response('\n'.join(lines), mimetype='text/plain',
                        headers={'Content-Disposition': 'attachment;filename=scan_ips.txt'})
    elif fmt == 'excel':
        try:
            import openpyxl
            from io import BytesIO
        except ImportError:
            return jsonify({'error': 'Excel export requires openpyxl'}), 500
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Results'
        headers = EXPORT_HEADERS[lang]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        for row_idx, r in enumerate(results, 2):
            ports_list = json.loads(r.open_ports) if r.open_ports else []
            ports_str = ' '.join(str(p) for p in ports_list) if ports_list else ''
            ws.cell(row=row_idx, column=1, value=row_idx - 1)
            ws.cell(row=row_idx, column=2, value=r.ip or '')
            ws.cell(row=row_idx, column=3, value=round(r.ping, 1) if r.ping is not None else '')
            ws.cell(row=row_idx, column=4, value=ports_str)
            ws.cell(row=row_idx, column=5, value=round(r.score, 1) if r.score is not None else '')
            ws.cell(row=row_idx, column=6, value=(r.operator or ''))
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            buf.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment;filename=scan_results.xlsx'}
        )
    else:
        return jsonify({'error': 'Unsupported format'}), 400


# ========== Reset ==========

@api_bp.route('/reset', methods=['POST'])
def reset_data():
    ScanResult.query.delete()
    ScanSession.query.delete()
    ClosedIP.query.delete()
    ScanLog.query.delete()
    db.session.commit()
    return jsonify({'status': 'ok'})


# ========== Info ==========

@api_bp.route('/info', methods=['GET'])
def app_info():
    return jsonify({
        'version': current_app.config.get('VERSION', '2.0'),
        'author': current_app.config.get('AUTHOR', 'shahinst'),
        'github': current_app.config.get('GITHUB_REPO_URL', ''),
        'speed_modes': {k: v['label'] for k, v in SPEED_MODES.items()},
    })


# ========== Update System ==========

@api_bp.route('/check-update', methods=['POST'])
def check_update():
    """Check GitHub for newer version."""
    import requests as req
    current_version = current_app.config.get('VERSION', '2.0')
    version_url = 'https://raw.githubusercontent.com/shahinst/cdn-ip-scanner/main/version'
    try:
        r = req.get(version_url, timeout=10)
        if r.status_code != 200:
            return jsonify({'error': f'Failed to fetch version (HTTP {r.status_code})'}), 200
        remote_version = r.text.strip()
        if not remote_version:
            return jsonify({'error': 'Empty version file'}), 200

        # Compare versions (simple string/numeric compare)
        def parse_ver(v):
            parts = []
            for p in v.replace('-', '.').split('.'):
                try:
                    parts.append(int(p))
                except ValueError:
                    parts.append(0)
            return parts

        rv = parse_ver(remote_version)
        cv = parse_ver(current_version)
        update_available = rv > cv

        return jsonify({
            'current_version': current_version,
            'remote_version': remote_version,
            'update_available': update_available,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 200


@api_bp.route('/do-update', methods=['POST'])
def do_update():
    """Pull latest from GitHub and restart."""
    import subprocess
    import sys
    import os

    base_dir = current_app.config.get('BASE_DIR', os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

    try:
        # Check if git is available and this is a git repo
        git_check = subprocess.run(['git', 'status'], cwd=base_dir,
                                   capture_output=True, text=True, timeout=10)
        if git_check.returncode != 0:
            # Not a git repo - try git clone approach
            return jsonify({'error': 'Not a git repository. Please install via git clone.'}), 200

        # Pull latest changes
        pull_result = subprocess.run(['git', 'pull', 'origin', 'main'], cwd=base_dir,
                                     capture_output=True, text=True, timeout=60)
        if pull_result.returncode != 0:
            # Try with --rebase
            pull_result = subprocess.run(['git', 'pull', '--rebase', 'origin', 'main'], cwd=base_dir,
                                         capture_output=True, text=True, timeout=60)

        # Install updated dependencies
        req_file = os.path.join(base_dir, 'requirements.txt')
        if os.path.exists(req_file):
            subprocess.run([sys.executable, '-m', 'pip', 'install', '-r', req_file, '-q'],
                          cwd=base_dir, capture_output=True, text=True, timeout=120)

        # Schedule restart in a background thread
        def restart_app():
            import time as _time
            _time.sleep(2)
            os.execv(sys.executable, [sys.executable] + sys.argv)

        restart_thread = threading.Thread(target=restart_app, daemon=True)
        restart_thread.start()

        return jsonify({
            'status': 'ok',
            'message': 'Update complete. Restarting...',
            'git_output': pull_result.stdout or pull_result.stderr,
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 200
