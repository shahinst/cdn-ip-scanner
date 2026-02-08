#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CDN IP Scanner V1.6
Author: شاهین سالک توتونچی (shahin salek tootoonchi)
GitHub: github.com/shahinst
Website: digicloud.tr
YouTube: https://www.youtube.com/@shaahinst
"""

import os
import re
import sys
import socket
import math
import json
import time
import random
from queue import Queue
from datetime import datetime

import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, filedialog
from tkinter import font as tkfont

import ipaddress
import requests
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
import webbrowser
import threading
import zipfile
import tempfile
import subprocess
import shutil
from concurrent.futures import ThreadPoolExecutor, as_completed

# مسیر پایه و مسیر منابع: در حالت exe (one-file) منابع در _MEIPASS، دادهٔ کاربر در پوشهٔ exe
if getattr(sys, "frozen", False):
    BASE_DIR = os.path.dirname(sys.executable)
    RESOURCE_DIR = getattr(sys, "_MEIPASS", BASE_DIR)
else:
    BASE_DIR = RESOURCE_DIR = os.path.dirname(os.path.abspath(__file__))

# بارگذاری فونت از فایل در ویندوز (Windows API) — Vazirmatn، EV Sam، Font Awesome 6
def _load_fonts_win(font_dir):
    if sys.platform != "win32":
        return
    try:
        from ctypes import windll, byref, create_unicode_buffer
        FR_PRIVATE = 0x10
        names = (
            "Vazirmatn-Regular.ttf", "Vazirmatn-Bold.ttf",
            "EV-Sam-Regular.ttf", "EV-Sam-Bold.ttf", "EVSam-Regular.ttf", "EVSam-Bold.ttf",
            "fa-solid-900.ttf", "fa-brands-400.ttf",  # Font Awesome 6 Free
        )
        for name in names:
            path = os.path.join(font_dir, name)
            if os.path.isfile(path):
                path_buf = create_unicode_buffer(path)
                windll.gdi32.AddFontResourceExW(byref(path_buf), FR_PRIVATE, None)
    except Exception:
        pass

# Font Awesome 6 Free — یونیکد آیکن‌ها (PUA)
FA = {
    "rocket": "\uf135", "gear": "\uf013", "bolt": "\uf0e7", "check": "\uf00c", "clock": "\uf017",
    "bullseye": "\uf140", "satellite": "\uf7c0", "robot": "\uf544", "save": "\uf0c7", "stop": "\uf04d",
    "play": "\uf04b", "chart": "\uf201", "star": "\uf005", "lock_open": "\uf3c1", "location": "\uf3c5",
    "trophy": "\uf091", "gem": "\uf3a5",     "moon": "\uf186", "sun": "\uf185",
    "github": "\uf09b", "youtube": "\uf167",  # brands
    "download": "\uf019",  # بروزرسانی / دانلود
}
# رنگ پیشنهادی برای آیکن‌ها (نام کلید تم یا هگز)
FA_COLORS = {
    "rocket": "#6366f1", "gear": "#8b5cf6", "bolt": "#eab308", "check": "#22c55e", "clock": "#0ea5e9",
    "bullseye": "#f97316", "satellite": "#06b6d4", "robot": "#a855f7", "save": "#3b82f6", "stop": "#ef4444",
    "play": "#22c55e", "chart": "#8b5cf6", "star": "#eab308", "lock_open": "#10b981", "location": "#ec4899",
    "trophy": "#f59e0b", "gem": "#06b6d4", "moon": "#64748b", "sun": "#eab308",
    "github": "#f0f6fc", "youtube": "#ef4444",
    "download": "#6b7280",
}

# Try to import openpyxl for Excel export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# ===== Author & Links =====
AUTHOR_NAME = "شاهین سالک توتونچی"
GITHUB_URL = "https://github.com/shahinst"
GITHUB_REPO_URL = "https://github.com/shahinst/cdn-ip-scanner"  # پروژه گیت‌هاب
GITHUB_RAW_VERSION_URL = "https://raw.githubusercontent.com/shahinst/cdn-ip-scanner/main/version"
GITHUB_ZIP_URL = "https://github.com/shahinst/cdn-ip-scanner/archive/refs/heads/main.zip"
WEBSITE_URL = "https://digicloud.tr"
YOUTUBE_URL = "https://www.youtube.com/@shaahinst"

# ===== Translations (en, fa, zh, ru) =====
TRANSLATIONS = {
    "en": {
        "app_title": "CDN IP Scanner V1.6",
        "app_subtitle": "High accuracy • Ultra fast • AI powered ⚡",
        "loading": "Loading...",
        "choose_language": "Choose language",
        "target": "Target",
        "speed": "Speed",
        "found": "Found",
        "time": "Time",
        "settings": "Settings ⚙️",
        "mode": "Mode 🔥",
        "ai": "AI 🤖",
        "count": "Count 🎯",
        "fetch_ranges_btn": "Fetch Ranges 📡",
        "analyze_btn": "AI Analyze 🤖",
        "save_btn": "Save 💾",
        "stop_btn": "Stop ⏹️",
        "start_btn": "Start Scan 🚀",
        "progress_title": "Progress 📊",
        "ready_status": "Ready to start ⚡",
        "results_title": "Results (by speed) 🎯",
        "scan_status_log": "Scan status (each IP):",
        "score_hdr": "Score ⭐",
        "ports_hdr": "Ports 🔓",
        "ping_hdr": "Ping ⚡",
        "ip_hdr": "IP 📍",
        "rank_hdr": "Rank 🏆",
        "made_by": "Made with ❤️ by {}",
        "author_display": "shahin salek tootoonchi",
        "website_link_text": "Buy server",
        "donate_text": "💎 Donate (USDT): 0xde200d...c269",
        "github_text": "GitHub ⭐",
        "youtube_text": "YouTube 📺",
        "donate_title": "Support Development 💎",
        "donate_desc": "If this tool helped you, you can support with a donation:",
        "copy_btn": "Copy Address 📋",
        "close_btn": "Close",
        "thanks": "🙏 Thank you for your support!",
        "range_fetcher_title": "Fetch IP Ranges 📡",
        "range_sources_title": "Fetch IP ranges from sources 📡",
        "ranges_received": "Ranges received 📋:",
        "ready_fetch": "Ready to fetch...",
        "fetching": "Fetching...",
        "error_fetch": "Error fetching ranges!",
        "cf_api": "Official API - Cloudflare ☁️",
        "cf_asn": "ASN - Cloudflare 🔢",
        "cf_github": "GitHub - Cloudflare 🐙",
        "fastly_api": "Official API - Fastly ⚡",
        "fastly_asn": "ASN - Fastly 🔢",
        "all_sources": "All sources (recommended) 🌐",
        "export_menu_title": "Choose export format 💾",
        "excel_btn": "Excel (xlsx) 📊",
        "json_btn": "JSON (json) 📄",
        "txt_btn": "TXT (IPs only, one per line) 📄",
        "settings_ping_range": "Ping range filter (ms)",
        "ping_min_label": "Min ping (ms):",
        "ping_max_label": "Max ping (ms):",
        "ping_filter_display": "Ping filter: {}–{} ms",
        "ping_filter_none": "No ping filter",
        "settings_ports": "Scan ports",
        "ports_label": "Ports (comma-separated, e.g. 443,80,2010):",
        "ports_hint": "Default: 443,80,8443,2053,2083,2087,2096",
        "ranges_hint_paste_fetch": "You can paste your IPs here or fetch IP ranges. One IP or CIDR per line.",
        "add_range_manual": "Add range / IP manually:",
        "add_range_btn": "Add",
        "box_range_and_scan": "Select range and scan method",
        "scan_method_label": "Scan method",
        "scan_method_cloud": "Cloud scan",
        "scan_method_operators": "With operator IP ranges",
        "scan_method_v2ray": "With custom config in V2rayN",
        "scan_method_cloud_tooltip": "Through this method, only IPs that you enter manually or IPs that you select via \"Fetch Ranges\" are checked using your local internet.",
        "coming_soon_scan": "Updating; will be added soon.",
        "operator_fetch_btn": "Fetch",
        "fetch_all_operators_btn": "Fetch all operator IPs 📡",
        "filter_only_clean_label": "Results: only IPs with open port + ping",
        "filter_show_all_label": "Results: show all (including closed)",
        "scan_from_operator": "Scan from operator:",
        "scan_from_operator_all": "All",
        "operator_scan_basis": "Like cf-ip-scanner: each IP is tested with HTTP/HTTPS GET to /cdn-cgi/trace (Cloudflare). Several attempts per port; only if enough succeed and latency is under max, the IP/port is marked active. Scan runs from your current connection (one operator at a time). To see status on all operators, run a separate scan from each.",
        "operator_ranges_count": "{} ranges",
        "operator_fetch_error": "Fetch error",
        "operator_fetching": "Fetching...",
        "invalid_ip_cidr": "Invalid IP or CIDR",
        "msg_no_results": "No results to save!",
        "msg_saved": "File saved!",
        "msg_error": "Error",
        "msg_copied": "Copied!",
        "ip_copied": "IP copied: {}",
        "error_title": "Error",
        "success_title": "Success",
        "analyze_title": "AI Analysis",
        "no_results_analyze": "No results to analyze!",
        "openpyxl_missing": "openpyxl not installed. Run: pip install openpyxl",
        "file_saved": "File saved!\n{}",
        "save_error": "Save error:\n{}",
        "scan_done": "Done! {} IP in {}s ✅",
        "mode_turbo": "Turbo (10s)", "mode_hyper": "Hyper (5s)", "mode_ultra": "Ultra (15s)", "mode_deep": "Deep (30s)",
        "mode_turbo_desc": "700 processes, 30 IPs per range",
        "mode_hyper_desc": "600 processes, 50 IPs per range",
        "mode_ultra_desc": "500 processes, 70 IPs per range",
        "mode_deep_desc": "450 processes, 100 IPs per range",
        "ai_basic": "Basic", "ai_smart": "Smart", "ai_advanced": "Advanced", "ai_expert": "Expert",
        "target_50": "50", "target_100": "100", "target_150": "150", "target_200": "200", "target_500": "500", "target_1000": "1000", "target_all": "All",
        "count_desc": "Number of IPs to scan (exact limit).",
        "coming_future_updates": "Will be added in future updates.",
        "range_source_label": "Range source:",
        "range_source_none": "—",
        "fetch_status": "{} ranges from {} ✅",
        "source_cf_api": "Cloudflare API", "source_cf_asn": "Cloudflare ASN", "source_cf_github": "Cloudflare GitHub",
        "source_fastly_api": "Fastly API", "source_fastly_asn": "Fastly ASN", "source_all": "All Sources",
        "status_fetch": "Fetching ranges... 🔍",
        "status_analyze": "Analyzing {} ranges... 🤖",
        "status_scan": "Scanning {} ranges... ⚡",
        "status_pct": "Scan... {}% ⚡",
        "made_by_simple": "Programming and design by: shahin salek tootoonchi",
        "settings_main": "Main settings",
        "settings_theme": "Theme / Color",
        "settings_ip_check": "IP check method",
        "ip_check_my_internet": "Check IPs with my internet",
        "ip_check_agent": "Use agent for search",
        "ip_check_agent_message": "This feature will be added in future updates. For now you can only search using your own internet.",
        "settings_saved": "Settings saved.",
        "rescan_closed_title": "Previously scanned IPs",
        "rescan_closed_message": "You have {} IP(s) that were previously scanned and had no open ports. Do you want to rescan them too?",
        "rescan_yes": "Yes, rescan them",
        "rescan_no": "No, skip them",
        "ping_unsuitable": "Unsuitable",
        "update_btn": "Update",
        "update_available_title": "Update available",
        "update_available_message": "A new version is available. Update now?",
        "update_confirm_question": "Do you want to update to the latest version?",
        "update_downloading": "Downloading update...",
        "update_done": "Update completed.",
        "update_restart_question": "Restart to use the new version?",
        "update_restart_yes": "Yes, restart",
        "update_restart_no": "No",
        "update_no_new": "You are on the latest version.",
        "update_checking": "Checking for updates...",
        "update_error": "Update error",
        "operator_country": "Operator country:",
        "country_iran": "Iran",
        "country_russia": "Russia",
        "country_china": "China",
        "select_range_first": "Select your range from the active ranges first, then start the scan.",
        "ranges_to_scan_label": "Ranges to scan (check/uncheck):",
        "no_ranges_loaded": "No ranges loaded. Fetch ranges from Cloudflare or Fastly API first.",
        "rescan_blocked_message": "Last time you had {} blocked IP(s) (no ping, no open ports). Rescan all of them or scan the rest only?",
        "target_operators": "Operators",
        "operators_summary": "{}: {}",
        "white_hdr": "White",
        "white_yes": "Yes",
        "white_no": "No",
        "operator_hdr": "Operator",
        "reset_data_btn": "Reset data 🔄",
        "reset_data_confirm": "Delete all config files (settings, last scan, cache, operator data) and restart the app?",
        "reset_data_done": "Data deleted. Restarting...",
    },
    "fa": {
        "app_title": "CDN IP Scanner V1.6",
        "app_subtitle": "دقت بالا • سرعت فوق‌العاده • قدرت هوش مصنوعی ⚡",
        "loading": "در حال بارگذاری...",
        "choose_language": "زبان را انتخاب کنید",
        "target": "هدف",
        "speed": "سرعت",
        "found": "یافت شده",
        "time": "زمان",
        "settings": "تنظیمات ⚙️",
        "mode": "حالت 🔥",
        "ai": "AI 🤖",
        "count": "تعداد 🎯",
        "fetch_ranges_btn": "دریافت رنج‌ها 📡",
        "analyze_btn": "تحلیل AI 🤖",
        "save_btn": "ذخیره 💾",
        "stop_btn": "توقف ⏹️",
        "start_btn": "شروع اسکن 🚀",
        "progress_title": "پیشرفت 📊",
        "ready_status": "آماده برای شروع ⚡",
        "results_title": "نتایج (مرتب شده بر اساس سرعت) 🎯",
        "scan_status_log": "وضعیت اسکن (هر آی‌پی):",
        "score_hdr": "امتیاز ⭐",
        "ports_hdr": "پورت‌ها 🔓",
        "ping_hdr": "Ping ⚡",
        "ip_hdr": "آدرس IP 📍",
        "rank_hdr": "رتبه 🏆",
        "made_by": "ساخته شده با ❤️ توسط {}",
        "author_display": AUTHOR_NAME,
        "website_link_text": "خرید سرور",
        "donate_text": "💎 Donate (USDT): 0xde200d...c269",
        "github_text": "GitHub ⭐",
        "youtube_text": "یوتیوب 📺",
        "donate_title": "حمایت از توسعه 💎",
        "donate_desc": "اگر این ابزار برات مفید بود، می‌تونی با دونیت حمایت کنی:",
        "copy_btn": "کپی آدرس 📋",
        "close_btn": "بستن",
        "thanks": "🙏 ممنون از حمایتت!",
        "range_fetcher_title": "دریافت رنج‌های IP 📡",
        "range_sources_title": "دریافت رنج‌های IP از منابع مختلف 📡",
        "ranges_received": "رنج‌های دریافت شده 📋:",
        "ready_fetch": "آماده برای دریافت...",
        "fetching": "در حال دریافت...",
        "error_fetch": "خطا در دریافت رنج‌ها!",
        "cf_api": "API رسمی - Cloudflare ☁️",
        "cf_asn": "ASN - Cloudflare 🔢",
        "cf_github": "GitHub - Cloudflare 🐙",
        "fastly_api": "API رسمی - Fastly ⚡",
        "fastly_asn": "ASN - Fastly 🔢",
        "all_sources": "(همه منابع (توصیه میشه! 🌐",
        "export_menu_title": "انتخاب فرمت خروجی 💾",
        "excel_btn": "Excel (xlsx) 📊",
        "json_btn": "JSON (json) 📄",
        "txt_btn": "TXT (فقط آی‌پی، هر خط یک آی‌پی) 📄",
        "settings_ping_range": "فیلتر محدوده پینگ (ms)",
        "ping_min_label": "کمترین پینگ (ms):",
        "ping_max_label": "بیشترین پینگ (ms):",
        "ping_filter_display": "فیلتر پینگ: {}–{} ms",
        "ping_filter_none": "بدون فیلتر پینگ",
        "settings_ports": "پورت‌های اسکن",
        "ports_label": "پورت‌ها (با کاما، مثلاً 443,80,2010):",
        "ports_hint": "پیش‌فرض: 443,80,8443,2053,2083,2087,2096",
        "ranges_hint_paste_fetch": "می‌توانید آی‌پی‌های خود را اینجا پیست کنید یا رنج آی‌پی بگیرید. هر خط یک آی‌پی یا CIDR.",
        "add_range_manual": "افزودن رنج / آی‌پی دستی:",
        "add_range_btn": "افزودن",
        "box_range_and_scan": "انتخاب رنج و نحوه اسکن",
        "scan_method_label": "نحوه اسکن",
        "scan_method_cloud": "اسکن کلود",
        "scan_method_operators": "با رنج آی پی اپراتور ها",
        "scan_method_v2ray": "با کانفیگ دلخواه در V2rayN",
        "scan_method_cloud_tooltip": "از این طریق تنها آی پی هایی که خودتان به صورت دستی وارد می کنید و یا آی پی هایی که از طریق گزینه دریافت رنج ها انتخاب می کنید را با استفاده از اینترنت لوکال خودتان بررسی می کند.",
        "coming_soon_scan": "در حال آپدیت و به زودی قرار داده می شود.",
        "operator_fetch_btn": "دریافت",
        "fetch_all_operators_btn": "دریافت همهٔ آی‌پی‌های اپراتورها 📡",
        "filter_only_clean_label": "نتایج: فقط آی‌پی‌های تمیز (باز و پینگ‌دار)",
        "filter_show_all_label": "نتایج: همه (از جمله بسته)",
        "scan_from_operator": "اسکن از اپراتور:",
        "scan_from_operator_all": "همه",
        "operator_scan_basis": "مشابه cf-ip-scanner: هر آی‌پی با درخواست HTTP/HTTPS به /cdn-cgi/trace (کلودفلیر) تست می‌شود. برای هر پورت چند بار تلاش می‌شود؛ فقط در صورت موفقیت کافی و تأخیر زیر حداکثر، آی‌پی/پورت «فعال» در نظر گرفته می‌شود. اسکن از اتصال فعلی شما (در هر لحظه یک اپراتور) انجام می‌شود. برای وضعیت روی همهٔ اپراتورها، از هر اپراتور جداگانه اسکن کنید.",
        "operator_ranges_count": "{} رنج دریافت شده",
        "operator_fetch_error": "خطا در دریافت",
        "operator_fetching": "در حال دریافت...",
        "invalid_ip_cidr": "آی‌پی یا CIDR نامعتبر",
        "msg_no_results": "نتیجه‌ای برای ذخیره وجود ندارد!",
        "msg_saved": "فایل ذخیره شد!",
        "msg_error": "خطا",
        "msg_copied": "کپی شد!",
        "ip_copied": "IP کپی شد: {}",
        "error_title": "خطا",
        "success_title": "موفق",
        "analyze_title": "تحلیل AI",
        "no_results_analyze": "نتیجه‌ای برای تحلیل وجود ندارد!",
        "openpyxl_missing": "کتابخانه openpyxl نصب نیست. برای نصب: pip install openpyxl",
        "file_saved": "فایل ذخیره شد!\n{}",
        "save_error": "خطا در ذخیره فایل:\n{}",
        "scan_done": "اتمام! {} IP در {}s ✅",
        "mode_turbo": "Turbo (۱۰s)", "mode_hyper": "Hyper (۵s)", "mode_ultra": "Ultra (۱۵s)", "mode_deep": "Deep (۳۰s)",
        "mode_turbo_desc": "۷۰۰ پروسس، ۳۰ آی‌پی در هر رنج",
        "mode_hyper_desc": "۶۰۰ پروسس، ۵۰ آی‌پی در هر رنج",
        "mode_ultra_desc": "۵۰۰ پروسس، ۷۰ آی‌پی در هر رنج",
        "mode_deep_desc": "۴۵۰ پروسس، ۱۰۰ آی‌پی در هر رنج",
        "ai_basic": "Basic", "ai_smart": "Smart", "ai_advanced": "Advanced", "ai_expert": "Expert",
        "target_50": "۵۰", "target_100": "۱۰۰", "target_150": "۱۵۰", "target_200": "۲۰۰", "target_500": "۵۰۰", "target_1000": "۱۰۰۰", "target_all": "همه",
        "count_desc": "تعداد آی‌پی که اسکن می‌شود (حد دقیق).",
        "coming_future_updates": "در آپدیت‌های بعدی قرار داده خواهد شد.",
        "range_source_label": "منبع رنج:",
        "range_source_none": "—",
        "fetch_status": "{} رنج از {} ✅",
        "source_cf_api": "Cloudflare API", "source_cf_asn": "Cloudflare ASN", "source_cf_github": "Cloudflare GitHub",
        "source_fastly_api": "Fastly API", "source_fastly_asn": "Fastly ASN", "source_all": "همه منابع",
        "status_fetch": "دریافت رنج‌ها... 🔍",
        "status_analyze": "تحلیل {} رنج... 🤖",
        "status_scan": "اسکن {} رنج... ⚡",
        "status_pct": "اسکن... {}% ⚡",
        "made_by_simple": "برنامه نویسی و طراحی توسط : شاهین سالک توتونچی",
        "settings_main": "تنظیمات اصلی",
        "settings_theme": "تغییر رنگ / تم",
        "settings_ip_check": "تغییر بررسی آی پی ها",
        "ip_check_my_internet": "بررسی آی پی ها با اینترنت خودم",
        "ip_check_agent": "جستجو از طریق ایجنت",
        "ip_check_agent_message": "این قابلیت در به‌روزرسانی‌های بعدی اضافه خواهد شد. در حال حاضر فقط از طریق اینترنت خودتان می‌توانید جستجو را انجام دهید.",
        "settings_saved": "تنظیمات ذخیره شد.",
        "rescan_closed_title": "آی‌پی‌های اسکن‌شدهٔ قبلی",
        "rescan_closed_message": "{} آی‌پی که قبلاً اسکن شده و پورت باز نداشتند در حافظه هست. می‌خواهید آن‌ها را دوباره هم اسکن کنم؟",
        "rescan_yes": "بله، دوباره اسکن کن",
        "rescan_no": "خیر، ردشون کن",
        "ping_unsuitable": "نامناسب",
        "update_btn": "بروزرسانی",
        "update_available_title": "بروزرسانی موجود",
        "update_available_message": "نسخهٔ جدید موجود است. الان بروزرسانی شود؟",
        "update_confirm_question": "می‌خواهید به آخرین نسخه بروزرسانی کنید؟",
        "update_downloading": "در حال دریافت بروزرسانی...",
        "update_done": "بروزرسانی انجام شد.",
        "update_restart_question": "برنامه ریستارت شود تا نسخهٔ جدید اجرا شود؟",
        "update_restart_yes": "بله، ریستارت",
        "update_restart_no": "خیر",
        "update_no_new": "همین الان روی آخرین نسخه هستید.",
        "update_checking": "در حال بررسی بروزرسانی...",
        "update_error": "خطای بروزرسانی",
        "operator_country": "کشور اپراتور:",
        "country_iran": "ایران",
        "country_russia": "روسیه",
        "country_china": "چین",
        "select_range_first": "رنجت رو از بین رنج‌های فعال انتخاب کن، بعد اسکن رو شروع کن.",
        "ranges_to_scan_label": "رنج‌هایی که اسکن شوند:",
        "no_ranges_loaded": "رنجی بارگذاری نشده. اول از Cloudflare یا Fastly API رنج بگیر.",
        "rescan_blocked_message": "دفعهٔ قبل {} آی‌پی بلاک‌شده داشتی (بدون پینگ و بدون پورت باز). همه رو دوباره اسکن کنم یا فقط بقیه رو اسکن کنم؟",
        "target_operators": "اپراتورها",
        "operators_summary": "{}: {}",
        "white_hdr": "وایت",
        "white_yes": "بله",
        "white_no": "خیر",
        "operator_hdr": "اپراتور",
        "reset_data_btn": "حذف داده‌های قبلی 🔄",
        "reset_data_confirm": "همهٔ فایل‌های کانفیگ (تنظیمات، آخرین اسکن، کش، دادهٔ اپراتورها) حذف شوند و برنامه دوباره راه‌اندازی شود؟",
        "reset_data_done": "داده‌ها حذف شد. در حال راه‌اندازی مجدد...",
    },
    "zh": {
        "app_title": "CDN IP 扫描器 V1.6",
        "app_subtitle": "高精度 • 超快 • AI 驱动 ⚡",
        "loading": "加载中...",
        "choose_language": "选择语言",
        "target": "目标",
        "speed": "速度",
        "found": "已找到",
        "time": "时间",
        "settings": "设置 ⚙️",
        "mode": "模式 🔥",
        "ai": "AI 🤖",
        "count": "数量 🎯",
        "fetch_ranges_btn": "获取范围 📡",
        "analyze_btn": "AI 分析 🤖",
        "save_btn": "保存 💾",
        "stop_btn": "停止 ⏹️",
        "start_btn": "开始扫描 🚀",
        "progress_title": "进度 📊",
        "ready_status": "准备就绪 ⚡",
        "results_title": "结果（按速度）🎯",
        "scan_status_log": "扫描状态（每个IP）：",
        "score_hdr": "分数 ⭐",
        "ports_hdr": "端口 🔓",
        "ping_hdr": "Ping ⚡",
        "ip_hdr": "IP 📍",
        "rank_hdr": "排名 🏆",
        "made_by": "由 {} 用 ❤️ 制作",
        "author_display": "shahin salek tootoonchi",
        "website_link_text": "购买服务器",
        "donate_text": "💎 Donate (USDT): 0xde200d...c269",
        "github_text": "GitHub ⭐",
        "youtube_text": "YouTube 📺",
        "donate_title": "支持开发 💎",
        "donate_desc": "如果此工具对您有帮助，欢迎捐赠支持：",
        "copy_btn": "复制地址 📋",
        "close_btn": "关闭",
        "thanks": "🙏 感谢您的支持！",
        "range_fetcher_title": "获取 IP 范围 📡",
        "range_sources_title": "从各来源获取 IP 范围 📡",
        "ranges_received": "已获取范围 📋:",
        "ready_fetch": "准备获取...",
        "fetching": "获取中...",
        "error_fetch": "获取范围出错！",
        "cf_api": "官方 API - Cloudflare ☁️",
        "cf_asn": "ASN - Cloudflare 🔢",
        "cf_github": "GitHub - Cloudflare 🐙",
        "fastly_api": "官方 API - Fastly ⚡",
        "fastly_asn": "ASN - Fastly 🔢",
        "all_sources": "全部来源（推荐）🌐",
        "export_menu_title": "选择导出格式 💾",
        "excel_btn": "Excel (xlsx) 📊",
        "json_btn": "JSON (json) 📄",
        "txt_btn": "TXT（仅IP，每行一个）📄",
        "settings_ping_range": "Ping 范围过滤 (ms)",
        "ping_min_label": "最小 Ping (ms):",
        "ping_max_label": "最大 Ping (ms):",
        "ping_filter_display": "Ping 过滤: {}–{} ms",
        "ping_filter_none": "无 Ping 过滤",
        "settings_ports": "扫描端口",
        "ports_label": "端口（逗号分隔，如 443,80,2010）:",
        "ports_hint": "默认: 443,80,8443,2053,2083,2087,2096",
        "ranges_hint_paste_fetch": "可在此粘贴 IP 或获取 IP 范围。每行一个 IP 或 CIDR。",
        "add_range_manual": "手动添加范围 / IP:",
        "add_range_btn": "添加",
        "box_range_and_scan": "选择范围与扫描方式",
        "scan_method_label": "扫描方式",
        "scan_method_cloud": "云扫描",
        "scan_method_operators": "运营商 IP 范围",
        "scan_method_v2ray": "V2rayN 自定义配置",
        "scan_method_cloud_tooltip": "通过此方式，仅使用您的本地网络检查您手动输入的 IP 或通过「获取范围」选择的 IP。",
        "coming_soon_scan": "正在更新，即将推出。",
        "operator_fetch_btn": "Fetch",
        "fetch_all_operators_btn": "Fetch all operator IPs 📡",
        "filter_only_clean_label": "Results: only IPs with open port + ping",
        "filter_show_all_label": "Results: show all (including closed)",
        "scan_from_operator": "Scan from operator:",
        "scan_from_operator_all": "All",
        "operator_scan_basis": "Like cf-ip-scanner: each IP is tested with HTTP/HTTPS GET to /cdn-cgi/trace (Cloudflare). Several attempts per port; only if enough succeed and latency is under max, the IP/port is marked active. Scan runs from your current connection (one operator at a time). To see status on all operators, run a separate scan from each.",
        "operator_ranges_count": "{} ranges",
        "operator_fetch_error": "Fetch error",
        "operator_fetching": "Fetching...",
        "invalid_ip_cidr": "无效的 IP 或 CIDR",
        "msg_no_results": "没有可保存的结果！",
        "msg_saved": "文件已保存！",
        "msg_error": "错误",
        "msg_copied": "已复制！",
        "ip_copied": "IP 已复制: {}",
        "error_title": "错误",
        "success_title": "成功",
        "analyze_title": "AI 分析",
        "no_results_analyze": "没有可分析的结果！",
        "openpyxl_missing": "未安装 openpyxl。运行: pip install openpyxl",
        "file_saved": "文件已保存！\n{}",
        "save_error": "保存错误:\n{}",
        "scan_done": "完成！{} 个 IP，耗时 {}s ✅",
        "mode_turbo": "Turbo (10s)", "mode_hyper": "Hyper (5s)", "mode_ultra": "Ultra (15s)", "mode_deep": "Deep (30s)",
        "mode_turbo_desc": "700 processes, 30 IPs per range",
        "mode_hyper_desc": "600 processes, 50 IPs per range",
        "mode_ultra_desc": "500 processes, 70 IPs per range",
        "mode_deep_desc": "450 processes, 100 IPs per range",
        "ai_basic": "Basic", "ai_smart": "Smart", "ai_advanced": "Advanced", "ai_expert": "Expert",
        "target_50": "50", "target_100": "100", "target_150": "150", "target_200": "200", "target_500": "500", "target_1000": "1000", "target_all": "All",
        "count_desc": "要扫描的 IP 数量（精确限制）。",
        "coming_future_updates": "将在后续更新中添加。",
        "select_range_first": "请先从可用范围中选择范围，然后再开始扫描。",
        "ranges_to_scan_label": "要扫描的范围（勾选/取消）：",
        "no_ranges_loaded": "未加载范围。请先从 Cloudflare 或 Fastly API 获取范围。",
        "rescan_blocked_message": "上次您有 {} 个被阻止的 IP（无 ping，无开放端口）。重新扫描全部还是仅扫描其余？",
        "target_operators": "运营商",
        "operators_summary": "{}: {}",
        "white_hdr": "白名单",
        "white_yes": "是",
        "white_no": "否",
        "operator_hdr": "运营商",
        "reset_data_btn": "重置数据 🔄",
        "reset_data_confirm": "删除所有配置文件（设置、上次扫描、缓存、运营商数据）并重启应用？",
        "reset_data_done": "已删除。正在重启...",
        "range_source_label": "范围来源：",
        "range_source_none": "—",
        "fetch_status": "{} 个范围来自 {} ✅",
        "source_cf_api": "Cloudflare API", "source_cf_asn": "Cloudflare ASN", "source_cf_github": "Cloudflare GitHub",
        "source_fastly_api": "Fastly API", "source_fastly_asn": "Fastly ASN", "source_all": "All Sources",
        "status_fetch": "获取范围... 🔍",
        "status_analyze": "分析 {} 个范围... 🤖",
        "status_scan": "扫描 {} 个范围... ⚡",
        "status_pct": "扫描... {}% ⚡",
        "made_by_simple": "编程与设计：shahin salek tootoonchi",
        "settings_main": "主要设置",
        "settings_theme": "主题 / 颜色",
        "settings_ip_check": "IP 检查方式",
        "ip_check_my_internet": "使用我的网络检查 IP",
        "ip_check_agent": "通过代理搜索",
        "ip_check_agent_message": "此功能将在后续更新中添加。目前您只能使用自己的网络进行搜索。",
        "settings_saved": "设置已保存。",
        "rescan_closed_title": "之前扫描过的 IP",
        "rescan_closed_message": "有 {} 个 IP 之前扫描过且没有开放端口。是否要重新扫描它们？",
        "rescan_yes": "是，重新扫描",
        "rescan_no": "否，跳过",
        "ping_unsuitable": "不合适",
        "update_btn": "更新",
        "update_available_title": "有更新",
        "update_available_message": "有新版本可用。是否立即更新？",
        "update_confirm_question": "是否要更新到最新版本？",
        "update_downloading": "正在下载更新...",
        "update_done": "更新完成。",
        "update_restart_question": "是否重启以使用新版本？",
        "update_restart_yes": "是，重启",
        "update_restart_no": "否",
        "update_no_new": "您已是最新版本。",
        "update_checking": "正在检查更新...",
        "update_error": "更新错误",
        "operator_country": "运营商国家:",
        "country_iran": "伊朗",
        "country_russia": "俄罗斯",
        "country_china": "中国",
    },
    "ru": {
        "app_title": "CDN IP Сканер V1.6",
        "app_subtitle": "Точность • Скорость • ИИ ⚡",
        "loading": "Загрузка...",
        "choose_language": "Выберите язык",
        "target": "Цель",
        "speed": "Скорость",
        "found": "Найдено",
        "time": "Время",
        "settings": "Настройки ⚙️",
        "mode": "Режим 🔥",
        "ai": "AI 🤖",
        "count": "Кол-во 🎯",
        "fetch_ranges_btn": "Получить диапазоны 📡",
        "analyze_btn": "Анализ ИИ 🤖",
        "save_btn": "Сохранить 💾",
        "stop_btn": "Стоп ⏹️",
        "start_btn": "Сканировать 🚀",
        "progress_title": "Прогресс 📊",
        "ready_status": "Готов к запуску ⚡",
        "results_title": "Результаты (по скорости) 🎯",
        "scan_status_log": "Статус сканирования (каждый IP):",
        "score_hdr": "Оценка ⭐",
        "ports_hdr": "Порты 🔓",
        "ping_hdr": "Ping ⚡",
        "ip_hdr": "IP 📍",
        "rank_hdr": "Ранг 🏆",
        "made_by": "Сделано с ❤️ {}",
        "author_display": "shahin salek tootoonchi",
        "website_link_text": "Купить сервер",
        "donate_text": "💎 Donate (USDT): 0xde200d...c269",
        "github_text": "GitHub ⭐",
        "youtube_text": "YouTube 📺",
        "donate_title": "Поддержать разработку 💎",
        "donate_desc": "Если программа полезна, можно поддержать донатом:",
        "copy_btn": "Копировать адрес 📋",
        "close_btn": "Закрыть",
        "thanks": "🙏 Спасибо за поддержку!",
        "range_fetcher_title": "Получить диапазоны IP 📡",
        "range_sources_title": "Получить диапазоны IP из источников 📡",
        "ranges_received": "Полученные диапазоны 📋:",
        "ready_fetch": "Готов к получению...",
        "fetching": "Получение...",
        "error_fetch": "Ошибка получения диапазонов!",
        "cf_api": "Официальный API - Cloudflare ☁️",
        "cf_asn": "ASN - Cloudflare 🔢",
        "cf_github": "GitHub - Cloudflare 🐙",
        "fastly_api": "Официальный API - Fastly ⚡",
        "fastly_asn": "ASN - Fastly 🔢",
        "all_sources": "Все источники (рекомендуется) 🌐",
        "export_menu_title": "Формат экспорта 💾",
        "excel_btn": "Excel (xlsx) 📊",
        "json_btn": "JSON (json) 📄",
        "txt_btn": "TXT (только IP, по одному на строку) 📄",
        "settings_ping_range": "Фильтр диапазона Ping (мс)",
        "ping_min_label": "Мин. Ping (мс):",
        "ping_max_label": "Макс. Ping (мс):",
        "ping_filter_display": "Фильтр Ping: {}–{} мс",
        "ping_filter_none": "Без фильтра Ping",
        "settings_ports": "Порты сканирования",
        "ports_label": "Порты (через запятую, напр. 443,80,2010):",
        "ports_hint": "По умолчанию: 443,80,8443,2053,2083,2087,2096",
        "ranges_hint_paste_fetch": "Вставьте IP или загрузите диапазоны. Один IP или CIDR на строку.",
        "add_range_manual": "Добавить диапазон / IP вручную:",
        "add_range_btn": "Добавить",
        "box_range_and_scan": "Выбор диапазона и способа сканирования",
        "scan_method_label": "Способ сканирования",
        "scan_method_cloud": "Облачное сканирование",
        "scan_method_operators": "С диапазонами IP операторов",
        "scan_method_v2ray": "С пользовательской конфигурацией в V2rayN",
        "scan_method_cloud_tooltip": "Этим способом проверяются только те IP, которые вы вводите вручную или выбираете через «Получить диапазоны», с использованием вашего локального интернета.",
        "coming_soon_scan": "В разработке; скоро будет добавлено.",
        "operator_fetch_btn": "Fetch",
        "fetch_all_operators_btn": "Fetch all operator IPs 📡",
        "filter_only_clean_label": "Results: only IPs with open port + ping",
        "filter_show_all_label": "Results: show all (including closed)",
        "scan_from_operator": "Scan from operator:",
        "scan_from_operator_all": "All",
        "operator_scan_basis": "Like cf-ip-scanner: each IP is tested with HTTP/HTTPS GET to /cdn-cgi/trace (Cloudflare). Several attempts per port; only if enough succeed and latency is under max, the IP/port is marked active. Scan runs from your current connection (one operator at a time). To see status on all operators, run a separate scan from each.",
        "operator_ranges_count": "{} ranges",
        "operator_fetch_error": "Fetch error",
        "operator_fetching": "Fetching...",
        "invalid_ip_cidr": "Неверный IP или CIDR",
        "msg_no_results": "Нет результатов для сохранения!",
        "msg_saved": "Файл сохранён!",
        "msg_error": "Ошибка",
        "msg_copied": "Скопировано!",
        "ip_copied": "IP скопирован: {}",
        "error_title": "Ошибка",
        "success_title": "Успех",
        "analyze_title": "Анализ ИИ",
        "no_results_analyze": "Нет результатов для анализа!",
        "openpyxl_missing": "openpyxl не установлен. Выполните: pip install openpyxl",
        "file_saved": "Файл сохранён!\n{}",
        "save_error": "Ошибка сохранения:\n{}",
        "scan_done": "Готово! {} IP за {}s ✅",
        "mode_turbo": "Turbo (10s)", "mode_hyper": "Hyper (5s)", "mode_ultra": "Ultra (15s)", "mode_deep": "Deep (30s)",
        "mode_turbo_desc": "700 processes, 30 IPs per range",
        "mode_hyper_desc": "600 processes, 50 IPs per range",
        "mode_ultra_desc": "500 processes, 70 IPs per range",
        "mode_deep_desc": "450 processes, 100 IPs per range",
        "ai_basic": "Basic", "ai_smart": "Smart", "ai_advanced": "Advanced", "ai_expert": "Expert",
        "target_50": "50", "target_100": "100", "target_150": "150", "target_200": "200", "target_500": "500", "target_1000": "1000", "target_all": "All",
        "count_desc": "Количество IP для сканирования (точный лимит).",
        "coming_future_updates": "Будет добавлено в следующих обновлениях.",
        "range_source_label": "Источник диапазонов:",
        "range_source_none": "—",
        "fetch_status": "{} диапазонов из {} ✅",
        "source_cf_api": "Cloudflare API", "source_cf_asn": "Cloudflare ASN", "source_cf_github": "Cloudflare GitHub",
        "source_fastly_api": "Fastly API", "source_fastly_asn": "Fastly ASN", "source_all": "All Sources",
        "status_fetch": "Получение диапазонов... 🔍",
        "status_analyze": "Анализ {} диапазонов... 🤖",
        "status_scan": "Сканирование {} диапазонов... ⚡",
        "status_pct": "Сканирование... {}% ⚡",
        "made_by_simple": "Программирование и дизайн: shahin salek tootoonchi",
        "settings_main": "Основные настройки",
        "settings_theme": "Тема / Цвет",
        "settings_ip_check": "Способ проверки IP",
        "ip_check_my_internet": "Проверять IP через мой интернет",
        "ip_check_agent": "Поиск через агент",
        "ip_check_agent_message": "Эта функция будет добавлена в следующих обновлениях. Пока вы можете искать только через свой интернет.",
        "settings_saved": "Настройки сохранены.",
        "rescan_closed_title": "Ранее сканированные IP",
        "rescan_closed_message": "У вас {} IP ранее сканировались и не имели открытых портов. Сканировать их снова?",
        "rescan_yes": "Да, сканировать",
        "rescan_no": "Нет, пропустить",
        "ping_unsuitable": "Непригодно",
        "update_btn": "Обновить",
        "update_available_title": "Доступно обновление",
        "update_available_message": "Доступна новая версия. Обновить сейчас?",
        "update_confirm_question": "Обновить до последней версии?",
        "update_downloading": "Загрузка обновления...",
        "update_done": "Обновление завершено.",
        "update_restart_question": "Перезапустить для использования новой версии?",
        "update_restart_yes": "Да, перезапустить",
        "update_restart_no": "Нет",
        "update_no_new": "У вас последняя версия.",
        "update_checking": "Проверка обновлений...",
        "update_error": "Ошибка обновления",
        "operator_country": "Страна оператора:",
        "country_iran": "Иран",
        "country_russia": "Россия",
        "country_china": "Китай",
        "select_range_first": "Сначала выберите диапазон из активных, затем начните сканирование.",
        "ranges_to_scan_label": "Диапазоны для сканирования (отметьте снять):",
        "no_ranges_loaded": "Диапазоны не загружены. Сначала получите их из Cloudflare или Fastly API.",
        "rescan_blocked_message": "В прошлый раз у вас было {} заблокированных IP (без пинга, без открытых портов). Сканировать все снова или только остальные?",
        "target_operators": "Операторы",
        "operators_summary": "{}: {}",
        "white_hdr": "Белый",
        "white_yes": "Да",
        "white_no": "Нет",
        "operator_hdr": "Оператор",
        "reset_data_btn": "Сброс данных 🔄",
        "reset_data_confirm": "Удалить все конфиги (настройки, последнее сканирование, кэш, данные операторов) и перезапустить приложение?",
        "reset_data_done": "Данные удалены. Перезапуск...",
    },
}

def _persian_digit(s):
    """Convert 0-9 to Persian digits ۰-۹ in string s."""
    if not isinstance(s, str):
        s = str(s)
    table = str.maketrans("0123456789", "۰۱۲۳۴۵۶۷۸۹")
    return s.translate(table)

# ===== UI constants — استانداردهای ظاهری مدرن =====
RADIUS_BOX = 8
CARD_PADDING_X = 14
CARD_PADDING_Y = 10
SECTION_PADDING = 10
MIN_STAT_CARD_WIDTH = 110

class RoundedFrame(tk.Canvas):
    """Frame with rounded corners; use only where size is fixed (e.g. donate box)."""
    def __init__(self, parent, radius=None, bg=None, **kwargs):
        self.radius = radius if radius is not None else RADIUS_BOX
        self._bg = bg or parent.cget("bg")
        super().__init__(parent, highlightthickness=0, bg=parent.cget("bg"), **kwargs)
        self.inner = tk.Frame(self, bg=self._bg)
        self._win_id = None
        self.bind("<Configure>", self._on_configure)
    
    def _on_configure(self, event):
        w, h = event.width, event.height
        r = self.radius
        self.delete("all")
        bg = self._bg
        self.create_arc(0, 0, 2*r, 2*r, start=90, extent=90, fill=bg, outline=bg)
        self.create_arc(w-2*r, 0, w, 2*r, start=0, extent=90, fill=bg, outline=bg)
        self.create_arc(w-2*r, h-2*r, w, h, start=270, extent=90, fill=bg, outline=bg)
        self.create_arc(0, h-2*r, 2*r, h, start=180, extent=90, fill=bg, outline=bg)
        self.create_rectangle(r, 0, w-r, h, fill=bg, outline=bg)
        self.create_rectangle(0, r, w, h-r, fill=bg, outline=bg)
        if self._win_id:
            self.delete(self._win_id)
        self._win_id = self.create_window(r, r, window=self.inner, anchor="nw", width=max(1, w-2*r), height=max(1, h-2*r))

# ===== AI Optimizer =====
class AIOptimizer:
    def __init__(self):
        self.successful_patterns = {
            'cloudflare': ['104.16.0.0/13', '104.24.0.0/14', '162.159.0.0/19', '172.64.0.0/13'],
            'fastly': ['151.101.0.0/16', '199.232.0.0/16']
        }
        self.priority_ports = [443, 80, 8443, 2053, 2083, 2087, 2096]
    
    def predict_best_ranges(self, all_ranges, limit=100):
        scored = [(r, self._score(r)) for r in all_ranges]
        scored.sort(key=lambda x: x[1], reverse=True)
        return [r[0] for r in scored[:limit]]
    
    def _score(self, cidr):
        score = random.uniform(5, 15)
        try:
            net = ipaddress.IPv4Network(cidr, strict=False)
            score += 5.0 / math.log10(max(net.num_addresses, 10))
        except Exception:
            pass
        return score
    
    def smart_sample(self, cidr, max_ips=50):
        try:
            net = ipaddress.IPv4Network(cidr, strict=False)
            hosts = list(net.hosts())
            if len(hosts) <= max_ips:
                return hosts
            selected = [hosts[0], hosts[-1], hosts[len(hosts)//2]]
            step = len(hosts) // (max_ips - 10)
            for i in range(0, len(hosts), max(step, 1)):
                if len(selected) >= max_ips - 7:
                    break
                selected.append(hosts[i])
            remaining = max_ips - len(selected)
            if remaining > 0:
                selected.extend(random.sample(hosts, min(remaining, len(hosts))))
            return list(set(selected))[:max_ips]
        except Exception:
            return []

# ===== اسکنر مشابه cf-ip-scanner (JS) =====
# اساس اسکن (مثل اسکریپت cf-ip-scanner):
# - درخواست به IP مستقیم باید با Host: دامنهٔ کلودفلیر باشد تا سرور edge پاسخ /cdn-cgi/trace بدهد.
# - اگر /cdn-cgi/trace جواب نداد، اتصال TCP به پورت به‌عنوان fallback (برای هر سروری که پورت باز دارد).
CDN_TRACE_PATH = "/cdn-cgi/trace"
HTTPS_PORTS = {443, 8443, 2053, 2083, 2087, 2096}
HTTP_PORTS = {80, 8080, 2052, 2082, 2086, 2095}
CDN_CHECK_ATTEMPTS = 5
CDN_CHECK_MIN_SUCCESS = 1
CDN_HOST_HEADER = "www.cloudflare.com"


def _cdn_trace_url(ip_str, port):
    """بر اساس پورت پروتکل و URL برای /cdn-cgi/trace."""
    scheme = "https" if port in HTTPS_PORTS else "http"
    return f"{scheme}://{ip_str}:{port}{CDN_TRACE_PATH}"


class UltraScanner:
    def __init__(self):
        self.ai = AIOptimizer()
        self.max_workers = 550
        self.timeout = 2.0
        self.max_latency_ms = 9999
        self.failed_cache = set()
        self.update_queue = None
    
    def _http_check_once(self, ip_str, port, timeout_sec):
        """یک بار GET به /cdn-cgi/trace با Host کلودفلیر؛ برگردان (موفق: bool، تأخیر ms)."""
        url = _cdn_trace_url(ip_str, port)
        headers = {"Host": CDN_HOST_HEADER, "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
        start = time.time()
        try:
            r = requests.get(url, timeout=timeout_sec, verify=False, headers=headers)
            elapsed_ms = (time.time() - start) * 1000
            return (r.status_code == 200, elapsed_ms)
        except Exception:
            return (False, (time.time() - start) * 1000)
    
    def _tcp_connect(self, ip_str, port, timeout_sec):
        """اتصال TCP به پورت؛ اگر باز باشد برگردان (True، تأخیر ms)، وگرنه (False، ۰)."""
        start = time.time()
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(timeout_sec)
            sock.connect((ip_str, port))
            sock.close()
            return (True, (time.time() - start) * 1000)
        except Exception:
            return (False, 0)
    
    def _check_port_parallel(self, ip_str, port, timeout_sec):
        """چند بار GET به یک پورت به‌صورت موازی؛ برگردان (تعداد موفق، لیست تأخیرها)."""
        successes = 0
        latencies = []
        with ThreadPoolExecutor(max_workers=CDN_CHECK_ATTEMPTS) as ex:
            futures = [ex.submit(self._http_check_once, ip_str, port, timeout_sec) for _ in range(CDN_CHECK_ATTEMPTS)]
            for f in as_completed(futures, timeout=timeout_sec * 2):
                try:
                    ok, lat = f.result()
                    if ok:
                        successes += 1
                        latencies.append(lat)
                except Exception:
                    pass
        return successes, latencies
    
    def check(self, ip, ports):
        """
        برای هر پورت: اول GET به /cdn-cgi/trace (با Host کلودفلیر). اگر جواب نداد، اتصال TCP به پورت (fallback).
        پورت «باز» = حداقل یک پاسخ HTTP موفق یا اتصال TCP موفق، با تأخیر زیر max_latency_ms.
        """
        ip_str = str(ip)
        if ip_str in self.failed_cache:
            if self.update_queue:
                self.update_queue.put(("ip_status", (ip_str, False, None, [])))
            return None
        
        result = {'ip': ip_str, 'open_ports': [], 'ping': None}
        timeout_sec = min(10, max(2.0, self.max_latency_ms / 1000.0 * 1.5))
        
        try:
            for port in ports:
                successes, latencies = self._check_port_parallel(ip_str, port, timeout_sec)
                if successes >= CDN_CHECK_MIN_SUCCESS and latencies:
                    avg_lat = sum(latencies) / len(latencies)
                    if avg_lat <= self.max_latency_ms:
                        result['open_ports'].append(port)
                        if result['ping'] is None or avg_lat < result['ping']:
                            result['ping'] = avg_lat
                else:
                    ok_tcp, lat_tcp = self._tcp_connect(ip_str, port, min(5, timeout_sec))
                    if ok_tcp and lat_tcp <= self.max_latency_ms:
                        result['open_ports'].append(port)
                        if result['ping'] is None or lat_tcp < result['ping']:
                            result['ping'] = lat_tcp
            
            if result['open_ports']:
                if self.update_queue:
                    self.update_queue.put(("ip_status", (ip_str, True, result['ping'], result['open_ports'])))
                return result
            self.failed_cache.add(ip_str)
            if self.update_queue:
                self.update_queue.put(("ip_status", (ip_str, False, None, [])))
            return None
        except Exception:
            if self.update_queue:
                self.update_queue.put(("ip_status", (ip_str, False, None, [])))
            return None
    
    def batch_scan(self, ips, ports):
        results = []
        n = len(ips)
        # برای یک یا دو آی‌پی تایم‌اوت طولانی‌تر تا خطای «futures unfinished» ندهد
        timeout_sec = max(120, self.timeout * max(n, 1) + 60) if n <= 2 else max(60, self.timeout * n // 10 + 30)
        with ThreadPoolExecutor(max_workers=self.max_workers) as ex:
            futures = {ex.submit(self.check, ip, ports): ip for ip in ips}
            try:
                for future in as_completed(futures, timeout=timeout_sec):
                    try:
                        result = future.result()
                        if result:
                            results.append(result)
                    except Exception:
                        pass
            except Exception:
                for f in futures:
                    try:
                        f.cancel()
                    except Exception:
                        pass
        return results

# ===== Operators by country — ایران، چین، روسیه =====
OPERATORS_IRAN = {
    "irancell": {"name": "ایرانسل", "asn": 44244},
    "mci": {"name": "همراه اول", "asn": 197207},
    "rightel": {"name": "رایتل", "asn": 57218},
    "shuttle": {"name": "شاتل", "asn": 12880},
}

# چین — ASN از BGPView (مستندات اپراتورهای چین)
OPERATORS_CHINA = {
    "china_mobile": {"name": "China Mobile (中国移动)", "asn": [9808, 56040, 56041, 56042, 56044, 56046, 56047, 56048, 58453, 58807, 9231]},
    "china_unicom": {"name": "China Unicom (中国联通)", "asn": [4837, 4808, 9800, 9929, 10099, 17621, 17623, 17816]},
    "china_telecom": {"name": "China Telecom (中国电信)", "asn": [4134, 4809, 4812, 23724, 58466, 133774, 133775, 136958]},
    "china_broadnet": {"name": "China Broadnet (中国广电)", "asn": [58542]},
}

# روسیه — ASN از BGPView (مستندات اپراتورهای روسیه)
OPERATORS_RUSSIA = {
    "mts": {"name": "MTS (МТС)", "asn": [8359, 13174, 28840, 29226, 31163, 35807, 42511, 50544]},
    "megafon": {"name": "MegaFon (МегаФон)", "asn": [25159, 31133, 31200, 43478, 44843, 49037, 50716]},
    "beeline": {"name": "Beeline (Билайн)", "asn": [3216, 8402, 12389, 21453, 28917, 35000, 41733, 42668, 48642]},
    "tele2": {"name": "Tele2 Russia (Теле2)", "asn": [12958, 35104, 41668, 44746, 48287, 48715, 50384, 51547]},
}

OPERATORS_BY_COUNTRY = {"ir": OPERATORS_IRAN, "cn": OPERATORS_CHINA, "ru": OPERATORS_RUSSIA}

# هدر مرورگر برای جلوگیری از بلاک توسط APIها
OPERATOR_FETCH_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "application/json",
    "Accept-Language": "en-US,en;q=0.9",
}

def _normalize_ipv4_prefixes(prefixes):
    """Extract IPv4 prefix strings from list of dicts or strings. Filter out IPv6."""
    out = []
    for p in prefixes or []:
        if isinstance(p, dict):
            s = p.get("prefix") or p.get("ip") or ""
        else:
            s = str(p).strip()
        if not s or ":" in s:
            continue
        if "/" not in s:
            continue
        out.append(s)
    return out


def fetch_operator_prefixes_bgpview(asn):
    """Fetch IPv4 prefixes from BGPView API. Returns (count, prefix_list, error_msg or None)."""
    try:
        r = requests.get(
            f"https://api.bgpview.io/asn/{asn}/prefixes",
            headers=OPERATOR_FETCH_HEADERS,
            timeout=25
        )
        r.raise_for_status()
        data = r.json()
        if data.get("status") != "ok":
            return 0, [], data.get("status_message", "API error")
        payload = data.get("data") or {}
        ipv4 = payload.get("ipv4_prefixes")
        if ipv4 is None and isinstance(payload.get("prefixes"), dict):
            ipv4 = payload["prefixes"].get("ipv4") or payload["prefixes"].get("ipv4_prefixes") or []
        if ipv4 is None:
            ipv4 = []
        if not isinstance(ipv4, list):
            ipv4 = []
        prefixes = _normalize_ipv4_prefixes(ipv4)
        return len(prefixes), prefixes, None
    except requests.RequestException as e:
        return 0, [], getattr(e, "message", None) or str(e)
    except Exception as e:
        return 0, [], str(e)


def fetch_operator_prefixes_ripe(asn):
    """Fetch IPv4 prefixes from RIPE Stat API. Returns (count, prefix_list, error_msg or None)."""
    try:
        r = requests.get(
            f"https://stat.ripe.net/data/announced-prefixes/data.json?resource=AS{asn}",
            headers=OPERATOR_FETCH_HEADERS,
            timeout=25
        )
        r.raise_for_status()
        data = r.json()
        payload = data.get("data") or {}
        raw = payload.get("prefixes") or []
        prefixes = _normalize_ipv4_prefixes(raw)
        if not prefixes:
            return 0, [], "no IPv4 prefixes"
        return len(prefixes), prefixes, None
    except requests.RequestException as e:
        return 0, [], getattr(e, "message", None) or str(e)
    except Exception as e:
        return 0, [], str(e)


def fetch_operator_prefixes_github_fallback(asn):
    """
    Fallback: try GitHub raw files for Iran IP ranges (all operators in one file).
    Maps ASN to known repo paths. Returns (count, prefix_list, error_msg or None).
    """
    # لیست فایل‌های شناخته‌شده برای رنج ایران (ممکن است کل ایران باشد نه هر ASN جدا)
    urls = [
        "https://raw.githubusercontent.com/arastu/iran_ip_ranges/main/ip_ranges.json",
        "https://raw.githubusercontent.com/arastu/iran_ip_ranges/main/ip_ranges.txt",
    ]
    for url in urls:
        try:
            r = requests.get(url, headers=OPERATOR_FETCH_HEADERS, timeout=15)
            if r.status_code != 200:
                continue
            text = r.text.strip()
            prefixes = []
            if url.endswith(".json"):
                try:
                    j = r.json()
                    if isinstance(j, list):
                        prefixes = _normalize_ipv4_prefixes([p if isinstance(p, dict) else {"prefix": p} for p in j])
                    elif isinstance(j, dict):
                        for v in j.values():
                            if isinstance(v, list):
                                prefixes.extend(_normalize_ipv4_prefixes(v))
                    if prefixes:
                        return len(prefixes), prefixes[:500], None
                except Exception:
                    pass
            else:
                for line in text.splitlines():
                    line = line.strip()
                    if line and not line.startswith("#") and "/" in line and ":" not in line:
                        prefixes.append(line)
                if prefixes:
                    return len(prefixes), prefixes[:500], None
        except Exception:
            continue
    return 0, [], "GitHub fallback failed"


def fetch_operator_prefixes_smart(asn):
    """
    دریافت هوشمند رنج آی‌پی اپراتور از چند منبع.
    اول RIPE Stat (معمولاً در ایران در دسترس)، بعد BGPView، در صورت نیاز GitHub.
    Returns (count, prefix_list, error_msg or None).
    """
    # ۱) RIPE Stat — معمولاً از ایران در دسترس و پایدار است
    cnt, prefixes, err = fetch_operator_prefixes_ripe(asn)
    if not err and cnt > 0:
        return cnt, prefixes, None
    # ۲) BGPView
    time.sleep(0.3)
    cnt, prefixes, err = fetch_operator_prefixes_bgpview(asn)
    if not err and cnt > 0:
        return cnt, prefixes, None
    # ۳) تلاش مجدد RIPE Stat
    time.sleep(0.5)
    cnt, prefixes, err = fetch_operator_prefixes_ripe(asn)
    if not err and cnt > 0:
        return cnt, prefixes, None
    # ۴) تلاش مجدد BGPView
    time.sleep(0.5)
    cnt, prefixes, err = fetch_operator_prefixes_bgpview(asn)
    if not err and cnt > 0:
        return cnt, prefixes, None
    # ۵) fallback از GitHub (فقط اپراتورهای اصلی)
    if asn in (197207, 44244, 57218):
        cnt, prefixes, err = fetch_operator_prefixes_github_fallback(asn)
        if not err and cnt > 0:
            return cnt, prefixes, None
    return 0, [], err or "no data"

# ===== Range Fetcher =====
class RangeFetcher:
    @staticmethod
    def get_cloudflare_official():
        try:
            r = requests.get('https://api.cloudflare.com/client/v4/ips', timeout=10)
            return r.json()['result']['ipv4_cidrs']
        except Exception:
            return []
    
    @staticmethod
    def get_cloudflare_asn():
        cloudflare_asn_ranges = [
            '104.16.0.0/12', '172.64.0.0/13', '162.159.0.0/16',
            '173.245.48.0/20', '103.21.244.0/22', '103.22.200.0/22',
            '103.31.4.0/22', '141.101.64.0/18', '108.162.192.0/18',
            '190.93.240.0/20', '188.114.96.0/20', '197.234.240.0/22',
            '198.41.128.0/17', '131.0.72.0/22'
        ]
        return cloudflare_asn_ranges
    
    @staticmethod
    def get_cloudflare_github():
        try:
            r = requests.get('https://raw.githubusercontent.com/cloudflare/cloudflare-docs/production/data/ip-ranges.json', timeout=10)
            data = r.json()
            return data.get('ipv4_cidrs', [])
        except Exception:
            return []
    
    @staticmethod
    def get_fastly_official():
        try:
            r = requests.get('https://api.fastly.com/public-ip-list', timeout=10)
            return r.json()['addresses']
        except Exception:
            return []
    
    @staticmethod
    def get_fastly_asn():
        fastly_asn_ranges = [
            '151.101.0.0/16', '199.232.0.0/16',
            '103.244.50.0/24', '103.245.222.0/23',
            '103.245.224.0/24', '104.156.80.0/20'
        ]
        return fastly_asn_ranges
    
    @staticmethod
    def get_all_sources():
        all_ranges = []
        all_ranges.extend(RangeFetcher.get_cloudflare_official())
        all_ranges.extend(RangeFetcher.get_cloudflare_asn())
        all_ranges.extend(RangeFetcher.get_cloudflare_github())
        all_ranges.extend(RangeFetcher.get_fastly_official())
        all_ranges.extend(RangeFetcher.get_fastly_asn())
        return list(set(all_ranges))

# ===== Theme Manager (loads from assets/theme.json) =====
def _load_theme_from_assets():
    """بارگذاری تم و layout از assets/theme.json؛ در صورت خطا از مقادیر پیش‌فرض استفاده می‌شود."""
    _path = os.path.join(RESOURCE_DIR, "assets", "theme.json")
    default_themes = {
        'dark': {
            'bg': '#0a0a0a', 'fg': '#ffffff', 'accent': '#ffffff', 'success': '#ffffff',
            'error': '#6b7280', 'warning': '#6b7280', 'card_bg': '#1a1a1a', 'card_fg': '#e5e5e5'
        },
        'light': {
            'bg': '#ffffff', 'fg': '#0f0f0f', 'accent': '#0f0f0f', 'success': '#0f0f0f',
            'error': '#374151', 'warning': '#6b7280', 'card_bg': '#f5f5f5', 'card_fg': '#525252'
        }
    }
    default_layout = {
        'button_min_width': 140, 'button_height': 40, 'button_padx': 20, 'button_pady': 10,
        'button_border': '#525252', 'button_border_radius': 5,
        'card_padding_x': 16, 'card_padding_y': 12, 'section_padding': 12,
        'border_highlight': '#525252',
        'font_size_title': 24, 'font_size_heading': 12, 'font_size_body': 10, 'font_size_small': 9
    }
    if not os.path.isfile(_path):
        return default_themes, default_layout
    try:
        with open(_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        themes = data.get("themes") or default_themes
        layout = data.get("layout") or default_layout
        for k, v in default_layout.items():
            if k not in layout:
                layout[k] = v
        for name in ("dark", "light"):
            if name not in themes:
                themes[name] = default_themes[name]
            for key in default_themes["dark"]:
                if key not in themes[name]:
                    themes[name][key] = default_themes[name].get(key, default_themes["dark"][key])
        return themes, layout
    except Exception:
        return default_themes, default_layout

class ThemeManager:
    _THEMES, _LAYOUT = _load_theme_from_assets()
    THEMES = _THEMES
    
    @staticmethod
    def get_theme(name='dark'):
        return ThemeManager.THEMES.get(name, ThemeManager.THEMES['dark'])
    
    @staticmethod
    def get_layout():
        return getattr(ThemeManager, '_LAYOUT', {
            'button_min_width': 140, 'button_height': 40, 'button_padx': 20, 'button_pady': 10,
            'button_border': '#525252', 'card_padding_x': 16, 'card_padding_y': 12, 'section_padding': 12,
            'border_highlight': '#525252'
        })

# ===== Excel Exporter =====
class ExcelExporter:
    @staticmethod
    def export(results, filename, app=None):
        """app: instance of CDNScannerPro for t() and num(); if None, use plain strings."""
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl not installed. Run: pip install openpyxl")
        
        def _t(key):
            return app.t(key) if app else {"ping_unsuitable": "Unsuitable"}.get(key, key)
        def _num(n):
            return app.num(n) if app else str(n)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Clean IPs"
        
        # Header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(name='VazirMatn', size=12, bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Border
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ['رتبه', 'آدرس IP', 'Ping (ms)', 'پورت‌های باز', 'امتیاز', 'تاریخ', 'زمان']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data
        data_font = Font(name='VazirMatn', size=11)
        data_alignment = Alignment(horizontal="right", vertical="center")
        
        sorted_results = sorted(results, key=lambda x: x.get('score', 0), reverse=True)
        
        now = datetime.now()
        date_str = now.strftime('%Y/%m/%d')
        time_str = now.strftime('%H:%M:%S')
        
        for idx, result in enumerate(sorted_results, 1):
            row = idx + 1
            ports_str = ', '.join(map(str, result['open_ports']))
            
            ping_val = result.get('ping')
            if ping_val is None:
                ping_cell = 'N/A'
            elif ping_val >= 1000:
                ping_cell = _t("ping_unsuitable")
            else:
                ping_cell = _num(int(round(ping_val))) + " ms"
            data = [
                f"#{idx}",
                result['ip'],
                ping_cell,
                ports_str,
                _num(int(result.get('score', 0))) + "/" + _num(100),
                date_str,
                time_str
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border
                
                # Color coding based on score
                if col == 5:  # Score column
                    score = result.get('score', 0)
                    if score >= 80:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif score >= 60:
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        
        # Right-to-left
        ws.sheet_view.rightToLeft = True
        
        # Save
        wb.save(filename)

# ===== Main GUI =====
class CDNScannerPro:
    def __init__(self, root):
        self.root = root
        self.root.title("CDN IP Scanner V1.6")
        
        # مسیر فونت‌ها: Vazirmatn یا در صورت وجود EV Sam (فونت ای وی سام)
        self._font_dir = os.path.join(RESOURCE_DIR, "fonts")
        self._font_regular = os.path.join(self._font_dir, "Vazirmatn-Regular.ttf")
        self._font_bold = os.path.join(self._font_dir, "Vazirmatn-Bold.ttf")
        self._font_evsam_r = os.path.join(self._font_dir, "EV-Sam-Regular.ttf")
        self._font_evsam_b = os.path.join(self._font_dir, "EV-Sam-Bold.ttf")
        if not os.path.isfile(self._font_evsam_r):
            self._font_evsam_r = os.path.join(self._font_dir, "EVSam-Regular.ttf")
            self._font_evsam_b = os.path.join(self._font_dir, "EVSam-Bold.ttf")
        self._font_cache = {}
        self._fa_solid_path = os.path.join(self._font_dir, "fa-solid-900.ttf")
        self._fa_brands_path = os.path.join(self._font_dir, "fa-brands-400.ttf")
        self._fa_font_cache = {}  # (style, size) -> Font
        _load_fonts_win(self._font_dir)
        
        # ریسپانسیو: اندازه بر اساس مانیتور
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        window_width = max(800, min(1600, int(screen_width * 0.92)))
        window_height = max(500, min(900, int(screen_height * 0.88)))
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        self.root.geometry(f"{window_width}x{window_height}+{x}+{y}")
        self.root.minsize(750, 480)
        self.root.resizable(True, True)
        
        # Variables
        self.scanner = UltraScanner()
        self.ai = AIOptimizer()
        self.fetcher = RangeFetcher()
        
        self.current_theme = 'dark'
        self.theme = ThemeManager.get_theme(self.current_theme)
        self.layout = ThemeManager.get_layout()
        global CARD_PADDING_X, CARD_PADDING_Y, SECTION_PADDING
        CARD_PADDING_X = self.layout.get('card_padding_x', 16)
        CARD_PADDING_Y = self.layout.get('card_padding_y', 12)
        SECTION_PADDING = self.layout.get('section_padding', 12)
        
        self.lang = 'en'  # default English
        self.is_rtl = False
        self._loading_timer = None
        
        self.is_scanning = False
        self.results = []
        self.total_scanned = 0
        self.total_found = 0
        self.update_queue = Queue()
        self.start_time = None
        self.current_range_source = None  # نام منبع رنج فعلی
        
        # تنظیمات: متغیرها قبل از load_config (برای ذخیره/بارگذاری)
        self._target_internal = [50, 100, 150, 200, 500, 1000, "All"]
        self._target_keys = ["target_50", "target_100", "target_150", "target_200", "target_500", "target_1000", "target_all"]
        mode_vals_en = ["Turbo (10s)", "Hyper (5s)", "Ultra (15s)", "Deep (30s)"]
        ai_vals_en = ["Basic", "Smart", "Advanced", "Expert"]
        target_vals_en = ["50", "100", "150", "200", "500", "1000", "All"]
        self.mode_var = tk.StringVar(value=mode_vals_en[0])
        self.ai_var = tk.StringVar(value=ai_vals_en[1])
        self.target_var = tk.StringVar(value=target_vals_en[1])
        self.ip_check_method = tk.StringVar(value="internet")  # internet | agent
        
        self.scanned_closed = set()  # آی‌پی‌هایی که قبلاً اسکن شده و پورت باز نداشتند (ذخیرهٔ دائمی)
        self.ping_min = 0
        self.ping_max = 9999
        self.scan_ports = [443, 80, 8443, 2053, 2083, 2087, 2096]
        self.scan_method_var = tk.StringVar()
        self.current_operator_key = None
        self.current_operator_country = "ir"  # برای انگلیسی: ir | cn | ru
        self.filter_only_clean_var = tk.BooleanVar(value=True)
        self._update_available = False
        self._load_config()
        self._load_scan_cache()
        if not getattr(self, 'scan_ports', None) or not self.scan_ports:
            self.scan_ports = [443, 80, 8443, 2053, 2083, 2087, 2096]
        self.ai.priority_ports = self.scan_ports
        
        self.show_loading()
        # چک بروزرسانی در بک‌گراند (یک بار بعد از اجرا)
        threading.Thread(target=self._background_check_update, daemon=True).start()
    
    def get_operators_for_current_context(self):
        """بر اساس زبان و (در انگلیسی) کشور انتخاب‌شده، دیکشنری اپراتورها را برمی‌گرداند."""
        if self.lang == "fa":
            return OPERATORS_IRAN
        if self.lang == "zh":
            return OPERATORS_CHINA
        if self.lang == "ru":
            return OPERATORS_RUSSIA
        if self.lang == "en":
            return OPERATORS_BY_COUNTRY.get(getattr(self, "current_operator_country", "ir"), OPERATORS_IRAN)
        return OPERATORS_IRAN
    
    def _version_file(self):
        """مسیر فایل version در پوشهٔ برنامه."""
        return os.path.join(RESOURCE_DIR, "version")
    
    def _get_local_version(self):
        """خواندن نسخهٔ محلی از فایل version."""
        path = self._version_file()
        if not os.path.isfile(path):
            return "0"
        try:
            with open(path, "r", encoding="utf-8") as f:
                return (f.read() or "").strip() or "0"
        except Exception:
            return "0"
    
    def _background_check_update(self):
        """در بک‌گراند از گیت‌هاب نسخه را چک می‌کند؛ اگر جدیدتر بود _update_available را True می‌کند."""
        try:
            r = requests.get(GITHUB_RAW_VERSION_URL, timeout=10)
            if r.status_code != 200:
                return
            remote = (r.text or "").strip() or "0"
            local = self._get_local_version()
            if remote != local:
                try:
                    def _ver_tuple(s):
                        return tuple(int(x) for x in re.split(r"[^0-9]+", (s or "0").strip()) if x and x.isdigit())
                    rt, lt = _ver_tuple(remote), _ver_tuple(local)
                    if rt > lt:
                        self._update_available = True
                except Exception:
                    self._update_available = True
        except Exception:
            pass
    
    def _check_update_available(self):
        """چک هم‌زمان بروزرسانی؛ True اگر نسخهٔ جدید موجود باشد."""
        try:
            r = requests.get(GITHUB_RAW_VERSION_URL, timeout=10)
            if r.status_code != 200:
                return False
            remote = (r.text or "").strip() or "0"
            local = self._get_local_version()
            if remote == local:
                return False
            try:
                def _ver_tuple(s):
                    return tuple(int(x) for x in re.split(r"[^0-9]+", (s or "0").strip()) if x and x.isdigit())
                rt, lt = _ver_tuple(remote), _ver_tuple(local)
                return rt > lt
            except Exception:
                return True
        except Exception:
            return False
    
    def _do_update(self):
        """دانلود آخرین نسخه از گیت‌هاب، کپی فایل‌ها، سپس پرسش ریستارت."""
        app_dir = BASE_DIR
        try:
            r = requests.get(GITHUB_ZIP_URL, timeout=60, stream=True)
            r.raise_for_status()
        except Exception as e:
            messagebox.showerror(self.t("update_error"), str(e))
            return
        try:
            tmp = tempfile.mkdtemp()
            zip_path = os.path.join(tmp, "repo.zip")
            with open(zip_path, "wb") as f:
                for chunk in r.iter_content(chunk_size=8192):
                    f.write(chunk)
            with zipfile.ZipFile(zip_path, "r") as z:
                z.extractall(tmp)
            # پوشهٔ استخراج شده معمولاً cdn-ip-scanner-main است
            extracted = os.path.join(tmp, "cdn-ip-scanner-main")
            if not os.path.isdir(extracted):
                for name in os.listdir(tmp):
                    if name != "repo.zip" and os.path.isdir(os.path.join(tmp, name)):
                        extracted = os.path.join(tmp, name)
                        break
            if not os.path.isdir(extracted):
                messagebox.showerror(self.t("update_error"), "Invalid archive structure.")
                shutil.rmtree(tmp, ignore_errors=True)
                return
            # فایل‌هایی که بروزرسانی می‌شوند (بدون بازنویسی config/cache)
            to_copy = ["ip_scanner_pro.py", "requirements.txt", "Scanner_Pro.bat", "version"]
            for name in to_copy:
                src = os.path.join(extracted, name)
                dst = os.path.join(app_dir, name)
                if os.path.isfile(src):
                    shutil.copy2(src, dst)
            for folder in ["fonts", "assets"]:
                src_dir = os.path.join(extracted, folder)
                dst_dir = os.path.join(app_dir, folder)
                if os.path.isdir(src_dir):
                    if os.path.isdir(dst_dir):
                        shutil.rmtree(dst_dir, ignore_errors=True)
                    shutil.copytree(src_dir, dst_dir)
            shutil.rmtree(tmp, ignore_errors=True)
        except Exception as e:
            messagebox.showerror(self.t("update_error"), str(e))
            return
        if messagebox.askyesno(self.t("update_done"), self.t("update_restart_question"), default=tk.YES):
            self._restart_app()
    
    def _restart_app(self):
        """ریستارت برنامه و اجرای نسخهٔ جدید."""
        app_dir = BASE_DIR
        if getattr(sys, "frozen", False):
            cmd, cwd = [sys.executable], app_dir
        else:
            script = os.path.join(app_dir, "ip_scanner_pro.py")
            cmd, cwd = [sys.executable, script], app_dir
        try:
            if sys.platform == "win32" and hasattr(subprocess, "CREATE_NEW_CONSOLE"):
                subprocess.Popen(cmd, cwd=cwd, creationflags=subprocess.CREATE_NEW_CONSOLE)
            else:
                subprocess.Popen(cmd, cwd=cwd)
        except Exception:
            subprocess.Popen(cmd, cwd=cwd)
        self.root.quit()
        self.root.destroy()
        sys.exit(0)
    
    def _on_update_click(self):
        """کلیک روی دکمهٔ بروزرسانی: پرسش، سپس دانلود و نصب و در پایان پرسش ریستارت."""
        if self._update_available:
            msg = self.t("update_available_message")
        else:
            if not self._check_update_available():
                messagebox.showinfo(self.t("update_btn"), self.t("update_no_new"))
                return
            msg = self.t("update_confirm_question")
        if not messagebox.askyesno(self.t("update_available_title"), msg, default=tk.YES):
            return
        self._update_available = False
        # پنجره باز می‌ماند؛ در همین thread دانلود می‌کنیم (می‌توان بعداً با progress در thread انجام داد)
        self.root.config(cursor="watch")
        self.root.update()
        try:
            self._do_update()
        finally:
            self.root.config(cursor="")
    
    def t(self, key):
        """Translation for current language."""
        return TRANSLATIONS.get(self.lang, TRANSLATIONS["en"]).get(key, TRANSLATIONS["en"].get(key, key))
    
    def num(self, n):
        """Number as string; Persian digits if Farsi."""
        s = str(n) if isinstance(n, (int, float)) else n
        return _persian_digit(s) if self.lang == "fa" else s
    
    def font(self, size, bold=False):
        """Font: EV Sam or Vazirmatn for Persian; Segoe UI for others (modern)."""
        key = (size, bold)
        if key not in self._font_cache:
            if self.lang == "fa":
                use_evsam = os.path.isfile(self._font_evsam_r)
                if sys.platform == "win32":
                    fam = "EV Sam" if use_evsam else "Vazirmatn"
                    self._font_cache[key] = tkfont.Font(root=self.root, family=fam, size=size, weight="bold" if bold else "normal")
                elif use_evsam and os.path.isfile(self._font_evsam_b if bold else self._font_evsam_r):
                    path = self._font_evsam_b if bold else self._font_evsam_r
                    self._font_cache[key] = tkfont.Font(root=self.root, file=path, size=size)
                elif os.path.isfile(self._font_bold if bold else self._font_regular):
                    path = self._font_bold if bold else self._font_regular
                    self._font_cache[key] = tkfont.Font(root=self.root, file=path, size=size)
                else:
                    self._font_cache[key] = tkfont.Font(family="Tahoma", size=size, weight="bold" if bold else "normal")
            else:
                fam = "Segoe UI" if sys.platform == "win32" else "Tahoma"
                self._font_cache[key] = tkfont.Font(family=fam, size=size, weight="bold" if bold else "normal")
        return self._font_cache[key]
    
    def _fa_font_solid(self, size=14):
        """فونت Font Awesome 6 Solid برای آیکن‌ها."""
        key = ("solid", size)
        if key in self._fa_font_cache:
            return self._fa_font_cache[key]
        if not os.path.isfile(self._fa_solid_path):
            return None
        try:
            f = tkfont.Font(root=self.root, family="Font Awesome 6 Free", size=size, weight="bold")
        except Exception:
            try:
                f = tkfont.Font(root=self.root, file=self._fa_solid_path, size=size)
            except Exception:
                f = None
        self._fa_font_cache[key] = f
        return f
    
    def _fa_font_brands(self, size=14):
        """فونت Font Awesome 6 Brands برای آیکن‌های برند."""
        key = ("brands", size)
        if key in self._fa_font_cache:
            return self._fa_font_cache[key]
        if not os.path.isfile(self._fa_brands_path):
            return None
        try:
            f = tkfont.Font(root=self.root, family="Font Awesome 6 Brands", size=size)
        except Exception:
            try:
                f = tkfont.Font(root=self.root, file=self._fa_brands_path, size=size)
            except Exception:
                f = None
        self._fa_font_cache[key] = f
        return f
    
    def fa_icon(self, name, size=14, use_brands=False):
        """برگرداندن (کاراکتر آیکن، فونت، رنگ) برای نمایش؛ در صورت نبود فونت FA از ایموجی استفاده می‌شود."""
        char = FA.get(name, "")
        color = FA_COLORS.get(name, self.theme.get("accent", "#6366f1"))
        if use_brands and name in ("github", "youtube"):
            font = self._fa_font_brands(size)
        else:
            font = self._fa_font_solid(size)
        if font is None:
            fallback = {"rocket": "🚀", "gear": "⚙️", "bolt": "⚡", "check": "✅", "clock": "⏱️", "bullseye": "🎯", "satellite": "📡", "robot": "🤖", "save": "💾", "stop": "⏹️", "play": "🚀", "chart": "📊", "star": "⭐", "lock_open": "🔓", "location": "📍", "trophy": "🏆", "gem": "💎", "moon": "🌙", "sun": "☀️", "github": "⭐", "youtube": "📺", "download": "⬇️"}
            return (fallback.get(name, ""), None, color)
        return (char, font, color)
    
    def get_target_internal(self):
        """Return internal target value (50, 100, 150, 200, 500, 1000, or 'All') from combo display."""
        v = self.target_var.get()
        for i, k in enumerate(self._target_keys):
            if self.t(k) == v:
                return self._target_internal[i]
        return 100
    
    _MODE_INTERNAL = {"mode_turbo": "Turbo (10s)", "mode_hyper": "Hyper (5s)", "mode_ultra": "Ultra (15s)", "mode_deep": "Deep (30s)"}
    
    def get_mode_internal(self):
        """Return internal mode string (e.g. 'Turbo (10s)') from translated combo value."""
        v = self.mode_var.get()
        for key in ("mode_turbo", "mode_hyper", "mode_ultra", "mode_deep"):
            if self.t(key) == v:
                return self._MODE_INTERNAL[key]
        return "Turbo (10s)"
    
    def _config_file(self):
        return os.path.join(BASE_DIR, "config.json")
    
    def _scan_cache_file(self):
        return os.path.join(BASE_DIR, "scan_cache.json")
    
    def _load_scan_cache(self):
        """بارگذاری آی‌پی‌های اسکن‌شدهٔ قبلی (بسته) از فایل."""
        path = self._scan_cache_file()
        if not os.path.isfile(path):
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            self.scanned_closed.update(data.get("closed_ips", []))
        except Exception:
            pass
    
    def _save_scan_cache(self):
        """ذخیرهٔ آی‌پی‌های بسته در فایل (حتی بعد از بستن برنامه و خاموشی)."""
        path = self._scan_cache_file()
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump({"closed_ips": list(self.scanned_closed)}, f, ensure_ascii=False)
        except Exception:
            pass
    
    def _last_scan_file(self):
        return os.path.join(BASE_DIR, "last_scan.json")
    
    def _save_last_scan(self):
        """ذخیرهٔ آخرین نتایج اسکن برای بازیابی بعد از بستن برنامه."""
        if not getattr(self, "results", None):
            return
        path = self._last_scan_file()
        try:
            operator_mode = self._has_operator_display()
            data = {
                "results": self.results,
                "total_scanned": getattr(self, "total_scanned", 0),
                "total_found": getattr(self, "total_found", 0),
                "operator_mode": operator_mode,
                "timestamp": datetime.now().isoformat(),
            }
            with open(path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception:
            pass
    
    def _get_all_config_paths(self):
        """لیست مسیر همهٔ فایل‌های کانفیگ/کش در مسیر برنامه."""
        paths = [
            os.path.join(BASE_DIR, "config.json"),
            self._last_scan_file(),
            self._scan_cache_file(),
            self._operator_matrix_file(),
        ]
        for d in (OPERATORS_IRAN, OPERATORS_CHINA, OPERATORS_RUSSIA):
            for op_key in d:
                paths.append(self._operator_ranges_file(op_key))
        return paths
    
    def _reset_all_data(self):
        """حذف تمام فایل‌های کانفیگ و راه‌اندازی مجدد برنامه."""
        if not messagebox.askyesno(self.t("error_title"), self.t("reset_data_confirm")):
            return
        for path in self._get_all_config_paths():
            try:
                if os.path.isfile(path):
                    os.remove(path)
            except Exception:
                pass
        self.scanned_closed.clear()
        messagebox.showinfo(self.t("success_title"), self.t("reset_data_done"))
        try:
            subprocess.Popen([sys.executable, os.path.join(BASE_DIR, "ip_scanner_pro.py") if not getattr(sys, "frozen", False) else sys.executable], cwd=BASE_DIR)
        except Exception:
            pass
        self.root.quit()
        self.root.destroy()
        os._exit(0)
    
    def _is_operator_scan_mode(self):
        """آیا حالت اسکن «با رنج آی پی اپراتورها» انتخاب شده؟ (فارسی=ایران، چینی=چین، روسی=روسیه، انگلیسی=کشور انتخاب‌شده)"""
        return self.scan_method_var.get() == self.t("scan_method_operators")
    
    def _has_operator_display(self):
        """آیا ستون اپراتور و پورت در نتایج نمایش داده شود؟ فقط وقتی «با رنج آی پی اپراتورها» انتخاب شده."""
        return self._is_operator_scan_mode()
    
    def _matrix_cell_active(self, cell):
        """از سلول ماتریس (bool یا dict) مقدار فعال بودن را بردار. سازگار با فرمت قدیم."""
        if cell is None:
            return None
        if isinstance(cell, dict):
            return cell.get("active", False)
        return bool(cell)
    
    def _matrix_cell_ports(self, cell):
        """از سلول ماتریس لیست پورت‌های باز را بردار. سازگار با فرمت قدیم."""
        if cell is None or not isinstance(cell, dict):
            return []
        return list(cell.get("ports", []))
    
    def _format_operator_column(self, result):
        """ستون اپراتور: اپراتور(ها) با ✅ (پینگ دارد) یا ❌ (پینگ ندارد). وقتی «همه» انتخاب شده همهٔ اپراتورها با وضعیت نشان داده می‌شوند."""
        ip = result.get("ip", "")
        matrix = self._load_operator_matrix()
        row = matrix.get(ip, {})
        current_op_name = result.get("operator") or ""
        show_only_key = getattr(self, 'current_operator_key', None)
        has_ping = bool(result.get("open_ports")) and result.get("ping") is not None
        operators = self.get_operators_for_current_context()
        parts = []
        for op_key, op_info in operators.items():
            if show_only_key is not None and op_key != show_only_key:
                continue
            name = op_info["name"]
            cell = row.get(op_key)
            active = self._matrix_cell_active(cell)
            if active is None:
                if current_op_name == name or (show_only_key and not cell):
                    active = has_ping
                else:
                    active = None
            if active is None:
                parts.append(name + "—")
            elif active:
                parts.append(name + "✅")
            else:
                parts.append(name + "❌")
        if parts:
            return "، ".join(parts)
        name = (operators.get(show_only_key, {}).get("name") if show_only_key else None) or current_op_name or ""
        if name:
            return name + ("✅" if has_ping else "❌")
        return "—"
    
    def _format_ports_column(self, result):
        """ستون پورت: فقط پورت‌های باز (با ✅). پورت‌های بسته نشان داده نمی‌شوند."""
        ip = result.get("ip", "")
        matrix = self._load_operator_matrix()
        row = matrix.get(ip, {})
        open_ports = result.get("open_ports") or []
        current_op_name = result.get("operator")
        show_only_key = getattr(self, 'current_operator_key', None)
        operators = self.get_operators_for_current_context()
        parts = []
        for op_key, op_info in operators.items():
            if show_only_key is not None and op_key != show_only_key:
                continue
            cell = row.get(op_key)
            if isinstance(cell, dict):
                ports_list = self._matrix_cell_ports(cell)
                open_only = [p for p in ports_list if p]
                if open_only:
                    parts.append(op_info["name"] + ": " + " ".join([f"{p}✅" for p in open_only]))
            elif current_op_name == op_info["name"] and open_ports:
                parts.append(op_info["name"] + ": " + " ".join([f"{p}✅" for p in open_ports]))
        if not parts and current_op_name and open_ports:
            return current_op_name + ": " + " ".join([f"{p}✅" for p in open_ports])
        if not parts and open_ports:
            return " ".join([f"{p}✅" for p in open_ports])
        return "؛ ".join(parts) if parts else "—"
    
    def _update_target_display(self):
        """در حالت اپراتورها: نمایش «اپراتورها» و تعداد رنج هر اپراتور؛ وگرنه مقدار هدف (۵۰، ۱۰۰، ...)."""
        if not hasattr(self, "target_label") or not self.target_label.winfo_exists():
            return
        if self._is_operator_scan_mode():
            operators = self.get_operators_for_current_context()
            parts = []
            for op_key, op_info in operators.items():
                c = self._load_operator_count(op_key)
                parts.append(self.t("operators_summary").format(op_info["name"], self.num(c or 0)))
            self.target_label.config(text=self.t("target_operators") + "\n" + "، ".join(parts))
        else:
            tval = self.get_target_internal()
            tdisp = self.num(tval) if tval != "All" else self.t("target_all")
            self.target_label.config(text=tdisp)
    
    def _rebuild_results_from_matrix(self):
        """پر کردن جدول نتایج از ماتریس: یک ردیف به‌ازای هر (IP، اپراتور) که روی آن اپراتور باز است (یا همه اگر فیلتر خاموش)."""
        if not hasattr(self, "tree") or not self.tree.winfo_exists():
            return
        for item in self.tree.get_children():
            self.tree.delete(item)
        matrix = self._load_operator_matrix()
        only_clean = getattr(self, "filter_only_clean_var", None) and getattr(self.filter_only_clean_var, "get", lambda: True)()
        operators = self.get_operators_for_current_context()
        rows = []
        for ip_str, row in matrix.items():
            for op_key, cell in row.items():
                if not isinstance(cell, dict):
                    continue
                if only_clean and not (cell.get("active") and cell.get("ports")):
                    continue
                op_name = operators.get(op_key, {}).get("name", op_key)
                open_ports = cell.get("ports") or []
                ping_val = cell.get("ping")
                mini = {'ip': ip_str, 'open_ports': open_ports, 'ping': ping_val, 'operator': op_name}
                score = self._calc_score(mini)
                score_str = self.num(int(score)) + "/" + self.num(100)
                ports_str = " ".join([f"{p}✅" for p in open_ports]) if open_ports else "—"
                if ping_val is None or ping_val >= 1000:
                    ping_str = "—" if ping_val is None else self.t("ping_unsuitable")
                else:
                    ping_str = self.num(int(round(ping_val))) + " ms"
                op_str = op_name + "✅"
                rows.append((score_str, ports_str, ping_str, ip_str, op_str, score))
        for idx, (score_str, ports_str, ping_str, ip_str, op_str, score) in enumerate(sorted(rows, key=lambda x: -x[5]), 1):
            self.tree.insert("", tk.END, values=(score_str, ports_str, ping_str, ip_str, op_str, "#" + self.num(idx)))
        self.root.update_idletasks()
    
    def _load_config(self):
        """بارگذاری تنظیمات ذخیره‌شده از فایل."""
        path = self._config_file()
        if not os.path.isfile(path):
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                cfg = json.load(f)
            if cfg.get("theme") in ("dark", "light"):
                self.current_theme = cfg["theme"]
                self.theme = ThemeManager.get_theme(self.current_theme)
            mode_key = cfg.get("mode_key", "mode_turbo")
            ai_key = cfg.get("ai_key", "ai_smart")
            target_key = cfg.get("target_key", "target_100")
            self._saved_mode_key = mode_key
            self._saved_ai_key = ai_key
            self._saved_target_key = target_key
            self.ip_check_method.set(cfg.get("ip_check", "internet"))
            if "ping_min" in cfg and isinstance(cfg["ping_min"], (int, float)):
                self.ping_min = int(cfg["ping_min"])
            if "ping_max" in cfg and isinstance(cfg["ping_max"], (int, float)):
                self.ping_max = int(cfg["ping_max"])
            if "scan_ports" in cfg:
                if isinstance(cfg["scan_ports"], list):
                    self.scan_ports = [int(p) for p in cfg["scan_ports"] if str(p).strip().isdigit()]
                elif isinstance(cfg["scan_ports"], str) and cfg["scan_ports"].strip():
                    self.scan_ports = [int(p) for p in cfg["scan_ports"].replace(",", " ").split() if str(p).strip().isdigit()]
                if self.scan_ports:
                    self.ai.priority_ports = self.scan_ports
        except Exception:
            pass
    
    def _apply_loaded_config(self):
        """اعمال تنظیمات بارگذاری‌شده روی comboها (بعد از انتخاب زبان)."""
        if hasattr(self, "_saved_mode_key"):
            self.mode_var.set(self.t(self._saved_mode_key))
        if hasattr(self, "_saved_ai_key"):
            self.ai_var.set(self.t(self._saved_ai_key))
        if hasattr(self, "_saved_target_key"):
            self.target_var.set(self.t(self._saved_target_key))
    
    def _save_config(self, mode_display=None, ai_display=None, target_display=None):
        """ذخیره تنظیمات در فایل. اگر مقدار نمایشی داده شود از آن استفاده می‌شود (مثلاً از combo)."""
        mode_val = (mode_display or self.mode_var.get() or "").strip()
        ai_val = (ai_display or self.ai_var.get() or "").strip()
        target_val = (target_display or self.target_var.get() or "").strip()
        
        mode_key = None
        for k in ("mode_turbo", "mode_hyper", "mode_ultra", "mode_deep"):
            if (self.t(k) or "").strip() == mode_val:
                mode_key = k
                break
        ai_key = None
        for k in ("ai_basic", "ai_smart", "ai_advanced", "ai_expert"):
            if (self.t(k) or "").strip() == ai_val:
                ai_key = k
                break
        target_key = None
        for k in self._target_keys:
            if (self.t(k) or "").strip() == target_val:
                target_key = k
                break
        
        cfg = {
            "theme": self.current_theme,
            "mode_key": mode_key or "mode_turbo",
            "ai_key": ai_key or "ai_smart",
            "target_key": target_key or "target_100",
            "ip_check": self.ip_check_method.get(),
            "ping_min": getattr(self, "ping_min", 0),
            "ping_max": getattr(self, "ping_max", 9999),
            "scan_ports": getattr(self, "scan_ports", [443, 80, 8443, 2053, 2083, 2087, 2096]),
        }
        self._saved_mode_key = cfg["mode_key"]
        self._saved_ai_key = cfg["ai_key"]
        self._saved_target_key = cfg["target_key"]
        
        path_cfg = self._config_file()
        try:
            with open(path_cfg, "w", encoding="utf-8") as f:
                json.dump(cfg, f, ensure_ascii=False, indent=2)
                f.flush()
                try:
                    if hasattr(os, 'fsync'):
                        os.fsync(f.fileno())
                except Exception:
                    pass
        except Exception as e:
            try:
                messagebox.showerror(self.t("error_title"), self.t("save_error").format(str(e)))
            except Exception:
                pass

    def _on_lang_selected(self, lang):
        if self._loading_timer:
            self.root.after_cancel(self._loading_timer)
            self._loading_timer = None
        self.lang = lang
        self.is_rtl = (lang == "fa")
        self._font_cache.clear()
        self._fa_font_cache.clear()
        self.root.after(300, self.init_ui)
    
    def show_loading(self):
        self.loading_frame = tk.Frame(self.root, bg='#1e1e2e')
        self.loading_frame.pack(fill=tk.BOTH, expand=True)
        
        char, fa_f, color = self.fa_icon("rocket", 72)
        if fa_f:
            tk.Label(self.loading_frame, text=char, font=fa_f, bg='#1e1e2e', fg=color).pack(pady=(80, 15))
        else:
            tk.Label(self.loading_frame, text="🚀", font=("Segoe UI", 72), bg='#1e1e2e', fg='#89b4fa').pack(pady=(80, 15))
        tk.Label(self.loading_frame, text="CDN IP Scanner V1.6", font=("Segoe UI", 28, "bold"), bg='#1e1e2e', fg='#cdd6f4').pack()
        
        lang_frame = tk.Frame(self.loading_frame, bg='#1e1e2e')
        lang_frame.pack(pady=25)
        tk.Label(lang_frame, text="Choose language / انتخاب زبان / 选择语言 / Выберите язык", font=("Segoe UI", 11), bg='#1e1e2e', fg='#a6adc8').pack(pady=(0, 15))
        
        flags = [
            ("🇮🇷", "fa", "فارسی"),
            ("🇬🇧", "en", "English"),
            ("🇨🇳", "zh", "中文"),
            ("🇷🇺", "ru", "Русский"),
        ]
        btn_frame = tk.Frame(lang_frame, bg='#1e1e2e')
        btn_frame.pack()
        for flag, code, name in flags:
            b = tk.Button(
                btn_frame, text=f"{flag}  {name}",
                font=("Segoe UI", 12), bg='#313244', fg='#cdd6f4', activebackground='#89b4fa',
                relief=tk.FLAT, padx=20, pady=12, cursor="hand2",
                command=(lambda c=code: self._on_lang_selected(c))
            )
            b.pack(side=tk.LEFT, padx=8)
        
        footer_frame = tk.Frame(self.loading_frame, bg='#1e1e2e')
        footer_frame.pack(side=tk.BOTTOM, pady=25)
        tk.Label(footer_frame, text="Programming and design by: shahin salek tootoonchi", font=("Segoe UI", 10), bg='#1e1e2e', fg='#a6adc8').pack()
        link_f = tk.Frame(footer_frame, bg='#1e1e2e')
        link_f.pack(pady=4)
        for lbl, url in [("GitHub", GITHUB_REPO_URL), ("YouTube", YOUTUBE_URL), ("digicloud.tr", WEBSITE_URL)]:
            L = tk.Label(link_f, text=lbl, font=("Segoe UI", 10), bg='#1e1e2e', fg='#74c7ec', cursor="hand2")
            L.pack(side=tk.LEFT, padx=8)
            L.bind("<Button-1>", lambda e, u=url: webbrowser.open(u))
        # منتظر انتخاب زبان توسط کاربر می‌ماند
    
    def init_ui(self):
        self.loading_frame.destroy()
        self.root.title(self.t("app_title"))
        self.setup_ui()
        self.check_queue()
    
    def _setup_ttk_scrollbar_style(self):
        """استایل اسکرول‌بار ttk برای تم دارک: مرتب و مدرن."""
        if self.current_theme != 'dark':
            return
        try:
            style = ttk.Style()
            style.theme_use('clam')
            bg = self.theme.get('card_bg', '#1a1a1a')
            trough = self.theme.get('bg', '#0a0a0a')
            border = self.layout.get('border_highlight', '#525252')
            style.configure("Vertical.TScrollbar", background=trough, troughcolor=trough, bordercolor=border, arrowcolor=self.theme.get('fg', '#ffffff'), width=10)
            style.map("Vertical.TScrollbar", background=[('active', self.layout.get('button_border', '#525252'))])
        except Exception:
            pass
    
    def _scrolled_text_frame(self, parent, **text_kw):
        """Text با اسکرول‌بار ttk مدرن؛ برمی‌گرداند (text_widget, frame)."""
        frame = tk.Frame(parent, bg=text_kw.get('bg', self.theme.get('card_bg', '#1a1a1a')))
        text = tk.Text(frame, **text_kw)
        sb = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=text.yview)
        text.configure(yscrollcommand=sb.set)
        text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        return text, frame
    
    def setup_ui(self):
        self.root.configure(bg=self.theme['bg'])
        self._setup_ttk_scrollbar_style()
        
        outer = tk.Frame(self.root, bg=self.theme['bg'])
        outer.pack(fill=tk.BOTH, expand=True)
        
        # اسکرول عمودی برای ریسپانسیو بودن
        self.main_canvas = tk.Canvas(outer, bg=self.theme['bg'], highlightthickness=0)
        scrollbar = ttk.Scrollbar(outer)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.main_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.main_canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.configure(command=self.main_canvas.yview)
        
        content_frame = tk.Frame(self.main_canvas, bg=self.theme['bg'])
        self._content_window = self.main_canvas.create_window((0, 0), window=content_frame, anchor='nw')
        
        def on_frame_configure(e):
            self.main_canvas.configure(scrollregion=self.main_canvas.bbox("all"))
        content_frame.bind("<Configure>", on_frame_configure)
        
        def on_canvas_configure(e):
            self.main_canvas.itemconfig(self._content_window, width=e.width)
        self.main_canvas.bind("<Configure>", on_canvas_configure)
        
        def on_mousewheel(e):
            if hasattr(e, 'delta') and e.delta != 0:
                self.main_canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
            elif e.num == 5:
                self.main_canvas.yview_scroll(1, "units")
            elif e.num == 4:
                self.main_canvas.yview_scroll(-1, "units")
        def bind_mousewheel(e):
            self.main_canvas.bind_all("<MouseWheel>", on_mousewheel)
            self.main_canvas.bind_all("<Button-4>", on_mousewheel)
            self.main_canvas.bind_all("<Button-5>", on_mousewheel)
        def unbind_mousewheel(e):
            self.main_canvas.unbind_all("<MouseWheel>")
            self.main_canvas.unbind_all("<Button-4>")
            self.main_canvas.unbind_all("<Button-5>")
        self.main_canvas.bind("<Enter>", bind_mousewheel)
        self.main_canvas.bind("<Leave>", unbind_mousewheel)
        
        content_frame.columnconfigure(0, weight=1)
        for r in range(7):
            content_frame.rowconfigure(r, weight=0)
        content_frame.rowconfigure(5, weight=1, minsize=180)
        
        header = self.create_header(content_frame)
        stats = self.create_stats(content_frame)
        two_boxes_row = self.create_two_boxes_row(content_frame)
        btns = self.create_buttons(content_frame)
        prog = self.create_progress(content_frame)
        results = self.create_results(content_frame)
        footer = self.create_footer(content_frame)
        
        header.grid(row=0, column=0, sticky='ew', padx=20, pady=2)
        stats.grid(row=1, column=0, sticky='ew', padx=20, pady=2)
        two_boxes_row.grid(row=2, column=0, sticky='nsew', padx=20, pady=2)
        btns.grid(row=3, column=0, sticky='ew', padx=20, pady=2)
        prog.grid(row=4, column=0, sticky='ew', padx=20, pady=2)
        results.grid(row=5, column=0, sticky='nsew', padx=20, pady=4)
        footer.grid(row=6, column=0, sticky='ew', padx=20, pady=4)
        self._apply_loaded_config()
        self._update_target_display()
    
    def switch_lang(self, lang):
        self.lang = lang
        self.is_rtl = (lang == "fa")
        self._font_cache.clear()
        self._fa_font_cache.clear()
        for w in self.root.winfo_children():
            w.destroy()
        self.setup_ui()
    
    def create_header(self, parent):
        header = tk.Frame(parent, bg=self.theme['bg'])
        anchor_main = 'e' if self.is_rtl else 'w'
        title_frame = tk.Frame(header, bg=self.theme['bg'])
        title_frame.pack(side=tk.RIGHT if self.is_rtl else tk.LEFT)
        
        char, fa_f, fa_color = self.fa_icon("rocket", 24)
        if fa_f:
            title_row = tk.Frame(title_frame, bg=self.theme['bg'])
            title_row.pack(anchor=anchor_main)
            tk.Label(title_row, text=self.t("app_title"), font=self.font(24, bold=True), bg=self.theme['bg'], fg=self.theme['fg']).pack(side=tk.LEFT if not self.is_rtl else tk.RIGHT, padx=(0, 8) if not self.is_rtl else (8, 0))
            tk.Label(title_row, text=char, font=fa_f, bg=self.theme['bg'], fg=fa_color).pack(side=tk.LEFT if not self.is_rtl else tk.RIGHT)
        else:
            tk.Label(title_frame, text=self.t("app_title") + " 🚀", font=self.font(24, bold=True), bg=self.theme['bg'], fg=self.theme['fg']).pack(anchor=anchor_main)
        tk.Label(title_frame, text=self.t("app_subtitle"), font=self.font(11), bg=self.theme['bg'], fg=self.theme['fg']).pack(anchor=anchor_main)
        
        btn_frame = tk.Frame(header, bg=self.theme['bg'])
        btn_frame.pack(side=tk.LEFT if self.is_rtl else tk.RIGHT)
        border_gray = self.layout.get('button_border', '#525252')
        up_char, up_f, up_color = self.fa_icon("download", 12)
        up_text = (up_char if up_f else "⬇️") + " " + self.t("update_btn")
        update_btn = self._make_button(btn_frame, up_text, self._on_update_click, font=(up_f if up_f else self.font(10)), width=10, padx=8, pady=4, fg=(up_color if up_f else self.theme['fg']), border=border_gray)
        update_btn.pack(side=tk.RIGHT if self.is_rtl else tk.LEFT, padx=6)
        for flag, code in [("🇮🇷", "fa"), ("🇬🇧", "en"), ("🇨🇳", "zh"), ("🇷🇺", "ru")]:
            btn = self._make_button(btn_frame, flag, (lambda c=code: self.switch_lang(c)), font=("Segoe UI", 10), width=4, padx=4, pady=2, border=border_gray)
            btn.pack(side=tk.RIGHT if self.is_rtl else tk.LEFT, padx=2)
        th_char, th_f, th_color = self.fa_icon("moon" if self.current_theme == 'dark' else "sun", 16)
        th_text = th_char if th_f else ("🌙" if self.current_theme == 'dark' else "☀️")
        self.theme_btn = self._make_button(btn_frame, th_text, self.toggle_theme, font=(th_f if th_f else self.font(16)), width=4, padx=15, pady=5, fg=(th_color if th_f else self.theme['fg']), border=border_gray)
        self.theme_btn.pack(side=tk.RIGHT if self.is_rtl else tk.LEFT, padx=5)
        return header
    
    def create_stats(self, parent):
        stats_frame = tk.Frame(parent, bg=self.theme['bg'])
        tval = self.get_target_internal()
        tdisp = self.num(tval) if tval != "All" else self.t("target_all")
        stats_icons = [("bullseye", "target", tdisp, "target_label"), ("bolt", "speed", self.num(0) + " IP/s", "speed_label"), ("check", "found", self.num(0), "found_label"), ("clock", "time", self.num(0) + "s", "time_label")]
        for idx, (icon_name, key, value, var_name) in enumerate(stats_icons):
            card = tk.Frame(stats_frame, bg=self.theme['card_bg'], relief=tk.FLAT, bd=0, highlightbackground=self.layout.get('border_highlight', '#525252'), highlightthickness=1)
            card.grid(row=0, column=idx, padx=SECTION_PADDING // 2, pady=SECTION_PADDING // 2, sticky='ew')
            stats_frame.grid_columnconfigure(idx, weight=1, minsize=MIN_STAT_CARD_WIDTH)
            char, fa_f, fa_color = self.fa_icon(icon_name, 14)
            row_top = tk.Frame(card, bg=self.theme['card_bg'])
            row_top.pack(pady=(CARD_PADDING_Y, 0), anchor='center')
            if fa_f:
                tk.Label(row_top, text=char, font=fa_f, bg=self.theme['card_bg'], fg=fa_color).pack(side=tk.LEFT, padx=2)
            else:
                tk.Label(row_top, text=char, font=self.font(8), bg=self.theme['card_bg'], fg=self.theme['card_fg']).pack(side=tk.LEFT, padx=2)
            tk.Label(row_top, text=self.t(key), font=self.font(8), bg=self.theme['card_bg'], fg=self.theme['card_fg']).pack(side=tk.LEFT)
            label = tk.Label(card, text=value, font=self.font(11, bold=True), bg=self.theme['card_bg'], fg=self.theme['fg'])
            label.pack(pady=(0, CARD_PADDING_Y), anchor='center')
            setattr(self, var_name, label)
        return stats_frame
    
    def create_controls(self, parent):
        control_wrap = tk.Frame(parent, bg=self.theme['bg'])
        tk.Label(control_wrap, text=" " + self.t("box_range_and_scan") + " ", font=self.font(12, bold=True), bg=self.theme['bg'], fg=self.theme['fg']).pack(anchor='e' if self.is_rtl else 'w')
        center_frame = tk.Frame(control_wrap, bg=self.theme['bg'])
        sp = self.layout.get('section_padding', 12)
        cp_x, cp_y = self.layout.get('card_padding_x', 16), self.layout.get('card_padding_y', 12)
        center_frame.pack(anchor='e' if self.is_rtl else 'w', fill=tk.X, pady=(sp // 2, 0))
        btn_frame = tk.Frame(center_frame, bg=self.theme['bg'])
        btn_frame.pack(anchor='e' if self.is_rtl else 'w')
        pad_btn = cp_y
        pad_btn_h = cp_x
        s_char, s_f, s_color = self.fa_icon("satellite", 14)
        fetch_btn = tk.Frame(btn_frame, bg=self.theme['warning'], cursor="hand2")
        fetch_btn.pack(side=tk.RIGHT if self.is_rtl else tk.LEFT, padx=sp // 2)
        if s_f:
            tk.Label(fetch_btn, text=s_char, font=s_f, bg=self.theme['warning'], fg=s_color).pack(side=tk.LEFT, padx=(pad_btn_h, 4), pady=pad_btn)
        tk.Label(fetch_btn, text=self._strip_emoji(self.t("fetch_ranges_btn")), font=self.font(11, bold=True), bg=self.theme['warning'], fg='#000').pack(side=tk.LEFT, padx=(0, pad_btn_h), pady=pad_btn)
        for c in fetch_btn.winfo_children():
            c.bind("<Button-1>", lambda e: self.show_range_fetcher())
        side_label = tk.RIGHT if self.is_rtl else tk.LEFT
        scan_method_row = tk.Frame(center_frame, bg=self.theme['bg'])
        scan_method_row.pack(anchor='center', pady=(8, 0))
        tk.Label(scan_method_row, text=self.t("scan_method_label") + " ", font=self.font(10), bg=self.theme['bg'], fg=self.theme['fg']).pack(side=side_label)
        info_icon = tk.Label(scan_method_row, text="ⓘ", font=self.font(11), bg=self.theme['bg'], fg=self.theme['fg'], cursor="hand2")
        info_icon.pack(side=side_label, padx=(0, 4) if not self.is_rtl else (4, 0))
        def show_scan_method_info_ctl(e=None):
            if self.scan_method_var.get() == self.t("scan_method_operators"):
                messagebox.showinfo(self.t("scan_method_cloud"), self.t("operator_scan_basis"))
            else:
                messagebox.showinfo(self.t("scan_method_cloud"), self.t("scan_method_cloud_tooltip"))
        info_icon.bind("<Button-1>", show_scan_method_info_ctl)
        scan_method_vals = [self.t("scan_method_cloud"), self.t("scan_method_operators"), self.t("scan_method_v2ray")]
        if self.scan_method_var.get() not in scan_method_vals:
            self.scan_method_var.set(self.t("scan_method_cloud"))
        scan_method_combo = ttk.Combobox(scan_method_row, textvariable=self.scan_method_var, values=scan_method_vals, width=28, state="readonly", font=self.font(9))
        scan_method_combo.pack(side=side_label, padx=4)
        self.operators_frame_ctrl = None
        def on_scan_method_change(event=None):
            v = self.scan_method_var.get()
            if v == self.t("scan_method_v2ray"):
                messagebox.showwarning(self.t("error_title"), self.t("coming_soon_scan"))
                self.scan_method_var.set(self.t("scan_method_cloud"))
                scan_method_combo.set(self.t("scan_method_cloud"))
                return
            if hasattr(self, "operators_frame_ctrl") and self.operators_frame_ctrl is not None:
                if v == self.t("scan_method_operators"):
                    self.operators_frame_ctrl.pack(anchor='center', pady=(8, 0))
                else:
                    self.operators_frame_ctrl.pack_forget()
        scan_method_combo.bind("<<ComboboxSelected>>", on_scan_method_change)
        self.operators_frame_ctrl = tk.Frame(center_frame, bg=self.theme['bg'])
        operators = self.get_operators_for_current_context()
        if self.lang == "en":
            country_row = tk.Frame(self.operators_frame_ctrl, bg=self.theme['bg'])
            country_row.pack(anchor='center', pady=(0, 4))
            tk.Label(country_row, text=self.t("operator_country") + " ", font=self.font(9), bg=self.theme['bg'], fg=self.theme['fg']).pack(side=tk.LEFT, padx=(0, 4))
            country_vals = [self.t("country_iran"), self.t("country_russia"), self.t("country_china")]
            country_map = {self.t("country_iran"): "ir", self.t("country_russia"): "ru", self.t("country_china"): "cn"}
            self.current_operator_country_var = tk.StringVar(value=self.t("country_iran"))
            country_combo = ttk.Combobox(country_row, textvariable=self.current_operator_country_var, values=country_vals, width=12, state="readonly", font=self.font(9))
            country_combo.pack(side=tk.LEFT)
            def on_country_change_ctl(event=None):
                self.current_operator_country = country_map.get(self.current_operator_country_var.get(), "ir")
                ops = self.get_operators_for_current_context()
                op_names = [op_info["name"] for op_info in ops.values()]
                all_lbl = self.t("scan_from_operator_all")
                scan_from_combo_ctl["values"] = [all_lbl] + op_names
                self.current_operator_var.set(all_lbl)
                self.current_operator_key = None
                self._update_target_display()
            country_combo.bind("<<ComboboxSelected>>", on_country_change_ctl)
        all_label = self.t("scan_from_operator_all")
        op_names = [op_info["name"] for op_info in operators.values()]
        combo_vals = [all_label] + op_names
        self.current_operator_var = tk.StringVar(value=all_label)
        scan_from_row = tk.Frame(self.operators_frame_ctrl, bg=self.theme['bg'])
        scan_from_row.pack(anchor='center', pady=(0, 6))
        tk.Label(scan_from_row, text=self.t("scan_from_operator") + " ", font=self.font(9), bg=self.theme['bg'], fg=self.theme['fg']).pack(side=tk.LEFT, padx=(0, 4))
        scan_from_combo_ctl = ttk.Combobox(scan_from_row, textvariable=self.current_operator_var, values=combo_vals, width=14, state="readonly", font=self.font(9))
        scan_from_combo_ctl.pack(side=tk.LEFT)
        def on_operator_select_ctl(event=None):
            sel = self.current_operator_var.get()
            if sel == all_label:
                self.current_operator_key = None
            else:
                for k, v in self.get_operators_for_current_context().items():
                    if v["name"] == sel:
                        self.current_operator_key = k
                        break
            self._update_target_display()
        scan_from_combo_ctl.bind("<<ComboboxSelected>>", on_operator_select_ctl)
        on_operator_select_ctl()
        fetch_all_row = tk.Frame(self.operators_frame_ctrl, bg=self.theme['bg'])
        fetch_all_row.pack(anchor='center', pady=4)
        bs = self._button_style()
        fetch_all_btn = self._make_button(fetch_all_row, self.t("fetch_all_operators_btn"), self._fetch_all_operator_ranges, font=self.font(10), width=bs['width'], padx=bs['padx'], pady=bs['pady'])
        fetch_all_btn.pack()
        if self.scan_method_var.get() == self.t("scan_method_operators"):
            self.operators_frame_ctrl.pack(anchor='center', pady=(8, 0))
        else:
            self.operators_frame_ctrl.pack_forget()
        # نمایش منبع رنج فعلی (Cloudflare / Fastly API یا —)
        range_row = tk.Frame(center_frame, bg=self.theme['bg'])
        range_row.pack(anchor='center', pady=(8, 0))
        tk.Label(range_row, text=self.t("range_source_label") + " ", font=self.font(10), bg=self.theme['bg'], fg=self.theme['fg']).pack(side=side_label)
        self.range_source_value_label = tk.Label(range_row, text=(self.current_range_source or self.t("range_source_none")), font=self.font(10, bold=True), bg=self.theme['bg'], fg=self.theme['fg'])
        self.range_source_value_label.pack(side=side_label)
        ping_row_ctrl = tk.Frame(center_frame, bg=self.theme['bg'])
        ping_row_ctrl.pack(anchor='center', pady=(4, 0))
        self.ping_filter_label = tk.Label(ping_row_ctrl, text="", font=self.font(9), bg=self.theme['bg'], fg=self.theme['card_fg'])
        self.ping_filter_label.pack(side=side_label)
        self._update_ping_filter_label()
        return control_wrap
    
    def _update_ping_filter_label(self):
        """Update main page ping filter label from saved settings."""
        if not hasattr(self, "ping_filter_label") or not self.ping_filter_label.winfo_exists():
            return
        ping_min = getattr(self, "ping_min", 0)
        ping_max = getattr(self, "ping_max", 9999)
        if ping_min <= 0 and ping_max >= 9999:
            self.ping_filter_label.config(text=self.t("ping_filter_none"))
        else:
            self.ping_filter_label.config(text=self.t("ping_filter_display").format(ping_min, ping_max))
    
    def _operator_ranges_file(self, operator_key):
        """Path to save operator ranges JSON (only FA)."""
        return os.path.join(BASE_DIR, f"operator_ranges_{operator_key}.json")
    
    def _load_operator_count(self, operator_key):
        """Return saved count for operator from file, or None."""
        path = self._operator_ranges_file(operator_key)
        if not os.path.isfile(path):
            return None
        try:
            with open(path, "r", encoding="utf-8") as f:
                d = json.load(f)
            return d.get("count") or len(d.get("prefixes", []))
        except Exception:
            return None
    
    def _get_operator_ranges_for(self, op_key):
        """رنج‌های ذخیره‌شدهٔ یک اپراتور (لیست CIDR)."""
        if op_key not in OPERATORS_IRAN and op_key not in OPERATORS_CHINA and op_key not in OPERATORS_RUSSIA:
            return []
        path = self._operator_ranges_file(op_key)
        if not os.path.isfile(path):
            return []
        try:
            with open(path, "r", encoding="utf-8") as f:
                d = json.load(f)
            out = []
            for p in d.get("prefixes") or []:
                s = (p.get("prefix") if isinstance(p, dict) else str(p)).strip()
                if s and "/" in s and ":" not in s:
                    out.append(s)
            return out
        except Exception:
            return []
    
    def _get_operator_merged_ranges(self):
        """در حالت اپراتور: ادغام همهٔ رنج‌های ذخیره‌شدهٔ اپراتورها به‌صورت لیست CIDR."""
        out = []
        for op_key in self.get_operators_for_current_context():
            out.extend(self._get_operator_ranges_for(op_key))
        return out
    
    def _get_user_operator(self):
        """با IP عمومی کاربر، اپراتور فعلی را از روی رنج‌های ذخیره‌شده تشخیص بده. نام اپراتور یا None."""
        try:
            r = requests.get("https://api.ipify.org", timeout=6)
            r.raise_for_status()
            user_ip_s = r.text.strip()
        except Exception:
            return None
        if not user_ip_s:
            return None
        try:
            user_ip = ipaddress.ip_address(user_ip_s)
        except ValueError:
            return None
        for op_key, op_info in self.get_operators_for_current_context().items():
            path = self._operator_ranges_file(op_key)
            if not os.path.isfile(path):
                continue
            try:
                with open(path, "r", encoding="utf-8") as f:
                    d = json.load(f)
                prefixes = d.get("prefixes") or []
                for p in prefixes:
                    s = (p.get("prefix") if isinstance(p, dict) else str(p)).strip()
                    if not s or "/" not in s or ":" in s:
                        continue
                    try:
                        net = ipaddress.ip_network(s, strict=False)
                        if user_ip in net:
                            return op_info["name"]
                    except ValueError:
                        continue
            except Exception:
                continue
        return None
    
    def _get_user_operator_key(self):
        """همان تشخیص اپراتور ولی برگرداندن کلید (irancell, mci, ...) برای ذخیره در ماتریس."""
        try:
            r = requests.get("https://api.ipify.org", timeout=6)
            r.raise_for_status()
            user_ip_s = r.text.strip()
        except Exception:
            return None
        if not user_ip_s:
            return None
        try:
            user_ip = ipaddress.ip_address(user_ip_s)
        except ValueError:
            return None
        for op_key, op_info in self.get_operators_for_current_context().items():
            path = self._operator_ranges_file(op_key)
            if not os.path.isfile(path):
                continue
            try:
                with open(path, "r", encoding="utf-8") as f:
                    d = json.load(f)
                for p in d.get("prefixes") or []:
                    s = (p.get("prefix") if isinstance(p, dict) else str(p)).strip()
                    if not s or "/" not in s or ":" in s:
                        continue
                    try:
                        net = ipaddress.ip_network(s, strict=False)
                        if user_ip in net:
                            return op_key
                    except ValueError:
                        continue
            except Exception:
                continue
        return None
    
    def _operator_matrix_file(self):
        """مسیر فایل ذخیرهٔ وضعیت هر IP در هر اپراتور (وایت/بلاک از اسکن واقعی)."""
        return os.path.join(BASE_DIR, "operator_ip_matrix.json")
    
    def _load_operator_matrix(self):
        """بارگذاری ماتریس IP × اپراتور. ساختار: { "ip": { "irancell": true, "mci": false }, ... }."""
        path = self._operator_matrix_file()
        if not os.path.isfile(path):
            return {}
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    
    def _save_operator_matrix(self, matrix):
        """ذخیرهٔ ماتریس وضعیت اپراتورها."""
        path = self._operator_matrix_file()
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(matrix, f, ensure_ascii=False, indent=2)
        except Exception:
            pass
    
    def fetch_operator_ranges(self, operator_key, count_label_ref):
        """Fetch operator IP ranges from BGPView (or multi-ASN for China/Russia), save to file, update count label."""
        operators = self.get_operators_for_current_context()
        if operator_key not in operators:
            return
        asn = operators[operator_key]["asn"]
        path = self._operator_ranges_file(operator_key)
        if count_label_ref and count_label_ref.winfo_exists():
            count_label_ref.config(text=self.t("operator_fetching"))
        
        def do_fetch():
            if isinstance(asn, list):
                all_prefixes = []
                for a in asn:
                    cnt, prefixes, err = fetch_operator_prefixes_smart(a)
                    if not err and prefixes:
                        all_prefixes.extend(prefixes)
                    time.sleep(0.5)
                cnt, prefixes, err = len(all_prefixes), all_prefixes, None if all_prefixes else "no data"
            else:
                cnt, prefixes, err = fetch_operator_prefixes_smart(asn)
            display_text = self.t("operator_fetch_error") if err else self.t("operator_ranges_count").format(self.num(cnt))
            if not err:
                data = {"asn": asn, "count": cnt, "prefixes": prefixes}
                try:
                    with open(path, "w", encoding="utf-8") as f:
                        json.dump(data, f, ensure_ascii=False, indent=2)
                except Exception:
                    pass
            def update_ui():
                if count_label_ref and count_label_ref.winfo_exists():
                    count_label_ref.config(text=display_text)
                self._update_target_display()
            self.root.after(0, update_ui)
        
        threading.Thread(target=do_fetch, daemon=True).start()
    
    def _fetch_all_operator_ranges(self):
        """دریافت رنج همهٔ اپراتورهای کشور فعلی به‌ترتیب؛ بعد به‌روزرسانی باکس هدف."""
        operators = self.get_operators_for_current_context()
        def do_all():
            for op_key in operators:
                asn = operators[op_key]["asn"]
                path = self._operator_ranges_file(op_key)
                if isinstance(asn, list):
                    all_prefixes = []
                    for a in asn:
                        cnt, prefixes, err = fetch_operator_prefixes_smart(a)
                        if not err and prefixes:
                            all_prefixes.extend(prefixes)
                        time.sleep(0.5)
                    cnt, prefixes, err = len(all_prefixes), all_prefixes, None if all_prefixes else "no data"
                else:
                    cnt, prefixes, err = fetch_operator_prefixes_smart(asn)
                if not err:
                    data = {"asn": asn, "count": cnt, "prefixes": prefixes}
                    try:
                        with open(path, "w", encoding="utf-8") as f:
                            json.dump(data, f, ensure_ascii=False, indent=2)
                    except Exception:
                        pass
            self.root.after(0, self._update_target_display)
        threading.Thread(target=do_all, daemon=True).start()
    
    def create_two_boxes_row(self, parent):
        """یک ردیف با دو باکس: سمت چپ = رنج‌های اسکن، سمت راست = تنظیمات."""
        row = tk.Frame(parent, bg=self.theme['bg'])
        row.columnconfigure(0, weight=1)
        row.columnconfigure(1, weight=1)
        ranges_box = self.create_ranges_section(row)
        settings_box = self.create_controls_box(row)
        ranges_box.grid(row=0, column=0, sticky='nsew', padx=(0, 10))
        settings_box.grid(row=0, column=1, sticky='nsew', padx=(10, 0))
        return row
    
    def create_controls_box(self, parent):
        """باکس انتخاب رنج و نحوه اسکن (دکمه دریافت رنج + منبع رنج + نحوه اسکن)."""
        wrap = tk.Frame(parent, bg=self.theme['bg'])
        box = tk.Frame(wrap, bg=self.theme['card_bg'], relief=tk.FLAT, highlightbackground=self.layout.get('border_highlight', '#525252'), highlightthickness=1, padx=CARD_PADDING_X, pady=CARD_PADDING_Y)
        box.pack(fill=tk.BOTH, expand=True)
        tk.Label(box, text=" " + self.t("box_range_and_scan") + " ", font=self.font(11, bold=True), bg=self.theme['card_bg'], fg=self.theme['fg']).pack(anchor='e' if self.is_rtl else 'w')
        center_frame = tk.Frame(box, bg=self.theme['card_bg'])
        center_frame.pack(anchor='center', fill=tk.X, pady=(SECTION_PADDING // 2, 0))
        btn_frame = tk.Frame(center_frame, bg=self.theme['card_bg'])
        btn_frame.pack(anchor='center')
        pad_btn = CARD_PADDING_Y
        pad_btn_h = CARD_PADDING_X
        s_char, s_f, s_color = self.fa_icon("satellite", 14)
        border_gray = self.layout.get('button_border', '#525252')
        fetch_btn = tk.Frame(btn_frame, bg=self.theme['card_bg'], cursor="hand2", highlightbackground=border_gray, highlightthickness=1)
        fetch_btn.pack(side=tk.LEFT, padx=SECTION_PADDING // 2)
        if s_f:
            tk.Label(fetch_btn, text=s_char, font=s_f, bg=self.theme['card_bg'], fg=s_color).pack(side=tk.LEFT, padx=(pad_btn_h, 4), pady=pad_btn)
        tk.Label(fetch_btn, text=self._strip_emoji(self.t("fetch_ranges_btn")), font=self.font(11, bold=True), bg=self.theme['card_bg'], fg=self.theme['fg']).pack(side=tk.LEFT, padx=(0, pad_btn_h), pady=pad_btn)
        for c in fetch_btn.winfo_children():
            c.bind("<Button-1>", lambda e: self.show_range_fetcher())
        side_label = tk.RIGHT if self.is_rtl else tk.LEFT
        scan_method_row = tk.Frame(center_frame, bg=self.theme['card_bg'])
        scan_method_row.pack(anchor='center', pady=(8, 0))
        tk.Label(scan_method_row, text=self.t("scan_method_label") + " ", font=self.font(10), bg=self.theme['card_bg'], fg=self.theme['fg']).pack(side=side_label)
        info_icon = tk.Label(scan_method_row, text="ⓘ", font=self.font(11), bg=self.theme['card_bg'], fg=self.theme['fg'], cursor="hand2")
        info_icon.pack(side=side_label, padx=(0, 4) if not self.is_rtl else (4, 0))
        def show_scan_method_info(e=None):
            if self.scan_method_var.get() == self.t("scan_method_operators"):
                messagebox.showinfo(self.t("scan_method_cloud"), self.t("operator_scan_basis"))
            else:
                messagebox.showinfo(self.t("scan_method_cloud"), self.t("scan_method_cloud_tooltip"))
        info_icon.bind("<Button-1>", show_scan_method_info)
        scan_method_vals = [self.t("scan_method_cloud"), self.t("scan_method_operators"), self.t("scan_method_v2ray")]
        if self.scan_method_var.get() not in scan_method_vals:
            self.scan_method_var.set(self.t("scan_method_cloud"))
        scan_method_combo = ttk.Combobox(scan_method_row, textvariable=self.scan_method_var, values=scan_method_vals, width=28, state="readonly", font=self.font(9))
        scan_method_combo.pack(side=side_label, padx=4)
        self.operators_frame_box = None
        def on_scan_method_change_box(event=None):
            v = self.scan_method_var.get()
            if v == self.t("scan_method_v2ray"):
                messagebox.showwarning(self.t("error_title"), self.t("coming_soon_scan"))
                self.scan_method_var.set(self.t("scan_method_cloud"))
                scan_method_combo.set(self.t("scan_method_cloud"))
                return
            if hasattr(self, "operators_frame_box") and self.operators_frame_box is not None:
                if v == self.t("scan_method_operators"):
                    self.operators_frame_box.pack(anchor='center', pady=(8, 0))
                else:
                    self.operators_frame_box.pack_forget()
            self._update_target_display()
        scan_method_combo.bind("<<ComboboxSelected>>", on_scan_method_change_box)
        self.operators_frame_box = tk.Frame(center_frame, bg=self.theme['card_bg'])
        operators_box = self.get_operators_for_current_context()
        if self.lang == "en":
            country_row_box = tk.Frame(self.operators_frame_box, bg=self.theme['card_bg'])
            country_row_box.pack(anchor='center', pady=(0, 4))
            tk.Label(country_row_box, text=self.t("operator_country") + " ", font=self.font(9), bg=self.theme['card_bg'], fg=self.theme['fg']).pack(side=tk.LEFT, padx=(0, 4))
            country_vals = [self.t("country_iran"), self.t("country_russia"), self.t("country_china")]
            country_map = {self.t("country_iran"): "ir", self.t("country_russia"): "ru", self.t("country_china"): "cn"}
            self.current_operator_country_var = tk.StringVar(value=self.t("country_iran"))
            country_combo_box = ttk.Combobox(country_row_box, textvariable=self.current_operator_country_var, values=country_vals, width=12, state="readonly", font=self.font(9))
            country_combo_box.pack(side=tk.LEFT)
            def on_country_change_box(event=None):
                self.current_operator_country = country_map.get(self.current_operator_country_var.get(), "ir")
                ops = self.get_operators_for_current_context()
                op_names = [op_info["name"] for op_info in ops.values()]
                all_lbl = self.t("scan_from_operator_all")
                scan_from_combo_box["values"] = [all_lbl] + op_names
                self.current_operator_var.set(all_lbl)
                self.current_operator_key = None
                self._update_target_display()
            country_combo_box.bind("<<ComboboxSelected>>", on_country_change_box)
        all_label = self.t("scan_from_operator_all")
        op_names = [op_info["name"] for op_info in operators_box.values()]
        combo_vals = [all_label] + op_names
        self.current_operator_var = tk.StringVar(value=all_label)
        scan_from_row = tk.Frame(self.operators_frame_box, bg=self.theme['card_bg'])
        scan_from_row.pack(anchor='center', pady=(0, 6))
        tk.Label(scan_from_row, text=self.t("scan_from_operator") + " ", font=self.font(9), bg=self.theme['card_bg'], fg=self.theme['fg']).pack(side=tk.LEFT, padx=(0, 4))
        scan_from_combo_box = ttk.Combobox(scan_from_row, textvariable=self.current_operator_var, values=combo_vals, width=14, state="readonly", font=self.font(9))
        scan_from_combo_box.pack(side=tk.LEFT)
        def on_operator_select(event=None):
            sel = self.current_operator_var.get()
            if sel == all_label:
                self.current_operator_key = None
                return
            for k, v in self.get_operators_for_current_context().items():
                if v["name"] == sel:
                    self.current_operator_key = k
                    return
            self.current_operator_key = None
        scan_from_combo_box.bind("<<ComboboxSelected>>", on_operator_select)
        on_operator_select()
        fetch_all_row_box = tk.Frame(self.operators_frame_box, bg=self.theme['card_bg'])
        fetch_all_row_box.pack(anchor='center', pady=4)
        bs = self._button_style()
        self._make_button(fetch_all_row_box, self.t("fetch_all_operators_btn"), self._fetch_all_operator_ranges, font=self.font(10), width=bs['width'], padx=bs['padx'], pady=bs['pady'], bg=self.theme['card_bg'], fg=self.theme['fg'], border=bs['border']).pack()
        if self.scan_method_var.get() == self.t("scan_method_operators"):
            self.operators_frame_box.pack(anchor='center', pady=(8, 0))
        else:
            self.operators_frame_box.pack_forget()
        range_row = tk.Frame(center_frame, bg=self.theme['card_bg'])
        range_row.pack(anchor='center', pady=(8, 0))
        tk.Label(range_row, text=self.t("range_source_label") + " ", font=self.font(10), bg=self.theme['card_bg'], fg=self.theme['fg']).pack(side=side_label)
        self.range_source_value_label = tk.Label(range_row, text=(self.current_range_source or self.t("range_source_none")), font=self.font(10, bold=True), bg=self.theme['card_bg'], fg=self.theme['fg'])
        self.range_source_value_label.pack(side=side_label)
        ping_row2 = tk.Frame(center_frame, bg=self.theme['card_bg'])
        ping_row2.pack(anchor='center', pady=(4, 0))
        self.ping_filter_label = tk.Label(ping_row2, text="", font=self.font(9), bg=self.theme['card_bg'], fg=self.theme['card_fg'])
        self.ping_filter_label.pack(side=side_label)
        self._update_ping_filter_label()
        return wrap
    
    def create_ranges_section(self, parent):
        """باکس آی‌پی/رنج: پیست دستی یا دریافت رنج؛ هر خط یک آی‌پی یا CIDR."""
        wrap = tk.Frame(parent, bg=self.theme['bg'])
        box = tk.Frame(wrap, bg=self.theme['card_bg'], relief=tk.FLAT, highlightbackground=self.layout.get('border_highlight', '#525252'), highlightthickness=1, padx=CARD_PADDING_X, pady=CARD_PADDING_Y)
        box.pack(fill=tk.BOTH, expand=True)
        anchor_sec = 'e' if self.is_rtl else 'w'
        tk.Label(box, text=" " + self.t("ranges_to_scan_label") + " ", font=self.font(11, bold=True), bg=self.theme['card_bg'], fg=self.theme['fg']).pack(anchor=anchor_sec)
        tk.Label(box, text=self.t("ranges_hint_paste_fetch"), font=self.font(9), bg=self.theme['card_bg'], fg=self.theme['card_fg']).pack(anchor=anchor_sec, pady=(0, 4))
        list_container = tk.Frame(box, bg=self.theme['card_bg'], relief=tk.FLAT)
        list_container.pack(fill=tk.BOTH, expand=True, pady=(4, 0))
        self.ranges_text_main, st_frame = self._scrolled_text_frame(list_container, height=6, font=self.font(10), bg=self.theme['card_bg'], fg=self.theme['fg'], wrap=tk.WORD)
        if self.is_rtl:
            self.ranges_text_main.tag_configure("rtl", justify=tk.RIGHT)
        st_frame.pack(fill=tk.BOTH, expand=True)
        if not hasattr(self, "filter_only_clean_var"):
            self.filter_only_clean_var = tk.BooleanVar(value=True)
        filter_row = tk.Frame(box, bg=self.theme['card_bg'])
        filter_row.pack(anchor='e' if self.is_rtl else 'w', pady=(6, 0))
        tk.Checkbutton(filter_row, text=self.t("filter_only_clean_label"), variable=self.filter_only_clean_var, font=self.font(9), bg=self.theme['card_bg'], fg=self.theme['fg'], activebackground=self.theme['card_bg'], activeforeground=self.theme['fg'], selectcolor=self.theme['card_bg']).pack(side=tk.LEFT if not self.is_rtl else tk.RIGHT)
        add_row = tk.Frame(box, bg=self.theme['card_bg'])
        add_row.pack(fill=tk.X, pady=(8, 0))
        tk.Label(add_row, text=self.t("add_range_manual"), font=self.font(10), bg=self.theme['card_bg'], fg=self.theme['fg']).pack(side=tk.LEFT if not self.is_rtl else tk.RIGHT, padx=(0, 8) if self.is_rtl else (8, 0))
        self.add_range_entry = tk.Entry(add_row, width=28, font=self.font(10), bg=self.theme['bg'], fg=self.theme['fg'])
        self.add_range_entry.pack(side=tk.LEFT if not self.is_rtl else tk.RIGHT, padx=4, fill=tk.X, expand=True)
        def on_add_range():
            s = (self.add_range_entry.get() or "").strip()
            if not s:
                return
            if not self._is_valid_ip_or_cidr(s):
                messagebox.showwarning(self.t("error_title"), self.t("invalid_ip_cidr") + ": " + s)
                return
            self.ranges_text_main.insert(tk.END, s + "\n")
            self.add_range_entry.delete(0, tk.END)
        bs = self._button_style()
        self._make_button(add_row, self.t("add_range_btn"), on_add_range, font=self.font(10), width=bs['width'], padx=bs['padx'], pady=bs['pady'], bg=self.theme['card_bg'], fg=self.theme['fg'], border=bs['border']).pack(side=tk.LEFT if not self.is_rtl else tk.RIGHT)
        return wrap
    
    def _is_valid_ip_or_cidr(self, s):
        """Return True if s is a valid IPv4 address or CIDR."""
        s = (s or "").strip()
        if not s:
            return False
        try:
            if "/" in s:
                ipaddress.IPv4Network(s, strict=False)
            else:
                ipaddress.IPv4Address(s)
            return True
        except Exception:
            return False
    
    def _range_to_ips(self, range_str, max_ips):
        """Convert a single IP or CIDR string to list of IP objects to scan."""
        range_str = (range_str or "").strip()
        if not range_str:
            return []
        try:
            if "/" not in range_str:
                return [ipaddress.IPv4Address(range_str)]
            return self.ai.smart_sample(range_str, max_ips)
        except Exception:
            return []
    
    def refresh_ranges_selection_ui(self, ranges):
        """After fetch: set main ranges text to fetched ranges (one per line)."""
        if not hasattr(self, "ranges_text_main") or not self.ranges_text_main.winfo_exists():
            return
        self.ranges_text_main.delete(1.0, tk.END)
        if ranges:
            self.ranges_text_main.insert(tk.END, "\n".join(ranges))
            if not ranges[-1].endswith("\n"):
                self.ranges_text_main.insert(tk.END, "\n")
        if self.is_rtl:
            self.ranges_text_main.tag_add("rtl", "1.0", tk.END)
    
    def get_selected_ranges(self):
        """منبع آی‌پی برای اسکن: اول باکس (دستی/کلودفلیر/فستلی)؛ اگر خالی بود و حالت «با رنج اپراتورها»: همه = ادغام همهٔ اپراتورها، یک اپراتور = فقط همان اپراتور."""
        user_ranges = []
        if hasattr(self, "ranges_text_main") and self.ranges_text_main.winfo_exists():
            text = self.ranges_text_main.get(1.0, tk.END)
            lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
            user_ranges = [ln for ln in lines if self._is_valid_ip_or_cidr(ln)]
        if user_ranges:
            return user_ranges
        if self._is_operator_scan_mode():
            op_key = getattr(self, 'current_operator_key', None)
            if op_key is not None:
                return self._get_operator_ranges_for(op_key)
            return self._get_operator_merged_ranges()
        return []
    
    def show_settings_dialog(self):
        """پنجره تنظیمات: اصلی، تم، روش بررسی IP."""
        # نمایش همیشه مطابق _saved_* (همان چیزی که آخر بار ذخیره شد)
        # از فایل نخوان اینجا — _saved_* موقع ذخیره آپدیت شده؛ با همین comboها را پر کن
        if hasattr(self, "_saved_mode_key"):
            self.mode_var.set(self.t(self._saved_mode_key))
        if hasattr(self, "_saved_ai_key"):
            self.ai_var.set(self.t(self._saved_ai_key))
        if hasattr(self, "_saved_target_key"):
            self.target_var.set(self.t(self._saved_target_key))
        
        win = tk.Toplevel(self.root)
        win.title(self.t("settings"))
        win.configure(bg=self.theme['bg'])
        win.resizable(True, True)
        win.transient(self.root)
        win.grab_set()
        
        rtl = self.is_rtl
        anchor_sec = 'e' if rtl else 'w'
        side_1 = tk.RIGHT if rtl else tk.LEFT
        side_2 = tk.LEFT if rtl else tk.RIGHT
        
        main_f = tk.Frame(win, bg=self.theme['bg'], padx=20, pady=15)
        main_f.pack(fill=tk.BOTH, expand=True)
        
        # === تنظیمات اصلی (راست‌چین در فارسی) ===
        tk.Label(main_f, text=self.t("settings_main"), font=self.font(12, bold=True), bg=self.theme['bg'], fg=self.theme['fg']).pack(anchor=anchor_sec, pady=(0, 8))
        row1 = tk.Frame(main_f, bg=self.theme['bg'])
        row1.pack(fill=tk.X, pady=5, anchor=anchor_sec)
        mode_vals = [self.t("mode_turbo"), self.t("mode_hyper"), self.t("mode_ultra"), self.t("mode_deep")]
        ai_vals = [self.t("ai_basic"), self.t("ai_smart"), self.t("ai_advanced"), self.t("ai_expert")]
        target_vals = [self.t(k) for k in self._target_keys]
        justify_combo = 'right' if rtl else 'left'
        tk.Label(row1, text=self.t("mode") + " ", font=self.font(10), bg=self.theme['bg'], fg=self.theme['fg']).pack(side=side_1, padx=(0, 5) if rtl else (5, 0))
        combo_mode = ttk.Combobox(row1, textvariable=self.mode_var, values=mode_vals, width=18, state="readonly", justify=justify_combo, font=self.font(9))
        combo_mode.pack(side=side_1, padx=5)
        mode_desc_keys = {"mode_turbo": "mode_turbo_desc", "mode_hyper": "mode_hyper_desc", "mode_ultra": "mode_ultra_desc", "mode_deep": "mode_deep_desc"}
        def mode_desc_text():
            v = self.mode_var.get()
            for k in ("mode_turbo", "mode_hyper", "mode_ultra", "mode_deep"):
                if self.t(k) == v:
                    return self.t(mode_desc_keys[k])
            return ""
        mode_desc_label = tk.Label(main_f, text=mode_desc_text(), font=self.font(9), bg=self.theme['bg'], fg=self.theme['card_fg'])
        mode_desc_label.pack(anchor=anchor_sec, pady=(2, 0))
        def on_mode_change(e=None):
            mode_desc_label.config(text=mode_desc_text())
        combo_mode.bind("<<ComboboxSelected>>", on_mode_change)
        tk.Label(row1, text=self.t("ai") + " ", font=self.font(10), bg=self.theme['bg'], fg=self.theme['fg']).pack(side=side_1, padx=(15, 5) if rtl else (5, 15))
        combo_ai = ttk.Combobox(row1, textvariable=self.ai_var, values=ai_vals, width=14, state="readonly", justify=justify_combo, font=self.font(9))
        combo_ai.pack(side=side_1, padx=5)
        tk.Label(row1, text=self.t("count") + " ", font=self.font(10), bg=self.theme['bg'], fg=self.theme['fg']).pack(side=side_1, padx=(15, 5) if rtl else (5, 15))
        combo_target = ttk.Combobox(row1, textvariable=self.target_var, values=target_vals, width=12, state="readonly", justify=justify_combo, font=self.font(9))
        combo_target.pack(side=side_1, padx=5)
        count_desc_label = tk.Label(main_f, text=self.t("count_desc"), font=self.font(9), bg=self.theme['bg'], fg=self.theme['card_fg'])
        count_desc_label.pack(anchor=anchor_sec, pady=(2, 8))
        # مطمئن شو comboها مقدار فعلی var را نشان می‌دهند (ttk گاهی sync نمی‌کند)
        try:
            combo_mode.set(self.mode_var.get())
            combo_ai.set(self.ai_var.get())
            combo_target.set(self.target_var.get())
        except Exception:
            pass
        
        # === Ping range filter ===
        tk.Label(main_f, text=self.t("settings_ping_range"), font=self.font(12, bold=True), bg=self.theme['bg'], fg=self.theme['fg']).pack(anchor=anchor_sec, pady=(20, 8))
        ping_frame = tk.Frame(main_f, bg=self.theme['bg'])
        ping_frame.pack(fill=tk.X, pady=5, anchor=anchor_sec)
        ping_min_var = tk.StringVar(value=str(getattr(self, "ping_min", 0)))
        ping_max_var = tk.StringVar(value=str(getattr(self, "ping_max", 9999)))
        tk.Label(ping_frame, text=self.t("ping_min_label"), font=self.font(10), bg=self.theme['bg'], fg=self.theme['fg']).pack(side=side_1, padx=5)
        entry_ping_min = tk.Entry(ping_frame, textvariable=ping_min_var, width=8, font=self.font(10), bg=self.theme['card_bg'], fg=self.theme['fg'])
        entry_ping_min.pack(side=side_1, padx=5)
        tk.Label(ping_frame, text=self.t("ping_max_label"), font=self.font(10), bg=self.theme['bg'], fg=self.theme['fg']).pack(side=side_1, padx=(15, 5) if not rtl else (5, 15))
        entry_ping_max = tk.Entry(ping_frame, textvariable=ping_max_var, width=8, font=self.font(10), bg=self.theme['card_bg'], fg=self.theme['fg'])
        entry_ping_max.pack(side=side_1, padx=5)
        
        # === Scan ports (comma-separated) ===
        tk.Label(main_f, text=self.t("settings_ports"), font=self.font(12, bold=True), bg=self.theme['bg'], fg=self.theme['fg']).pack(anchor=anchor_sec, pady=(20, 8))
        ports_var = tk.StringVar(value=",".join(map(str, getattr(self, "scan_ports", [443, 80, 8443, 2053, 2083, 2087, 2096]))))
        tk.Label(main_f, text=self.t("ports_label"), font=self.font(10), bg=self.theme['bg'], fg=self.theme['fg']).pack(anchor=anchor_sec, pady=(0, 2))
        entry_ports = tk.Entry(main_f, textvariable=ports_var, width=45, font=self.font(10), bg=self.theme['card_bg'], fg=self.theme['fg'])
        entry_ports.pack(fill=tk.X, pady=2, anchor=anchor_sec)
        tk.Label(main_f, text=self.t("ports_hint"), font=self.font(9), bg=self.theme['bg'], fg=self.theme['card_fg']).pack(anchor=anchor_sec)
        
        # === تغییر رنگ / تم ===
        tk.Label(main_f, text=self.t("settings_theme"), font=self.font(12, bold=True), bg=self.theme['bg'], fg=self.theme['fg']).pack(anchor=anchor_sec, pady=(20, 8))
        theme_frame = tk.Frame(main_f, bg=self.theme['bg'])
        theme_frame.pack(fill=tk.X, pady=5, anchor=anchor_sec)
        theme_var = tk.StringVar(value=self.current_theme)
        tk.Radiobutton(theme_frame, text="🌙 Dark", variable=theme_var, value="dark", font=self.font(10), bg=self.theme['bg'], fg=self.theme['fg'], selectcolor=self.theme['card_bg'], activebackground=self.theme['bg'], activeforeground=self.theme['fg']).pack(side=side_1, padx=10)
        tk.Radiobutton(theme_frame, text="☀️ Light", variable=theme_var, value="light", font=self.font(10), bg=self.theme['bg'], fg=self.theme['fg'], selectcolor=self.theme['card_bg'], activebackground=self.theme['bg'], activeforeground=self.theme['fg']).pack(side=side_1, padx=10)
        
        # === تغییر بررسی آی پی ها: فقط متن؛ گزینه ایجنت قابل انتخاب نیست، با کلیک فقط پیام نمایش داده می‌شود ===
        tk.Label(main_f, text=self.t("settings_ip_check"), font=self.font(12, bold=True), bg=self.theme['bg'], fg=self.theme['fg']).pack(anchor=anchor_sec, pady=(20, 8))
        ip_check_frame = tk.Frame(main_f, bg=self.theme['bg'])
        ip_check_frame.pack(fill=tk.X, pady=5, anchor=anchor_sec)
        tk.Label(ip_check_frame, text=self.t("ip_check_my_internet"), font=self.font(10), bg=self.theme['bg'], fg=self.theme['fg']).pack(side=side_1, padx=5)
        agent_link = tk.Label(ip_check_frame, text=self.t("ip_check_agent") + " (?)", font=self.font(10), bg=self.theme['bg'], fg=self.theme['fg'], cursor="hand2", underline=True)
        agent_link.pack(side=side_1, padx=5)
        agent_link.bind("<Button-1>", lambda e: messagebox.showinfo(self.t("settings_ip_check"), self.t("ip_check_agent_message")))
        
        def on_save():
            win.update()
            self.current_theme = theme_var.get()
            self.theme = ThemeManager.get_theme(self.current_theme)
            mode_display = (combo_mode.get() or "").strip()
            ai_display = (combo_ai.get() or "").strip()
            target_display = (combo_target.get() or "").strip()
            try:
                self.ping_min = int(ping_min_var.get().strip() or "0")
            except ValueError:
                self.ping_min = 0
            try:
                self.ping_max = int(ping_max_var.get().strip() or "9999")
            except ValueError:
                self.ping_max = 9999
            if self.ping_min > self.ping_max:
                self.ping_min, self.ping_max = self.ping_max, self.ping_min
            ports_str = (entry_ports.get() or "").strip()
            if ports_str:
                self.scan_ports = [int(p) for p in ports_str.replace(",", " ").split() if str(p).strip().isdigit()]
            if not self.scan_ports:
                self.scan_ports = [443, 80, 8443, 2053, 2083, 2087, 2096]
            self.ai.priority_ports = self.scan_ports
            self._save_config(mode_display=mode_display, ai_display=ai_display, target_display=target_display)
            if hasattr(self, "ping_filter_label") and self.ping_filter_label.winfo_exists():
                self._update_ping_filter_label()
            messagebox.showinfo(self.t("success_title"), self.t("settings_saved"))
            win.destroy()
            # اعمال فوری تم با بازسازی UI
            self._font_cache.clear()
            self._fa_font_cache.clear()
            for w in self.root.winfo_children():
                w.destroy()
            self.setup_ui()
            self.check_queue()
        
        btn_row = tk.Frame(main_f, bg=self.theme['bg'])
        btn_row.pack(pady=25, anchor=anchor_sec)
        bs = self._button_style()
        self._make_button(btn_row, self.t("save_btn"), on_save, font=self.font(11, bold=True), width=bs['width'], padx=bs['padx'], pady=bs['pady']).pack(side=side_1, padx=8)
        self._make_button(btn_row, self.t("close_btn"), win.destroy, font=self.font(10), width=bs['width'], padx=bs['padx'], pady=bs['pady']).pack(side=side_1, padx=8)
    
    def _strip_emoji(self, s):
        """حذف ایموجی از انتهای متن برای نمایش کنار آیکن FA."""
        for c in "🤖💾⏹️🚀📡⚙️⚡✅⏱️🎯📊⭐🔓📍🏆📺💎🌙☀️":
            s = s.replace(c, "")
        return s.strip()
    
    def _button_style(self):
        """استایل مینیمال: حاشیه ۱px، گوشهٔ گرد ۵px در تم دارک."""
        w = self.layout.get('button_min_width', 140)
        return {
            'width': max(14, w // 8),
            'padx': self.layout.get('button_padx', 20),
            'pady': self.layout.get('button_pady', 10),
            'border': self.layout.get('button_border', '#525252'),
            'radius': self.layout.get('button_border_radius', 5),
        }
    
    def _draw_rounded_rect(self, canvas, x1, y1, x2, y2, r, fill, outline, outline_width=1):
        """رسم مستطیل با گوشهٔ گرد؛ fill و outline."""
        # پر کردن
        canvas.create_rectangle(x1 + r, y1, x2 - r, y2, fill=fill, outline="")
        canvas.create_rectangle(x1, y1 + r, x2, y2 - r, fill=fill, outline="")
        canvas.create_arc(x1, y1, x1 + 2 * r, y1 + 2 * r, start=90, extent=90, style=tk.PIESLICE, fill=fill, outline="")
        canvas.create_arc(x2 - 2 * r, y1, x2, y1 + 2 * r, start=0, extent=90, style=tk.PIESLICE, fill=fill, outline="")
        canvas.create_arc(x2 - 2 * r, y2 - 2 * r, x2, y2, start=270, extent=90, style=tk.PIESLICE, fill=fill, outline="")
        canvas.create_arc(x1, y2 - 2 * r, x1 + 2 * r, y2, start=180, extent=90, style=tk.PIESLICE, fill=fill, outline="")
        # حاشیه ۱ پیکسل
        canvas.create_line(x1 + r, y1, x2 - r, y1, fill=outline, width=outline_width)
        canvas.create_line(x2 - r, y1 + r, x2 - r, y2 - r, fill=outline, width=outline_width)
        canvas.create_line(x2 - r, y2 - r, x1 + r, y2 - r, fill=outline, width=outline_width)
        canvas.create_line(x1 + r, y2 - r, x1 + r, y1 + r, fill=outline, width=outline_width)
        canvas.create_arc(x1, y1, x1 + 2 * r, y1 + 2 * r, start=90, extent=90, style=tk.ARC, outline=outline, width=outline_width)
        canvas.create_arc(x2 - 2 * r, y1, x2, y1 + 2 * r, start=0, extent=90, style=tk.ARC, outline=outline, width=outline_width)
        canvas.create_arc(x2 - 2 * r, y2 - 2 * r, x2, y2, start=270, extent=90, style=tk.ARC, outline=outline, width=outline_width)
        canvas.create_arc(x1, y2 - 2 * r, x1 + 2 * r, y2, start=180, extent=90, style=tk.ARC, outline=outline, width=outline_width)
    
    def _make_button(self, parent, text, command, **kwargs):
        """در تم دارک: دکمهٔ گرد با حاشیه ۱px و radius ۵px؛ در تم روشن: دکمهٔ معمولی."""
        bs = self._button_style()
        bg = kwargs.get('bg', self.theme['bg'])
        fg = kwargs.get('fg', self.theme['fg'])
        font = kwargs.get('font', self.font(10))
        width = kwargs.get('width', bs['width'])
        padx = kwargs.get('padx', bs['padx'])
        pady = kwargs.get('pady', bs['pady'])
        border = kwargs.get('border', bs['border'])
        radius = min(bs.get('radius', 5), 20)
        if self.current_theme == 'dark' and radius > 0:
            # دکمهٔ گرد با Canvas
            approx_w = max(80, width * 8 + 2 * padx)
            approx_h = 2 * pady + 22
            c = tk.Canvas(parent, width=approx_w, height=approx_h, bg=parent.cget('bg') if parent.winfo_class() == 'Frame' else self.theme['bg'], highlightthickness=0, cursor="hand2")
            def draw():
                c.delete("all")
                w, h = c.winfo_width(), c.winfo_height()
                if w <= 1:
                    w, h = approx_w, approx_h
                r = min(radius, (min(w, h) // 2) - 1)
                self._draw_rounded_rect(c, 1, 1, w - 1, h - 1, r, fill=bg, outline=border, outline_width=1)
                c.create_text(w // 2, h // 2, text=text, fill=fg, font=font)
            c.after_idle(draw)
            c.bind("<Configure>", lambda e: draw())
            c.bind("<Button-1>", lambda e: command())
            return c
        else:
            return tk.Button(parent, text=text, font=font, bg=bg, fg=fg, activebackground=bg, activeforeground=fg, relief=tk.FLAT, width=width, padx=padx, pady=pady, cursor="hand2", command=command, highlightbackground=border, highlightthickness=1)
    
    def create_buttons(self, parent):
        btn_frame = tk.Frame(parent, bg=self.theme['bg'])
        right_buttons = tk.Frame(btn_frame, bg=self.theme['bg'])
        right_buttons.pack(anchor='center')
        bs = self._button_style()
        btn_bg = self.theme['bg']
        buttons = [
            ("analyze_btn", "robot", self.ai_analyze, "analyze_btn"),
            ("reset_data_btn", "gear", self._reset_all_data, "reset_btn"),
            ("settings", "gear", self.show_settings_dialog, "settings_btn"),
            ("save_btn", "save", self.export_menu, "export_btn"),
            ("stop_btn", "stop", self.stop_scan, "stop_btn"),
            ("start_btn", "play", self.start_scan, "start_btn")
        ]
        side_btn = tk.RIGHT if self.is_rtl else tk.LEFT
        for text_key, icon_name, command, var_name in buttons:
            row = tk.Frame(right_buttons, bg=btn_bg)
            row.pack(side=side_btn, padx=5)
            char, fa_f, fa_color = self.fa_icon(icon_name, 12)
            if fa_f:
                tk.Label(row, text=char, font=fa_f, bg=btn_bg, fg=fa_color).pack(side=tk.RIGHT if self.is_rtl else tk.LEFT, padx=(10, 4) if not self.is_rtl else (4, 10), pady=bs['pady'])
            btn = self._make_button(row, self._strip_emoji(self.t(text_key)), command, font=self.font(11, bold=True), width=bs['width'], padx=bs['padx'], pady=bs['pady'], bg=btn_bg, fg=self.theme['fg'], border=bs['border'])
            btn.pack(side=tk.RIGHT if self.is_rtl else tk.LEFT, padx=(0, 14) if not self.is_rtl else (14, 0))
            setattr(self, var_name, btn)
        if isinstance(self.stop_btn, tk.Button):
            self.stop_btn.config(state=tk.DISABLED)
        return btn_frame
    
    PROGRESS_BAR_WIDTH = 400
    
    def _set_progress_percent(self, percent):
        """تنظیم نوار پیشرفت بر اساس درصد ۰–۱۰۰؛ در ۱۰۰٪ حتماً تا انتها پر شود."""
        percent = max(0, min(100, float(percent)))
        if not hasattr(self, '_progress_inner') or not hasattr(self, '_progress_outer'):
            return
        try:
            if not self._progress_inner.winfo_exists():
                return
        except Exception:
            return
        try:
            total_w = self._progress_outer.winfo_width()
            if total_w <= 0:
                total_w = self.PROGRESS_BAR_WIDTH
            if percent >= 99.5:
                w = total_w
            else:
                w = int(total_w * percent / 100.0)
            self._progress_inner.place(x=0, y=0, width=w, relheight=1)
            self._progress_inner.update_idletasks()
        except Exception:
            pass
    
    def create_progress(self, parent):
        prog_wrap = tk.Frame(parent, bg=self.theme['bg'])
        tk.Label(prog_wrap, text=" " + self.t("progress_title") + " ", font=self.font(11, bold=True), bg=self.theme['bg'], fg=self.theme['fg']).pack(anchor='e' if self.is_rtl else 'w')
        progress_frame = tk.Frame(prog_wrap, bg=self.theme['card_bg'], relief=tk.FLAT, highlightbackground=self.layout.get('border_highlight', '#525252'), highlightthickness=1, padx=CARD_PADDING_X, pady=6)
        progress_frame.pack(fill=tk.X, pady=(SECTION_PADDING // 2, 0))
        # یک خط: متن آماده + نوار پیشرفت، ارتفاع کم
        row_prog = tk.Frame(progress_frame, bg=self.theme['card_bg'])
        row_prog.pack(fill=tk.X)
        self.status_label = tk.Label(row_prog, text=self.t("ready_status"), font=self.font(10, bold=True), bg=self.theme['card_bg'], fg=self.theme['fg'])
        self.status_label.pack(side=tk.RIGHT if self.is_rtl else tk.LEFT, padx=(8, 0) if self.is_rtl else (0, 8))
        self._progress_outer = tk.Frame(row_prog, width=self.PROGRESS_BAR_WIDTH, height=16, bg=self.theme.get('card_fg', '#4c4f69'))
        self._progress_outer.pack(side=tk.RIGHT if self.is_rtl else tk.LEFT, fill=tk.X, expand=True)
        self._progress_outer.pack_propagate(False)
        self._progress_inner = tk.Frame(self._progress_outer, width=0, height=16, bg=self.layout.get('button_border', '#525252'))
        self._progress_inner.place(x=0, y=0, width=0, relheight=1)
        return prog_wrap
    
    def create_results(self, parent):
        res_wrap = tk.Frame(parent, bg=self.theme['bg'])
        title_row = tk.Frame(res_wrap, bg=self.theme['bg'])
        title_row.pack(anchor='e' if self.is_rtl else 'w', fill=tk.X)
        tk.Label(title_row, text=" " + self.t("results_title") + " ", font=self.font(12, bold=True), bg=self.theme['bg'], fg=self.theme['fg']).pack(side=tk.RIGHT if self.is_rtl else tk.LEFT)
        results_frame = tk.Frame(res_wrap, bg=self.theme['card_bg'], relief=tk.FLAT, highlightbackground=self.layout.get('border_highlight', '#525252'), highlightthickness=1, padx=CARD_PADDING_X, pady=CARD_PADDING_Y)
        results_frame.pack(fill=tk.BOTH, expand=True, pady=(SECTION_PADDING // 2, 0))
        style = ttk.Style()
        style.configure("Treeview", font=self.font(10), rowheight=18)
        style.configure("Treeview.Heading", font=self.font(11, bold=True))
        columns = ("Score", "Ports", "Ping", "IP", "Operator", "Rank")
        self.tree = ttk.Treeview(results_frame, columns=columns, show='headings', height=8)
        headers = [(self.t("score_hdr"), 70), (self.t("ports_hdr"), 400), (self.t("ping_hdr"), 70), (self.t("ip_hdr"), 120), (self.t("operator_hdr"), 220), (self.t("rank_hdr"), 50)]
        for (text, width), col in zip(headers, columns):
            self.tree.heading(col, text=text)
            self.tree.column(col, width=width, anchor='center')
        scrollbar = ttk.Scrollbar(results_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        
        self.tree.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=4, pady=4)
        scrollbar.pack(side=tk.LEFT, fill=tk.Y)
        
        self.tree.bind("<ButtonRelease-1>", self.on_tree_click)
        self.tree.bind("<Button-3>", self.copy_ip)
        return res_wrap
    
    def create_footer(self, parent):
        footer = tk.Frame(parent, bg=self.theme['bg'])
        footer_center = tk.Frame(footer, bg=self.theme['bg'])
        footer_center.pack(anchor='center')
        tk.Label(footer_center, text=self.t("made_by_simple"), font=self.font(10), bg=self.theme['bg'], fg=self.theme['fg']).pack()
        link_row = tk.Frame(footer_center, bg=self.theme['bg'])
        link_row.pack(pady=4)
        gh_char, gh_f, gh_color = self.fa_icon("github", 12, use_brands=True)
        if gh_f:
            github_btn = tk.Label(link_row, text=gh_char, font=gh_f, bg=self.theme['bg'], fg=gh_color, cursor="hand2")
        else:
            github_btn = tk.Label(link_row, text=self.t("github_text"), font=self.font(10), bg=self.theme['bg'], fg=self.theme['fg'], cursor="hand2")
        github_btn.pack(side=tk.LEFT, padx=12)
        github_btn.bind("<Button-1>", lambda e: webbrowser.open(GITHUB_REPO_URL))
        yt_char, yt_f, yt_color = self.fa_icon("youtube", 12, use_brands=True)
        if yt_f:
            yt_btn = tk.Label(link_row, text=yt_char, font=yt_f, bg=self.theme['bg'], fg=yt_color, cursor="hand2")
        else:
            yt_btn = tk.Label(link_row, text=self.t("youtube_text"), font=self.font(10), bg=self.theme['bg'], fg=self.theme['fg'], cursor="hand2")
        yt_btn.pack(side=tk.LEFT, padx=12)
        yt_btn.bind("<Button-1>", lambda e: webbrowser.open(YOUTUBE_URL))
        web_btn = tk.Label(link_row, text=self.t("website_link_text"), font=self.font(10), bg=self.theme['bg'], fg=self.theme['fg'], cursor="hand2")
        web_btn.pack(side=tk.LEFT, padx=12)
        web_btn.bind("<Button-1>", lambda e: webbrowser.open(WEBSITE_URL))
        return footer
    
    def on_tree_click(self, event):
        """کلیک روی آیتم برای کپی IP"""
        region = self.tree.identify("region", event.x, event.y)
        if region == "cell":
            item = self.tree.identify_row(event.y)
            if item:
                ip = self.tree.item(item)['values'][3]  # IP column
                self.root.clipboard_clear()
                self.root.clipboard_append(ip)
                self.root.update()
                # Show tooltip
                self.show_tooltip(self.t("ip_copied").format(ip))
    
    def show_tooltip(self, text):
        """نمایش tooltip"""
        tooltip = tk.Toplevel(self.root)
        tooltip.wm_overrideredirect(True)
        tooltip.wm_geometry(f"+{self.root.winfo_pointerx()+10}+{self.root.winfo_pointery()+10}")
        
        label = tk.Label(
            tooltip,
            text=text,
            font=self.font(10),
            bg=self.theme['success'],
            fg='#000',
            padx=10,
            pady=5
        )
        label.pack()
        
        self.root.after(1500, tooltip.destroy)
    
    def copy_ip(self, event):
        """کپی IP با راست کلیک"""
        item = self.tree.selection()
        if item:
            ip = self.tree.item(item[0])['values'][3]
            self.root.clipboard_clear()
            self.root.clipboard_append(ip)
            self.root.update()
            self.show_tooltip(self.t("ip_copied").format(ip))
    
    def show_donate(self):
        donate_window = tk.Toplevel(self.root)
        donate_window.title("💎 " + self.t("donate_title"))
        donate_window.geometry("500x300")
        donate_window.configure(bg=self.theme['bg'])
        donate_window.resizable(False, False)
        
        tk.Label(donate_window, text=self.t("donate_title"), font=self.font(18, bold=True), bg=self.theme['bg'], fg=self.theme['fg']).pack(pady=20)
        tk.Label(donate_window, text=self.t("donate_desc"), font=self.font(11), bg=self.theme['bg'], fg=self.theme['fg']).pack(pady=10)
        
        address_frame = RoundedFrame(donate_window, radius=RADIUS_BOX, bg=self.theme['card_bg'])
        address_frame.pack(pady=20, padx=30, fill=tk.X)
        
        tk.Label(
            address_frame.inner,
            text="USDT (Ethereum Network)",
            font=self.font(10, bold=True),
            bg=self.theme['card_bg'],
            fg=self.theme['warning']
        ).pack(pady=(10, 5))
        
        address = "0xde200d902bcf7d2a4f4a17b337b93caa7f78c269"
        tk.Label(
            address_frame.inner,
            text=address,
            font=self.font(10),
            bg=self.theme['card_bg'],
            fg=self.theme['fg']
        ).pack(pady=(0, 10))
        
        def copy_address():
            donate_window.clipboard_clear()
            donate_window.clipboard_append(address)
            messagebox.showinfo(self.t("msg_copied"), self.t("msg_copied"))
        
        bs = self._button_style()
        self._make_button(donate_window, self.t("copy_btn"), copy_address, font=self.font(11, bold=True), width=bs['width'], padx=bs['padx'], pady=bs['pady']).pack(pady=10)
        
        tk.Label(donate_window, text=self.t("thanks"), font=self.font(10), bg=self.theme['bg'], fg=self.theme['fg']).pack()
    
    def show_range_fetcher(self):
        fetcher_window = tk.Toplevel(self.root)
        fetcher_window.title("📡 " + self.t("range_fetcher_title"))
        fetcher_window.geometry("700x600")
        fetcher_window.configure(bg=self.theme['bg'])
        
        tk.Label(
            fetcher_window,
            text=self.t("range_sources_title"),
            font=self.font(16, bold=True),
            bg=self.theme['bg'],
            fg=self.theme['fg']
        ).pack(pady=20)
        
        sources_frame = tk.Frame(fetcher_window, bg=self.theme['bg'])
        sources_frame.pack(pady=10)
        
        sources = [
            (self.t("cf_api"), lambda: self.fetch_ranges("cf_api", fetcher_window)),
            (self.t("cf_asn"), lambda: self.fetch_ranges("cf_asn", fetcher_window)),
            (self.t("cf_github"), lambda: self.fetch_ranges("cf_github", fetcher_window)),
            (self.t("fastly_api"), lambda: self.fetch_ranges("fastly_api", fetcher_window)),
            (self.t("fastly_asn"), lambda: self.fetch_ranges("fastly_asn", fetcher_window)),
            (self.t("all_sources"), lambda: self.fetch_ranges("all", fetcher_window)),
        ]
        
        bs = self._button_style()
        for idx, (text, command) in enumerate(sources):
            self._make_button(sources_frame, text, command, font=self.font(11), width=bs['width'], padx=bs['padx'], pady=bs['pady'], bg=self.theme['card_bg'], fg=self.theme['fg'], border=bs['border']).grid(row=idx//2, column=idx%2, padx=10, pady=5, sticky='ew')
        
        tk.Label(fetcher_window, text=self.t("ranges_received"), font=self.font(12, bold=True), bg=self.theme['bg'], fg=self.theme['fg']).pack(pady=(20, 5))
        
        self.ranges_text, st_frame_fetcher = self._scrolled_text_frame(fetcher_window, height=15, font=self.font(10), bg=self.theme['card_bg'], fg=self.theme['fg'])
        st_frame_fetcher.pack(padx=20, pady=10, fill=tk.BOTH, expand=True)
        
        self.ranges_stats_label = tk.Label(
            fetcher_window,
            text=self.t("ready_fetch"),
            font=self.font(10),
            bg=self.theme['bg'],
            fg=self.theme['fg']
        )
        self.ranges_stats_label.pack(pady=5)
    
    def fetch_ranges(self, source, window):
        self.ranges_text.delete(1.0, tk.END)
        self.ranges_stats_label.config(text=self.t("fetching"))
        
        def fetch_thread():
            ranges = []
            
            source_key = {"cf_api": "source_cf_api", "cf_asn": "source_cf_asn", "cf_github": "source_cf_github",
                         "fastly_api": "source_fastly_api", "fastly_asn": "source_fastly_asn", "all": "source_all"}.get(source, "source_all")
            if source == "cf_api":
                ranges = self.fetcher.get_cloudflare_official()
            elif source == "cf_asn":
                ranges = self.fetcher.get_cloudflare_asn()
            elif source == "cf_github":
                ranges = self.fetcher.get_cloudflare_github()
            elif source == "fastly_api":
                ranges = self.fetcher.get_fastly_official()
            elif source == "fastly_asn":
                ranges = self.fetcher.get_fastly_asn()
            else:
                ranges = self.fetcher.get_all_sources()
            
            self.root.after(0, lambda: self.display_ranges(ranges, source_key))
        
        threading.Thread(target=fetch_thread, daemon=True).start()
    
    def display_ranges(self, ranges, source_key):
        self.ranges_text.delete(1.0, tk.END)
        
        if not ranges:
            self.ranges_text.insert(tk.END, "❌ " + self.t("error_fetch"))
            self.ranges_stats_label.config(text=self.t("error_title"), fg=self.theme['error'])
            return
        
        for r in ranges:
            self.ranges_text.insert(tk.END, f"{r}\n")
        
        self.ranges_stats_label.config(
            text=self.t("fetch_status").format(self.num(len(ranges)), self.t(source_key)),
            fg=self.theme['fg']
        )
        source_display = {
            "source_cf_api": self.t("cf_api"), "source_cf_asn": self.t("cf_asn"), "source_cf_github": self.t("cf_github"),
            "source_fastly_api": self.t("fastly_api"), "source_fastly_asn": self.t("fastly_asn"), "source_all": self.t("all_sources")
        }
        self.current_range_source = source_display.get(source_key, self.t(source_key))
        if hasattr(self, "range_source_value_label") and self.range_source_value_label.winfo_exists():
            self.range_source_value_label.config(text=self.current_range_source)
        self.root.after(0, lambda: self.refresh_ranges_selection_ui(ranges))
    
    def toggle_theme(self):
        self.current_theme = 'light' if self.current_theme == 'dark' else 'dark'
        self.theme = ThemeManager.get_theme(self.current_theme)
        
        for widget in self.root.winfo_children():
            widget.destroy()
        
        self.setup_ui()
    
    def export_menu(self):
        if not self.results:
            messagebox.showwarning(self.t("error_title"), self.t("msg_no_results"))
            return
        
        menu_window = tk.Toplevel(self.root)
        menu_window.title(self.t("export_menu_title"))
        menu_window.geometry("400x320")
        menu_window.configure(bg=self.theme['bg'])
        menu_window.resizable(False, False)
        
        tk.Label(
            menu_window,
            text=self.t("export_menu_title"),
            font=self.font(16, bold=True),
            bg=self.theme['bg'],
            fg=self.theme['fg']
        ).pack(pady=20)
        
        def export_json():
            menu_window.destroy()
            self.export_json()
        
        def export_excel():
            menu_window.destroy()
            self.export_excel()
        
        def export_txt():
            menu_window.destroy()
            self.export_txt()
        
        bs = self._button_style()
        self._make_button(menu_window, self.t("excel_btn"), export_excel, font=self.font(13, bold=True), width=bs['width'], padx=bs['padx'], pady=bs['pady']).pack(pady=10)
        self._make_button(menu_window, self.t("txt_btn"), export_txt, font=self.font(13, bold=True), width=bs['width'], padx=bs['padx'], pady=bs['pady']).pack(pady=10)
        self._make_button(menu_window, self.t("json_btn"), export_json, font=self.font(13, bold=True), width=bs['width'], padx=bs['padx'], pady=bs['pady']).pack(pady=10)
    
    def export_excel(self):
        """اکسپورت به Excel"""
        if not EXCEL_AVAILABLE:
            messagebox.showerror(self.t("error_title"), self.t("openpyxl_missing"))
            return
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"scan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
        if filename:
            try:
                ExcelExporter.export(self.results, filename, app=self)
                messagebox.showinfo(self.t("success_title"), self.t("file_saved").format(filename))
            except Exception as e:
                messagebox.showerror(self.t("error_title"), self.t("save_error").format(str(e)))
    
    def export_txt(self):
        """Export only IPs that were tested and got valid output (open ports + ping), one IP per line."""
        valid_results = [r for r in self.results if r.get('open_ports') and r.get('ping') is not None]
        if not valid_results:
            messagebox.showwarning(self.t("error_title"), self.t("msg_no_results"))
            return
        filename = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
            initialfile=f"ips_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        )
        if filename:
            try:
                with open(filename, "w", encoding="utf-8") as f:
                    for r in valid_results:
                        f.write(r["ip"] + "\n")
                messagebox.showinfo(self.t("success_title"), self.t("file_saved").format(filename))
            except Exception as e:
                messagebox.showerror(self.t("error_title"), self.t("save_error").format(str(e)))
    
    def export_json(self):
        """اکسپورت به JSON"""
        filename = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            initialfile=f"scan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        )
        
        if filename:
            sorted_results = sorted(self.results, key=lambda x: x['score'], reverse=True)
            
            output = {
                'scan_info': {
                    'date': datetime.now().isoformat(),
                    'mode': self.mode_var.get(),
                    'ai_level': self.ai_var.get(),
                    'duration': time.time() - self.start_time if self.start_time else 0,
                    'author': AUTHOR_NAME,
                    'github': GITHUB_URL,
                    'website': WEBSITE_URL,
                    'youtube': YOUTUBE_URL,
                    'donate': '0xde200d902bcf7d2a4f4a17b337b93caa7f78c269'
                },
                'statistics': {
                    'total_scanned': self.total_scanned,
                    'total_found': self.total_found,
                    'success_rate': f"{(self.total_found / self.total_scanned * 100):.2f}%" if self.total_scanned > 0 else "0%"
                },
                'results': sorted_results
            }
            
            try:
                with open(filename, 'w', encoding='utf-8') as f:
                    json.dump(output, f, indent=2, ensure_ascii=False)
                messagebox.showinfo(self.t("success_title"), self.t("file_saved").format(filename))
            except Exception as e:
                messagebox.showerror(self.t("error_title"), self.t("save_error").format(str(e)))
    
    def start_scan(self):
        if self.is_scanning:
            return
        
        # حتماً باید حداقل یک رنج انتخاب شده باشد
        selected = self.get_selected_ranges()
        if not selected:
            messagebox.showwarning(self.t("error_title"), self.t("select_range_first"))
            return
        
        # آی‌پی‌های بلاک‌شدهٔ قبلی را پاک نمی‌کنیم؛ از کش بارگذاری‌شده استفاده می‌کنیم و در صورت وجود سؤال می‌پرسیم
        rescan_closed = True
        if self.scanned_closed:
            n = len(self.scanned_closed)
            msg = self.t("rescan_blocked_message").format(self.num(n))
            win = tk.Toplevel(self.root)
            win.title(self.t("rescan_closed_title"))
            win.configure(bg=self.theme['bg'])
            win.resizable(False, False)
            win.transient(self.root)
            win.grab_set()
            tk.Label(win, text=msg, font=self.font(11), bg=self.theme['bg'], fg=self.theme['fg'], wraplength=400).pack(padx=24, pady=16)
            btn_f = tk.Frame(win, bg=self.theme['bg'])
            btn_f.pack(pady=(0, 16))
            dialog_result = [True]
            def on_yes():
                dialog_result[0] = True
                win.destroy()
            def on_no():
                dialog_result[0] = False
                win.destroy()
            bs = self._button_style()
            self._make_button(btn_f, self.t("rescan_yes"), on_yes, font=self.font(10), width=bs['width'], padx=bs['padx'], pady=bs['pady']).pack(side=tk.LEFT, padx=8)
            self._make_button(btn_f, self.t("rescan_no"), on_no, font=self.font(10), width=bs['width'], padx=bs['padx'], pady=bs['pady']).pack(side=tk.LEFT, padx=8)
            win.wait_window()
            rescan_closed = dialog_result[0]
        
        self.is_scanning = True
        self.results = []
        self.total_scanned = 0
        self.total_found = 0
        self.start_time = time.time()
        
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        self.scanner.update_queue = self.update_queue
        self._set_progress_percent(0)
        self.root.update_idletasks()
        
        if isinstance(self.start_btn, tk.Button):
            self.start_btn.config(state=tk.DISABLED)
        if isinstance(self.stop_btn, tk.Button):
            self.stop_btn.config(state=tk.NORMAL)
        if isinstance(self.export_btn, tk.Button):
            self.export_btn.config(state=tk.DISABLED)
        if isinstance(self.analyze_btn, tk.Button):
            self.analyze_btn.config(state=tk.DISABLED)
        
        threading.Thread(target=self.scan_worker, args=(rescan_closed,), daemon=True).start()
    
    def scan_worker(self, rescan_closed=True):
        try:
            self.update_queue.put(("status", self.t("status_fetch")))
            user_operator = None
            effective_operator_key = None
            operator_matrix = {}
            self._last_scan_operator = ""
            # فقط در حالت «با رنج آی پی اپراتورها» ماتریس و برچسب اپراتور پر می‌شود
            if self._is_operator_scan_mode():
                ops_ctx = self.get_operators_for_current_context()
                effective_operator_key = getattr(self, 'current_operator_key', None) or self._get_user_operator_key()
                if effective_operator_key is None or effective_operator_key not in ops_ctx:
                    effective_operator_key = list(ops_ctx.keys())[0]
                user_operator = ops_ctx[effective_operator_key]["name"]
                operator_matrix = self._load_operator_matrix()
                self._last_scan_operator = user_operator
            
            all_ranges = self.get_selected_ranges()
            
            if not all_ranges:
                self.update_queue.put(("error", self.t("select_range_first")))
                self.update_queue.put(("done", None))
                return
            
            self.update_queue.put(("status", self.t("status_analyze").format(len(all_ranges))))
            
            target = self.get_target_internal()
            if target == "All":
                best_ranges = all_ranges
            else:
                # برای رسیدن به «تعداد» نتیجه، رنج کافی برمی‌داریم (حداکثر ۵۰۰ رنج)
                best_ranges = self.ai.predict_best_ranges(all_ranges, min(500, max(100, int(target) * 3)))
            
            mode = self.get_mode_internal()
            if "5s" in mode:
                max_ips = 30
                self.scanner.max_workers = 700
            elif "10s" in mode:
                max_ips = 50
                self.scanner.max_workers = 600
            elif "15s" in mode:
                max_ips = 70
                self.scanner.max_workers = 500
            else:
                max_ips = 100
                self.scanner.max_workers = 450
            
            ports = getattr(self, "scan_ports", None) or self.ai.priority_ports
            self.ai.priority_ports = ports
            target_count = None if target == "All" else int(target)
            ping_min = getattr(self, "ping_min", 0)
            ping_max = getattr(self, "ping_max", 9999)
            self.scanner.max_latency_ms = ping_max
            self.scanner.timeout = min(10, max(2, ping_max / 1000.0 * 1.5))
            
            ips_to_scan = []
            for cidr in best_ranges:
                if not self.is_scanning:
                    break
                if target_count is not None and len(ips_to_scan) >= target_count:
                    break
                ips = self._range_to_ips(cidr, max_ips)
                # فقط وقتی «ردشون کن» زده شده، آی‌پی‌های بستهٔ قبلی را حذف کن؛ آی‌پی دستی (تک‌آی‌پی) همیشه اسکن شود
                is_explicit_single_ip = "/" not in (cidr or "")
                if not rescan_closed and self.scanned_closed and not is_explicit_single_ip:
                    ips = [ip for ip in ips if str(ip) not in self.scanned_closed]
                for ip in ips:
                    ips_to_scan.append(ip)
                    if target_count is not None and len(ips_to_scan) >= target_count:
                        break
            
            if target_count is not None:
                ips_to_scan = ips_to_scan[:target_count]
            total_to_scan = len(ips_to_scan)
            self.update_queue.put(("status", self.t("status_scan").format(total_to_scan)))
            if total_to_scan == 0:
                self.update_queue.put(("progress", (1, 1, 0, 0)))
                self._save_scan_cache()
                self.update_queue.put(("done", None))
                return
            
            self.update_queue.put(("total", total_to_scan))
            batch_size = min(100, max(30, total_to_scan // 4))
            done_so_far = 0
            
            for start in range(0, total_to_scan, batch_size):
                if not self.is_scanning:
                    break
                batch_ips = ips_to_scan[start:start + batch_size]
                batch_results = self.scanner.batch_scan(batch_ips, ports)
                open_ips = {r['ip'] for r in batch_results}
                results_by_ip = {r['ip']: r for r in batch_results}
                for ip in batch_ips:
                    ip_str = str(ip)
                    if ip_str not in open_ips:
                        self.scanned_closed.add(ip_str)
                    res = results_by_ip.get(ip_str)
                    open_ports_list = list(res['open_ports']) if res else []
                    ping_val = res.get('ping') if res else None
                    cell_data = {"active": ip_str in open_ips, "ports": open_ports_list, "ping": ping_val}
                    operator_matrix[ip_str] = operator_matrix.get(ip_str, {})
                    if effective_operator_key is not None:
                        operator_matrix[ip_str][effective_operator_key] = cell_data
                if operator_matrix:
                    self._save_operator_matrix(operator_matrix)
                
                for result in batch_results:
                    ping_val = result.get('ping')
                    if ping_val is not None and (ping_val < ping_min or ping_val > ping_max):
                        continue
                    if user_operator is not None:
                        result['operator'] = user_operator
                    self.total_found += 1
                    score = self._calc_score(result)
                    result['score'] = score
                    self.results.append(result)
                    self.update_queue.put(("result", result))
                
                self.total_scanned += len(batch_ips)
                done_so_far = min(start + len(batch_ips), total_to_scan)
                elapsed = time.time() - self.start_time
                speed = self.total_scanned / elapsed if elapsed > 0 else 0
                self.update_queue.put(("progress", (done_so_far, total_to_scan, speed, elapsed)))
            
            elapsed = time.time() - self.start_time
            speed = self.total_scanned / elapsed if elapsed > 0 else 0
            self.update_queue.put(("progress", (total_to_scan, total_to_scan, speed, elapsed)))
            self._save_scan_cache()
            self.update_queue.put(("done", None))
            
        except Exception as e:
            self.update_queue.put(("error", str(e)))
            self.update_queue.put(("done", None))
    
    def _calc_score(self, result):
        """امتیازدهی هوشمند بر اساس پینگ و وضعیت پورت‌ها (باز بودن)."""
        score = 0.0
        ping_val = result.get('ping')
        if ping_val is not None:
            if ping_val < 50:
                score += 35
            elif ping_val < 100:
                score += 28
            elif ping_val < 200:
                score += 20
            elif ping_val < 500:
                score += 10
            elif ping_val < 1000:
                score += 3
        open_ports = result.get('open_ports') or []
        score += len(open_ports) * 3
        if 443 in open_ports:
            score += 12
        if 80 in open_ports:
            score += 10
        if 8080 in open_ports:
            score += 4
        if 8443 in open_ports:
            score += 4
        return max(0.0, min(100.0, score))
    
    def stop_scan(self):
        self.is_scanning = False
        if isinstance(self.stop_btn, tk.Button):
            self.stop_btn.config(state=tk.DISABLED)
    
    def ai_analyze(self):
        if not self.results:
            messagebox.showinfo(self.t("analyze_title"), self.t("no_results_analyze"))
            return
        
        sorted_results = sorted(self.results, key=lambda x: x['score'], reverse=True)
        avg_ping = sum(r['ping'] for r in self.results if r['ping']) / len(self.results)
        best = sorted_results[0]
        
        n_results = len(self.results)
        n_low_ping = len([r for r in self.results if r['ping'] and r['ping'] < 50])
        n_443 = len([r for r in self.results if 443 in r['open_ports']])
        n_3ports = len([r for r in self.results if len(r['open_ports']) >= 3])
        
        # گزارش تحلیل: خطوط جدا و مرتب (آخرین خط‌ها جدا تا در RTL قاطی نشوند)
        lines = [
            self.t("analyze_title"),
            "=" * 40,
            "",
            self.t("found") + ": " + self.num(n_results),
            "Ping: " + (self.t("ping_unsuitable") if avg_ping >= 1000 else self.num(int(round(avg_ping))) + " ms"),
            self.t("score_hdr") + ": " + self.num(int(best['score'])) + " / " + self.num(100),
            "",
            "Best IP: " + best['ip'],
            "Ping: " + (self.t("ping_unsuitable") if best['ping'] >= 1000 else self.num(int(round(best['ping']))) + " ms"),
            "Ports: " + ", ".join(map(str, best['open_ports'])),
            "",
            self.num(n_low_ping) + " IP < 50ms",
            self.num(n_443) + " with 443",
            self.num(n_3ports) + " with 3+ ports",
        ]
        analysis_text = "\n".join(lines)
        
        # پنجرهٔ سفارشی برای گزارش راست‌چین و مرتب
        win = tk.Toplevel(self.root)
        win.title(self.t("analyze_title"))
        win.configure(bg=self.theme['bg'])
        win.resizable(True, True)
        win.transient(self.root)
        
        main_f = tk.Frame(win, bg=self.theme['bg'], padx=24, pady=20)
        main_f.pack(fill=tk.BOTH, expand=True)
        
        txt, st_frame_ai = self._scrolled_text_frame(
            main_f,
            font=self.font(11),
            bg=self.theme['card_bg'],
            fg=self.theme['fg'],
            wrap=tk.WORD,
            width=48,
            height=18,
        )
        txt.configure(relief=tk.FLAT, padx=15, pady=15)
        st_frame_ai.pack(fill=tk.BOTH, expand=True)
        txt.insert(tk.END, analysis_text)
        txt.tag_configure("align", justify=tk.RIGHT if self.is_rtl else tk.LEFT)
        txt.tag_add("align", "1.0", tk.END)
        # کلیک روی IP → کپی خودکار در کلیپ‌بورد
        def on_ai_text_click(event):
            idx = txt.index(f"@{event.x},{event.y}")
            parts = idx.split(".")
            line_num, col = parts[0], int(parts[1])
            line_content = txt.get(f"{line_num}.0", f"{line_num}.end")
            for m in re.finditer(r"\b(?:\d{1,3}\.){3}\d{1,3}\b", line_content):
                start, end = m.span()
                if start <= col < end:
                    ip = m.group(0)
                    win.clipboard_clear()
                    win.clipboard_append(ip)
                    win.update()
                    self.show_tooltip(self.t("ip_copied").format(ip))
                    return
        txt.bind("<Button-1>", on_ai_text_click)
        txt.bind("<Key>", lambda e: "break")
        txt.config(cursor="hand2")
        
        bs = self._button_style()
        self._make_button(main_f, self.t("close_btn"), win.destroy, font=self.font(10), width=bs['width'], padx=bs['padx'], pady=bs['pady']).pack(pady=(CARD_PADDING_Y + 4, 0), anchor='e' if self.is_rtl else 'w')
    
    def check_queue(self):
        try:
            while True:
                msg_type, data = self.update_queue.get_nowait()
                
                if msg_type == "status":
                    self.status_label.config(text=data)
                elif msg_type == "total":
                    pass
                    # نوار پیشرفت بر اساس درصد ۰–۱۰۰ است؛ هدف را عوض نکن
                elif msg_type == "progress":
                    done, total, speed, elapsed = data
                    percent = (float(done) / float(total) * 100.0) if total and total > 0 else 0.0
                    percent = min(100.0, percent)
                    self._set_progress_percent(percent)
                    self.speed_label.config(text=self.num(int(speed)) + " IP/s")
                    self.found_label.config(text=self.num(self.total_found))
                    self.time_label.config(text=self.num(f"{elapsed:.1f}") + "s")
                    self.status_label.config(text=self.t("status_pct").format(self.num(f"{percent:.1f}")))
                    self.root.update_idletasks()
                elif msg_type == "ip_status":
                    ip_str, success, latency, open_ports = data
                    try:
                        only_clean = getattr(self, "filter_only_clean_var", None) and self.filter_only_clean_var.get()
                        if only_clean and not (success and latency is not None and open_ports):
                            pass
                        elif success and latency is not None and open_ports is not None:
                            priority_ports = getattr(self, "scan_ports", None) or self.ai.priority_ports
                            mini = {'ip': ip_str, 'open_ports': open_ports, 'ping': latency}
                            mini['operator'] = getattr(self, '_last_scan_operator', "") if self._has_operator_display() else ""
                            score_str = self.num(int(self._calc_score(mini))) + "/" + self.num(100)
                            if self._has_operator_display():
                                ports_str = self._format_ports_column(mini)
                                op_str = self._format_operator_column(mini)
                            else:
                                ports_str = " ".join([f"{p}✅" for p in open_ports])
                                op_str = ""
                            ping_str = self.num(int(round(latency))) + " ms" if latency < 1000 else self.t("ping_unsuitable")
                            self.tree.insert("", tk.END, values=(score_str, ports_str, ping_str, ip_str, op_str, "—"))
                        else:
                            if not only_clean:
                                if self._has_operator_display():
                                    mini_fail = {'ip': ip_str, 'open_ports': [], 'ping': None, 'operator': getattr(self, '_last_scan_operator', '')}
                                    op_str = self._format_operator_column(mini_fail)
                                    ports_str = self._format_ports_column(mini_fail)
                                else:
                                    op_str = ""
                                    ports_str = "❌"
                                self.tree.insert("", tk.END, values=("—", ports_str, "—", ip_str, op_str, "—"))
                        for idx, item in enumerate(self.tree.get_children(), 1):
                            vals = list(self.tree.item(item)['values'])
                            vals[5] = "#" + self.num(idx)
                            self.tree.item(item, values=vals)
                        self.root.update_idletasks()
                    except Exception:
                        pass
                elif msg_type == "result":
                    result = data
                    ip_val = result['ip']
                    only_clean = getattr(self, "filter_only_clean_var", None) and self.filter_only_clean_var.get()
                    if self._has_operator_display() and getattr(self, 'current_operator_key', None) is None:
                        self._rebuild_results_from_matrix()
                    elif only_clean and not (result.get('open_ports') and result.get('ping')):
                        pass
                    else:
                        found_item = None
                        for item in self.tree.get_children():
                            if self.tree.item(item)['values'][3] == ip_val:
                                found_item = item
                                break
                        if self._has_operator_display():
                            ports_str = self._format_ports_column(result)
                            op_str = self._format_operator_column(result)
                        else:
                            ports_str = " ".join([f"{p}✅" for p in result.get('open_ports', [])])
                            op_str = ""
                        score_str = self.num(int(result['score'])) + "/" + self.num(100)
                        if not result.get('ping'):
                            ping_str = "N/A"
                        elif result['ping'] >= 1000:
                            ping_str = self.t("ping_unsuitable")
                        else:
                            ping_str = self.num(int(round(result['ping']))) + " ms"
                        if found_item:
                            self.tree.item(found_item, values=(score_str, ports_str, ping_str, ip_val, op_str, "—"))
                        else:
                            self.tree.insert("", tk.END, values=(score_str, ports_str, ping_str, ip_val, op_str, "—"))
                        target = self.get_target_internal()
                        items = self.tree.get_children()
                        if target != "All" and len(items) > int(target):
                            self.tree.delete(items[-1])
                        for idx, item in enumerate(self.tree.get_children(), 1):
                            vals = list(self.tree.item(item)['values'])
                            vals[5] = "#" + self.num(idx)
                            self.tree.item(item, values=vals)
                elif msg_type == "error":
                    messagebox.showerror(self.t("error_title"), data)
                elif msg_type == "done":
                    self.is_scanning = False
                    self._set_progress_percent(100)
                    if self._has_operator_display() and getattr(self, 'current_operator_key', None) is None:
                        self._rebuild_results_from_matrix()
                    try:
                        self.root.update_idletasks()
                        self.root.update()
                    except Exception:
                        pass
                    self._save_last_scan()
                    self.root.after(50, lambda: self._set_progress_percent(100))
                    if isinstance(self.start_btn, tk.Button):
                        self.start_btn.config(state=tk.NORMAL)
                    if isinstance(self.stop_btn, tk.Button):
                        self.stop_btn.config(state=tk.DISABLED)
                    if isinstance(self.export_btn, tk.Button):
                        self.export_btn.config(state=tk.NORMAL)
                    if isinstance(self.analyze_btn, tk.Button):
                        self.analyze_btn.config(state=tk.NORMAL)
                    elapsed = time.time() - self.start_time
                    self.status_label.config(text=self.t("scan_done").format(self.num(self.total_found), self.num(f"{elapsed:.1f}")))
                    if self.total_found > 0:
                        self.root.after(1000, self.ai_analyze)
        except Exception:
            pass

        self.root.after(50, self.check_queue)


def _show_splash_then_run():
    """نمایش پنجرهٔ بارگذاری با نوار پیشرفت، سپس اجرای برنامهٔ اصلی."""
    splash = tk.Tk()
    splash.title("")
    splash.overrideredirect(True)
    splash.configure(bg="#0a0a0a")
    w, h = 420, 160
    sw = splash.winfo_screenwidth()
    sh = splash.winfo_screenheight()
    splash.geometry(f"{w}x{h}+{(sw - w) // 2}+{(sh - h) // 2}")
    tk.Label(splash, text="CDN IP Scanner", font=("Segoe UI", 16, "bold"), bg="#0a0a0a", fg="#ffffff").pack(pady=(24, 8))
    tk.Label(splash, text="Loading... Preparing environment", font=("Segoe UI", 9), bg="#0a0a0a", fg="#a0a0a0").pack(pady=(0, 12))
    progress_var = tk.DoubleVar(value=0.0)
    progress = ttk.Progressbar(splash, variable=progress_var, maximum=100, length=360, mode="determinate")
    progress.pack(pady=(0, 20))
    splash.update()

    def step(percent, _splash=splash, _pv=progress_var):
        _pv.set(percent)
        _splash.update()

    step(5)
    _load_theme_from_assets()
    step(25)
    time.sleep(0.08)
    step(50)
    if os.path.isdir(os.path.join(RESOURCE_DIR, "fonts")):
        _load_fonts_win(os.path.join(RESOURCE_DIR, "fonts"))
    step(75)
    time.sleep(0.06)
    step(100)
    time.sleep(0.05)
    splash.destroy()
    splash.update()

    root = tk.Tk()
    app = CDNScannerPro(root)
    root.mainloop()


if __name__ == "__main__":
    _show_splash_then_run()
